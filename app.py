import streamlit as st
import json, os, uuid
import pandas as pd
from datetime import date, datetime, timedelta
from io import BytesIO

st.set_page_config(page_title="BP/BM 재고·손익 관리", layout="wide", initial_sidebar_state="expanded")
st.markdown("""
<style>
    section[data-testid="stSidebar"]{min-width:270px}
    .mbox{background:#f0f2f6;border-radius:8px;padding:12px 16px;margin:4px 0}
    .ph{font-size:1.35rem;font-weight:700;color:#1f4e79}
    .sp{font-size:0.88rem;color:#555;margin-bottom:2px}
    .b-bp{background:#d6e4f0;color:#1a5276;padding:2px 8px;border-radius:4px;font-size:.78rem;font-weight:600}
    .b-bm{background:#d5f5e3;color:#1e8449;padding:2px 8px;border-radius:4px;font-size:.78rem;font-weight:600}
    .b-sc{background:#fdebd0;color:#784212;padding:2px 8px;border-radius:4px;font-size:.78rem;font-weight:600}
    .b-ok{background:#d4efdf;color:#1d6a39;padding:2px 8px;border-radius:4px;font-size:.78rem;font-weight:600}
    .b-wn{background:#fef9e7;color:#7d6608;padding:2px 8px;border-radius:4px;font-size:.78rem;font-weight:600}
    .b-ng{background:#fadbd8;color:#922b21;padding:2px 8px;border-radius:4px;font-size:.78rem;font-weight:600}
    div[data-testid="stTabs"] button{font-size:.9rem;font-weight:600}
</style>""", unsafe_allow_html=True)

CONFIG_FILE  = os.path.join(os.path.dirname(__file__), "config.json")
_GSHEET_CREDS = os.path.join(os.path.dirname(__file__), "bp-calculator-498206-4308cbd64cba.json")

# ── 로컬 vs 클라우드 감지 ────────────────────────────────────────────────────
# 로컬 인증 파일 있음 → 파일로 인증, Drive를 데이터 저장소로 사용
# 클라우드(Streamlit Cloud) → Streamlit Secrets로 인증, Drive 사용
_HAS_LOCAL_CREDS = os.path.exists(_GSHEET_CREDS)
_IS_CLOUD        = not _HAS_LOCAL_CREDS   # 로컬 인증 파일 없으면 클라우드 모드

# ── 읽기 전용 배포 모드 ──────────────────────────────────────────────────────
# secrets.toml(또는 Streamlit Cloud Secrets)에 `read_only = true` 를 설정한
# 별도 배포본에서만 켜짐. 원본 앱의 secrets에는 이 키가 없으므로 항상 False.
try:
    READ_ONLY = bool(st.secrets.get("read_only", False))
except Exception:
    READ_ONLY = False

if READ_ONLY:
    st.warning(
        "🔒 **읽기 전용 모드** — 모든 저장·삭제·동기화 버튼이 비활성화되어 있습니다. "
        "데이터는 변경되지 않습니다.",
        icon="🔒",
    )
    _ro_orig_button       = st.button
    _ro_orig_form_submit  = st.form_submit_button

    def _ro_button(*args, **kwargs):
        kwargs["disabled"] = True
        _ro_orig_button(*args, **kwargs)
        return False

    def _ro_form_submit(*args, **kwargs):
        kwargs["disabled"] = True
        _ro_orig_form_submit(*args, **kwargs)
        return False

    st.button = _ro_button
    st.form_submit_button = _ro_form_submit

def _get_gcp_creds(scopes):
    """로컬: 인증 JSON 파일 / 클라우드: Streamlit Secrets."""
    from google.oauth2.service_account import Credentials
    import json as _json
    if _HAS_LOCAL_CREDS:
        return Credentials.from_service_account_file(_GSHEET_CREDS, scopes=scopes)
    info = _json.loads(st.secrets["gcp_service_account_json"])
    return Credentials.from_service_account_info(info, scopes=scopes)

def _get_drive_service():
    """Google Drive API 서비스 반환."""
    from googleapiclient.discovery import build
    creds = _get_gcp_creds(["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=creds)


@st.cache_data(ttl=30, show_spinner=False)
def _load_cfg_drive():
    """Google Drive에서 config.json 내용을 읽어 dict 반환 (30초 캐시)."""
    from googleapiclient.http import MediaIoBaseDownload
    import io
    svc     = _get_drive_service()
    file_id = st.secrets["drive_config_file_id"]
    req     = svc.files().get_media(fileId=file_id)
    buf     = io.BytesIO()
    dl      = MediaIoBaseDownload(buf, req)
    done    = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return json.loads(buf.read().decode("utf-8"))

def _save_cfg_drive(c):
    """dict를 JSON으로 직렬화해 Google Drive 파일에 덮어씁니다."""
    from googleapiclient.http import MediaInMemoryUpload
    svc     = _get_drive_service()
    file_id = st.secrets["drive_config_file_id"]
    content = json.dumps(c, ensure_ascii=False, indent=2).encode("utf-8")
    media   = MediaInMemoryUpload(content, mimetype="application/json")
    svc.files().update(fileId=file_id, media_body=media).execute()
    _load_cfg_drive.clear()   # Drive 캐시 무효화
    _fifo_lot_trace.clear()   # FIFO 계산 캐시 무효화

_DOCS_DRIVE_FILE_ID = "1lZd2EVs9T9OJL21AcPA49_to2Bhkdeup"

@st.cache_data(ttl=30, show_spinner=False)
def _load_docs_drive():
    """Google Drive에서 bp_documents.json 내용을 읽어 list 반환 (30초 캐시)."""
    from googleapiclient.http import MediaIoBaseDownload
    import io
    svc = _get_drive_service()
    req = svc.files().get_media(fileId=_DOCS_DRIVE_FILE_ID)
    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    data = json.loads(buf.read().decode("utf-8"))
    return data.get("documents", []) if isinstance(data, dict) else data

def _save_docs_drive(docs_list):
    """문서 목록을 bp_documents.json에 저장."""
    from googleapiclient.http import MediaInMemoryUpload
    svc     = _get_drive_service()
    content = json.dumps({"documents": docs_list}, ensure_ascii=False, indent=2).encode("utf-8")
    media   = MediaInMemoryUpload(content, mimetype="application/json")
    svc.files().update(fileId=_DOCS_DRIVE_FILE_ID, media_body=media).execute()
    _load_docs_drive.clear()

def load_cfg():
    """항상 Google Drive에서 로드 (로컬·클라우드 공통 원본)."""
    return _load_cfg_drive()

def save_cfg(c):
    """항상 Google Drive에 저장 (로컬·클라우드 공통 원본)."""
    if READ_ONLY:
        st.error("🔒 읽기 전용 모드 — 저장이 차단되었습니다.")
        st.stop()
    _save_cfg_drive(c)
    # 로컬 개발 편의용: Drive와 별도로 로컬 백업 유지
    if _HAS_LOCAL_CREDS and os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(c, f, ensure_ascii=False, indent=2)

def bp_price(ni_i,co_i,ni_c,co_c,ni_p,co_p):
    nv=ni_i*(ni_c/100)*ni_p; cv=co_i*(co_c/100)*co_p; t=nv+cv
    return nv,cv,t,t/1000

# ── 처리 이력 헬퍼 (모듈 레벨 — t_proc / t_pnl 공용) ──────────────────────────
from collections import defaultdict

def _ph_input_kg(rec):
    """투입량 반환 — 저장값 우선, 없으면 output_kg ÷ conversion_rate 역산"""
    if rec.get("input_kg") is not None:   # 0.0도 저장값으로 취급 (falsy-zero 방지)
        return rec["input_kg"]
    out  = rec.get("output_kg", 0) or 0
    conv = rec.get("conversion_rate_pct") or rec.get("conversion_rate")
    return out / (conv / 100) if (out and conv and conv > 0) else 0

def _ph_export_usd(rec, cfg=None):
    """수출비 총액(USD).
    우선순위: ① 배치 직접값(구버전 호환) → ② HBL 수출비 생산량 비례 배분 → ③ 구포맷 per_kg
    """
    # ① 배치에 직접 저장된 값 (구버전 호환)
    if rec.get("export_cost_usd") is not None:
        return float(rec["export_cost_usd"])
    # ② HBL 레벨 수출비 비례 배분
    if cfg and rec.get("shipment_id"):
        ship = next((s for s in cfg.get("shipments", []) if s.get("id") == rec["shipment_id"]), {})
        hbl_eu = float(ship.get("export_cost_usd") or 0)
        if hbl_eu > 0:
            total_out = sum(
                float(r.get("output_kg", 0) or 0)
                for r in cfg.get("processing_history", [])
                if r.get("shipment_id") == rec["shipment_id"]
            )
            rec_out = float(rec.get("output_kg", 0) or 0)
            if total_out > 0 and rec_out > 0:
                return round(hbl_eu * rec_out / total_out, 2)
    # ③ 구포맷 fallback
    per_kg = rec.get("export_cost_per_kg_bp", 0) or 0
    out    = rec.get("output_kg", 0) or 0
    return per_kg * out

def _inv_moving_avg(cfg, scrap_id, as_of_date=None):
    """이동평균 단가·누적 입고량 반환.
    as_of_date: 'YYYY-MM-DD' 또는 'YYYY-MM' — 해당 월 말까지의 입고만 반영.
    반환: (avg_cost, cumulative_qty)  둘 다 None/0 이면 기초재고 미설정.
    """
    inv = cfg.get("raw_material_inventory", {}).get(scrap_id, {})
    op  = inv.get("opening")
    if not op or not op.get("quantity_kg"):
        return None, 0.0

    qty = float(op["quantity_kg"])
    avg = float(op["unit_cost"])
    cutoff = str(as_of_date)[:7] if as_of_date and str(as_of_date).strip() else None  # YYYY-MM 비교

    op_date = (op.get("date") or "")[:10]   # 기초재고 기준일 (YYYY-MM-DD)
    for p in sorted(inv.get("purchases", []), key=lambda x: x.get("date", "")):
        p_raw  = (p.get("date") or "")
        p_date = p_raw[:10]
        if op_date:
            # 입고일이 YYYY-MM(일 미입력)인 경우: 월 단위로만 비교 — 기초재고 이전 월만 제외
            # 입고일이 YYYY-MM-DD(일 포함)인 경우: 전체 날짜로 비교 — 기초재고 당일·이전 제외
            if len(p_raw.strip()) <= 7:
                if p_date[:7] < op_date[:7]:
                    continue   # 기초재고 월보다 이전 월 입고 → 이미 기초재고에 포함
            else:
                if p_date <= op_date:
                    continue   # 기초재고 날짜 이전·당일 입고 → 이미 기초재고에 포함
        if cutoff and p_date[:7] > cutoff:
            break
        pq = float(p.get("quantity_kg") or 0)
        pc = float(p.get("unit_cost")   or 0)
        if pq > 0 and pc > 0:
            avg = (qty * avg + pq * pc) / (qty + pq)
            qty += pq

    return round(avg, 5), round(qty, 3)

def _inv_balance(cfg, scrap_id, ph_list=None, as_of_date=None):
    """창고 실물 잔량 = 누적 입고 − 임가공 출고 − 직접 판매
    dispatch_records·direct_sales 가 실물 반출 기준이므로 이를 차감.
    ph_list 인자는 하위호환용으로 유지하되 계산에는 미사용.
    """
    _, total_in = _inv_moving_avg(cfg, scrap_id, as_of_date)
    dispatched  = sum(float(dr.get("quantity_kg") or 0)
                      for dr in cfg.get("dispatch_records", [])
                      if dr.get("scrap_type_id") == scrap_id)
    direct_sold = sum(float(ds.get("quantity_kg") or 0)
                      for ds in cfg.get("direct_sales", [])
                      if ds.get("scrap_type_id") == scrap_id)
    return round(total_in - dispatched - direct_sold, 3)

def _get_eur_usd(cfg, month=None):
    """월별 EUR/USD 환율 조회.
    month: 'YYYY-MM' — 해당 월 이하 가장 최근 환율 반환.
    등록된 환율이 없으면 기본값 1.10 반환.
    """
    rates = sorted(cfg.get("eur_usd_rates", []), key=lambda x: x["month"])
    if not rates:
        return 1.10
    if not month:
        return float(rates[-1]["rate"])
    match = [r for r in rates if r["month"] <= month]
    return float(match[-1]["rate"]) if match else float(rates[0]["rate"])

def _rec_ref_date(rec, cfg):
    """배치 기준일 결정: batch_date 직접 입력 > 연결 HBL 선적일 순서로 반환 (YYYY-MM-DD or None)."""
    if rec.get("batch_date"):
        return rec["batch_date"][:10]
    ship_id = rec.get("shipment_id", "")
    if ship_id:
        ship = next((s for s in cfg.get("shipments", []) if s.get("id") == ship_id), {})
        return (ship.get("loading_date") or "")[:10] or None
    return None

def _storage_rate_eur(cfg, scrap_id):
    """스크랩 유형별 창고 보관비율 (EUR/톤백/day). 미설정 시 1.5."""
    sc = next((s for s in cfg.get("scrap_types", []) if s.get("id") == scrap_id), {})
    return float(sc.get("storage_rate_eur") or 1.5)

def _ph_storage_cost(rec, cfg):
    """scrap 보관비 자동계산 (수동 storage_days 입력 배치용).
    = 톤백수 × 보관일수 × EUR/톤백/day × EUR/USD
    톤백수: ton_bags 직접 입력 우선, 없으면 투입량 ÷ 510
    storage_days 미입력 시 0 반환.
    """
    days = rec.get("storage_days") or 0
    if not days:
        return 0.0
    inp_kg = _ph_input_kg(rec)
    if inp_kg <= 0:
        return 0.0
    # 톤백: 직접 입력 우선, 없으면 중량 역산
    tb = float(rec.get("ton_bags") or 0) or (inp_kg / 510.0)
    ref_date  = _rec_ref_date(rec, cfg)
    month     = ref_date[:7] if ref_date else None
    eur_rate  = _get_eur_usd(cfg, month)
    stor_rate = _storage_rate_eur(cfg, rec.get("scrap_type_id",""))
    return round(tb * float(days) * stor_rate * eur_rate, 2)

@st.cache_data(show_spinner=False)
def _fifo_lot_trace(cfg, scrap_id):
    """2단계 FIFO Lot 추적.

    Level 1  입고 Lot 큐 → 전체 출고 이벤트(임가공 출고 + 직접판매) — 날짜순
             각 임가공 출고 이벤트(dispatch_records)에 lot 귀속 정보 부여
    Level 2  임가공 출고 풀(프로세서별) → B/L 배치(processing_history) — 선적일순
             프로세서별 FIFO로 dispatch pool 소진 → 각 B/L에 lot 귀속 전파

    반환: (bl_result, all_events, lot_remaining)
      bl_result    : {shipment_id → {hbl, load_date, input_kg,
                        lots: {lot_label → {qty, amount, unit_cost, lot_date}}}}
      all_events   : 전체 출고 이벤트 목록 (attributions 포함) — 임가공출고 + 직접판매
      lot_remaining: 미소진 Lot 잔량 리스트
    """
    inv = cfg.get("raw_material_inventory", {}).get(scrap_id, {})
    op  = inv.get("opening")
    if not op or not float(op.get("quantity_kg") or 0):
        return {}, [], []

    # ── Lot 큐 구성 (FIFO = 입고일 오름차순) ─────────────────────────────────
    op_date = (op.get("date") or "")[:10]
    lots = []
    if float(op.get("quantity_kg") or 0) > 0:
        op_qty = float(op["quantity_kg"])
        op_tb  = float(op.get("ton_bags") or 0)
        # kg/백: 기초재고에 ton_bags 입력된 경우 사용, 없으면 510 역산
        op_kpb = op_qty / op_tb if op_tb > 0 else 510.0
        lots.append({
            "label":      f"기초재고 ({op_date})",
            "date":       op_date,
            "unit_cost":  float(op.get("unit_cost") or 0),
            "remain":     op_qty,
            "kg_per_bag": op_kpb,
        })
    for p in sorted(inv.get("purchases", []), key=lambda x: x.get("date", "")):
        p_raw  = (p.get("date") or "")
        p_date = p_raw[:10]
        if op_date:
            if len(p_raw.strip()) <= 7:
                if p_date[:7] < op_date[:7]:
                    continue   # 기초재고 월보다 이전 월 → 기초재고에 이미 포함
            else:
                if p_date <= op_date:
                    continue   # 기초재고 날짜 이전·당일 → 기초재고에 이미 포함
        pq = float(p.get("quantity_kg") or 0)
        if pq > 0:
            p_tb  = float(p.get("ton_bags") or 0)
            p_kpb = pq / p_tb if p_tb > 0 else 510.0
            lots.append({
                "label":      f"{p_date[:7]} 매입 ({p_date})",
                "date":       p_date,
                "unit_cost":  float(p.get("unit_cost") or 0),
                "remain":     pq,
                "kg_per_bag": p_kpb,
            })

    proc_map = {p["id"]: p for p in cfg.get("processors", [])}
    ship_map = {s["id"]: s for s in cfg.get("shipments", [])}

    # ── 출고 이벤트 구성 (임가공 출고 + 직접 판매) ───────────────────────────
    outflow_events = []
    for dr in cfg.get("dispatch_records", []):
        if dr.get("scrap_type_id") != scrap_id:
            continue
        qty = float(dr.get("quantity_kg") or 0)
        if qty <= 0:
            continue
        proc = proc_map.get(dr.get("processor_id", ""), {})
        outflow_events.append({
            "id":           dr.get("id", ""),
            "type":         "임가공출고",
            "date":         (dr.get("date") or "9999-12-31")[:10],
            "qty":          qty,
            "processor_id": dr.get("processor_id", ""),
            "processor":    proc.get("name", "—"),
            "notes":        dr.get("notes", ""),
        })
    for ds in cfg.get("direct_sales", []):
        if ds.get("scrap_type_id") != scrap_id:
            continue
        qty = float(ds.get("quantity_kg") or 0)
        if qty <= 0:
            continue
        outflow_events.append({
            "id":           ds.get("id", ""),
            "type":         "직접판매",
            "date":         (ds.get("date") or "9999-12-31")[:10],
            "qty":          qty,
            "processor_id": "",
            "processor":    "—",
            "notes":        ds.get("notes", ""),
        })
    outflow_events.sort(key=lambda x: (x["date"], x["type"]))

    # ── Level 1: lot 큐 → 출고 이벤트 FIFO 소진 ─────────────────────────────
    lot_q = [dict(l) for l in lots]
    all_events     = []
    dispatch_pool  = []   # 임가공 출고 이벤트에 lot_q_remain 부여한 풀

    for ev in outflow_events:
        need  = ev["qty"]
        attrs = []
        while need > 0.001 and lot_q:
            lot  = lot_q[0]
            take = min(lot["remain"], need)
            attrs.append({
                "lot_label":  lot["label"],
                "lot_date":   lot["date"],
                "unit_cost":  lot["unit_cost"],
                "qty":        round(take, 3),
                "amount":     round(take * lot["unit_cost"], 2),
                "kg_per_bag": lot.get("kg_per_bag", 510.0),  # Lot 고유 kg/백 비율 전파
            })
            lot["remain"] -= take
            need           -= take
            if lot["remain"] < 0.001:
                lot_q.pop(0)
        if need > 0.001:
            attrs.append({
                "lot_label":  "⚠️ 미기록 재고 (입고 이력 확인 필요)",
                "lot_date":   "",
                "unit_cost":  None,
                "qty":        round(need, 3),
                "amount":     None,
                "kg_per_bag": 510.0,
            })
        all_events.append({**ev, "attributions": attrs})

        if ev["type"] == "임가공출고":
            dispatch_pool.append({
                "dispatch_id":  ev["id"],
                "date":         ev["date"],
                "processor_id": ev["processor_id"],
                "total_qty":    ev["qty"],
                # lot 잔량 서브큐 (Level 2에서 소진) — kg_per_bag 포함
                "lot_q_remain": [
                    {"lot_label": a["lot_label"], "lot_date": a.get("lot_date",""),
                     "unit_cost": a["unit_cost"],  "remain":   a["qty"],
                     "kg_per_bag": a.get("kg_per_bag", 510.0)}
                    for a in attrs if (a.get("qty") or 0) > 0.001
                ],
            })

    # ── Level 2: 프로세서별 dispatch 풀 → B/L 배치 FIFO 소진 ─────────────────
    # 프로세서별 dispatch 큐 구성 (이미 날짜 오름차순)
    dq_by_proc = defaultdict(list)
    for de in dispatch_pool:
        dq_by_proc[de["processor_id"]].append(de)

    # processing_history → 선적일 오름차순 정렬
    ph_records = []
    for rec in cfg.get("processing_history", []):
        if rec.get("scrap_type_id") != scrap_id:
            continue
        inp_kg = float(rec.get("input_kg") or 0)
        if inp_kg <= 0:
            continue
        ship      = ship_map.get(rec.get("shipment_id", ""), {})
        load_date = (ship.get("loading_date") or "")[:10] or "9999-12-31"
        ph_records.append({
            "id":           rec.get("id", ""),
            "shipment_id":  rec.get("shipment_id", ""),
            "hbl":          ship.get("hbl", "미연결"),
            "load_date":    load_date,
            "input_kg":     inp_kg,
            "processor_id": rec.get("processor_id", ""),
        })
    ph_records.sort(key=lambda x: x["load_date"])

    bl_result = {}
    for ph in ph_records:
        dq      = dq_by_proc[ph["processor_id"]]   # mutable reference
        need    = ph["input_kg"]
        bl_lots = {}  # lot_label → {qty, amount, unit_cost, lot_date}

        while need > 0.001 and dq:
            de     = dq[0]
            avail  = sum(lr["remain"] for lr in de["lot_q_remain"])
            if avail < 0.001:
                dq.pop(0)
                continue
            take_dp = min(avail, need)
            dp_need = take_dp
            # dispatch 내부 lot 서브큐 FIFO 소진
            for lr in de["lot_q_remain"]:
                if dp_need < 0.001:
                    break
                lot_take = min(lr["remain"], dp_need)
                if lot_take > 0:
                    k = lr["lot_label"]
                    if k not in bl_lots:
                        bl_lots[k] = {"lot_label": k, "lot_date": lr.get("lot_date",""),
                                      "unit_cost": lr["unit_cost"], "qty": 0.0, "amount": 0.0,
                                      "storage_cost": 0.0, "storage_days_wsum": 0.0}
                    bl_lots[k]["qty"] += lot_take
                    if lr["unit_cost"] is not None:
                        bl_lots[k]["amount"] += round(lot_take * lr["unit_cost"], 4)
                    # ── 창고 보관비 자동 계산 (Lot 입고일 → 임가공 출고일) ─────
                    _ls_raw = lr.get("lot_date", "") or ""
                    _ll_chk = lr.get("lot_label", "")
                    # 기초재고 Lot은 기준일을 시작점으로 사용 (이전 보관비는 sunk cost)
                    if "기초재고" in _ll_chk:
                        _stor_start = op_date
                    elif len(_ls_raw) >= 10:
                        _stor_start = _ls_raw[:10]
                    elif len(_ls_raw) >= 7:
                        _stor_start = _ls_raw[:7] + "-15"  # 월만 있으면 중간값 근사
                    else:
                        _stor_start = ""
                    _sc_val = 0.0
                    _ldays  = 0
                    if _stor_start and de["date"] and de["date"] != "9999-12-31":
                        try:
                            _ldays = max(0, (
                                datetime.strptime(de["date"], "%Y-%m-%d") -
                                datetime.strptime(_stor_start, "%Y-%m-%d")
                            ).days)
                            _eur_r  = _get_eur_usd(cfg, de["date"][:7])
                            _stor_r = _storage_rate_eur(cfg, scrap_id)
                            # 톤백: Lot 고유 kg/백 비율 사용 (기초재고·입고 Lot별로 다를 수 있음)
                            _kpb   = lr.get("kg_per_bag", 510.0)
                            _tb    = lot_take / _kpb if _kpb > 0 else lot_take / 510.0
                            _sc_val = round(_tb * _ldays * _stor_r * _eur_r, 4)
                        except Exception:
                            _sc_val = 0.0
                            _ldays  = 0   # 예외 시 days도 초기화 — wsum 고아 누적 방지
                    bl_lots[k]["storage_cost"] += _sc_val
                    bl_lots[k]["storage_days_wsum"] += _ldays * lot_take  # 가중합: 나중에 평균 산출용
                    # ──────────────────────────────────────────────────────────
                    lr["remain"] -= lot_take
                    dp_need      -= lot_take
            de["lot_q_remain"] = [lr for lr in de["lot_q_remain"] if lr["remain"] > 0.001]
            need -= take_dp
            if not de["lot_q_remain"]:
                dq.pop(0)

        if need > 0.001:
            k = "⚠️ 출고 기록 미매칭 (출고 기록 탭에서 임가공 출고 입력 필요)"
            bl_lots[k] = {"lot_label": k, "lot_date": "", "unit_cost": None,
                          "qty": round(need, 3), "amount": None, "storage_cost": 0.0}

        sid = ph["shipment_id"] or f"__no_ship__{ph['id']}"
        if sid not in bl_result:
            bl_result[sid] = {"hbl": ph["hbl"], "load_date": ph["load_date"],
                              "input_kg": 0.0, "lots": {}, "storage_cost": 0.0}
        bl_result[sid]["input_kg"] += ph["input_kg"]
        for k, v in bl_lots.items():
            if k not in bl_result[sid]["lots"]:
                bl_result[sid]["lots"][k] = {**v}
            else:
                bl_result[sid]["lots"][k]["qty"] += v["qty"]
                if v.get("unit_cost") is not None:
                    bl_result[sid]["lots"][k]["amount"] = (
                        bl_result[sid]["lots"][k].get("amount", 0) + v.get("amount", 0))
                bl_result[sid]["lots"][k]["storage_cost"] = (
                    bl_result[sid]["lots"][k].get("storage_cost", 0.0) + v.get("storage_cost", 0.0))
                bl_result[sid]["lots"][k]["storage_days_wsum"] = (
                    bl_result[sid]["lots"][k].get("storage_days_wsum", 0.0) + v.get("storage_days_wsum", 0.0))
        bl_result[sid]["storage_cost"] += sum(v.get("storage_cost", 0.0) for v in bl_lots.values())

    return bl_result, all_events, lot_q   # lot_q: 미소진 잔량


def status_badge(s):
    m={"provisional":("Provisional 정산","b-wn"),"final":("최종정산","b-ok"),"paid":("입금완료","b-bp")}
    lbl,cls=m.get(s,("—","b-ng"))
    return f'<span class="{cls}">{lbl}</span>'

# ── Sidebar ───────────────────────────────────────────────────────────────────
cfg = load_cfg()
hist_opts = [h["month"] for h in sorted(cfg.get("index_history",[]),key=lambda x:x["month"],reverse=True)]

# NI/CO: 가장 최근 INDEX 이력 자동 적용 (사이드바 입력 제거)
_latest_idx = sorted(cfg.get("index_history", []), key=lambda x: x["month"], reverse=True)
NI = _latest_idx[0]["ni_index"] if _latest_idx else 17093.18
CO = _latest_idx[0]["co_index"] if _latest_idx else 56598.72

with st.sidebar:
    st.title("📊 현황")

    # ── 환율 ────────────────────────────────────────────────────────────────
    XR = st.number_input("💱 USD / KRW", value=1380.0, step=1.0, format="%.0f")
    st.divider()

    # ── 선적 현황 ────────────────────────────────────────────────────────────
    _sb_ships = cfg.get("shipments", [])
    _sb_today = date.today().isoformat()
    _sb_nosett = [s for s in _sb_ships if not s.get("status","") or s.get("status") == "provisional"]
    _sb_eta_soon = [
        s for s in _sb_ships
        if s.get("eta","") and s.get("eta","") >= _sb_today
        and s.get("eta","") <= (date.today() + timedelta(days=14)).isoformat()
        and s.get("status","") != "final"
    ]
    st.markdown("**🚢 선적 현황**")
    _sbc1, _sbc2 = st.columns(2)
    _sbc1.metric("전체", f"{len(_sb_ships)}건")
    _sbc2.metric("미확정", f"{len(_sb_nosett)}건")
    if _sb_eta_soon:
        st.warning(f"ETA 14일 이내  {len(_sb_eta_soon)}건")
        for _se in sorted(_sb_eta_soon, key=lambda x: x.get("eta","")):
            st.caption(f"· {_se.get('hbl','HBL미정')}  {_se.get('eta','')}")
    st.divider()

    # ── 재고 현황 ────────────────────────────────────────────────────────────
    st.markdown("**📦 재고 현황**")
    _sb_scraps = [s for s in cfg.get("scrap_types", []) if s.get("active", True)]
    if _sb_scraps:
        for _sc in _sb_scraps:
            _, _, _sb_lot_rem = _fifo_lot_trace(cfg, _sc["id"])
            _sb_rem_kg = sum(lot.get("remain", 0) for lot in _sb_lot_rem)
            st.metric(_sc["name"], f"{_sb_rem_kg:,.0f} kg")
    else:
        st.caption("스크랩 유형 없음")
    st.divider()

    # ── 경고 ────────────────────────────────────────────────────────────────
    _sb_ph_list    = cfg.get("processing_history", [])
    _sb_valid_sids = {s["id"] for s in _sb_ships}
    _sb_orphans    = sum(1 for p in _sb_ph_list
                         if p.get("shipment_id","") and p.get("shipment_id","") not in _sb_valid_sids)
    if _sb_orphans:
        st.error(f"⚠️ 고아 배치 {_sb_orphans}건\n임가공사 관리 > 세부 내역에서 확인")

    if _latest_idx:
        st.caption(f"INDEX 기준: {_latest_idx[0]['month']}  Ni ${NI:,.0f} / Co ${CO:,.0f}")

st.title("BP / BM 재고·손익 관리")

(t_report,t_bp,t_sens,
 t_ship,t_freight,t_pnl,
 t_buy,t_proc,t_stype,t_outflow,t_idx,t_docs,t_contract) = st.tabs([
    "📋 요약 보고서",
    "📊 BP/BM 매각 단가","📉 민감도 분석",
    "🚢 선적 정산 추적","🚛 포워더 운임 관리","💰 손익 분석",
    "🏢 매입사 관리","🏭 임가공사 관리","🗃️ 스크랩 유형 관리",
    "📥 입출고 기록","📈 INDEX 이력","📎 문서 보관함","📋 계약 이행"
])

active_buyers = [b for b in cfg["buyers"] if b.get("active",True)]
active_procs  = [p for p in cfg.get("processors",[]) if p.get("active",True)]
active_scraps = [s for s in cfg.get("scrap_types",[]) if s.get("active",True)]


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — BP/BM 매각 단가
# ══════════════════════════════════════════════════════════════════════════════
with t_bp:
    st.subheader("BP/BM 매각 단가 계산")
    with st.expander("⚙️ Metal INDEX ($/ton)", expanded=False):
        _bp_ref = st.selectbox("이력 불러오기", ["최신값 자동"] + hist_opts, key="bp_idx_ref")
        if _bp_ref == "최신값 자동":
            _bp_ni, _bp_co = NI, CO
        else:
            _hm2 = {h["month"]: h for h in cfg.get("index_history", [])}
            _bp_ni = _hm2[_bp_ref]["ni_index"] if _bp_ref in _hm2 else NI
            _bp_co = _hm2[_bp_ref]["co_index"] if _bp_ref in _hm2 else CO
        _bic1, _bic2 = st.columns(2)
        _bp_ni = _bic1.number_input("Ni INDEX (LME)", value=_bp_ni, step=10.0, format="%.2f", key="bp_ni")
        _bp_co = _bic2.number_input("Co INDEX (MB Rotterdam)", value=_bp_co, step=10.0, format="%.2f", key="bp_co")
    if not active_buyers: st.warning("매입사가 없습니다.")
    else:
        f1,f2=st.columns(2)
        with f1: fp=st.multiselect("품목",["BP","BM"],default=["BP","BM"],key="bp_fp")
        with f2:
            nm=sorted(set(b["name"] for b in active_buyers))
            fb=st.multiselect("매입사",nm,default=nm,key="bp_fb")
        show=[b for b in active_buyers if b["product"] in fp and b["name"] in fb]
        if show:
            rows=[]
            for b in show:
                nv,cv,tot,pkg=bp_price(_bp_ni,_bp_co,b["ni_content"],b["co_content"],b["ni_payable"],b["co_payable"])
                rows.append({"매입사":b["name"],"품목":b["product"],
                    "Ni 함유량":b["ni_content"],"Co 함유량":b["co_content"],
                    "Ni 지불율":b["ni_payable"],"Co 지불율":b["co_payable"],
                    "Ni Value($/ton)":round(nv,2),"Co Value($/ton)":round(cv,2),
                    "단가($/ton)":round(tot,2),"단가($/kg)":round(pkg,5),
                    "단가(원/ton)":round(tot*XR,0),"단가(원/kg)":round(pkg*XR,2)})
            cols=st.columns(len(show))
            for i,(b,r) in enumerate(zip(show,rows)):
                with cols[i]:
                    bd=f'<span class="b-{"bp" if b["product"]=="BP" else "bm"}">{b["product"]}</span>'
                    st.markdown(f"#### {b['name']}  {bd}",unsafe_allow_html=True)
                    st.markdown(f"""<div class="mbox"><div class="sp">매각 단가</div>
                      <div class="ph">${r['단가($/kg)']:.4f} / kg</div>
                      <div class="sp">${r['단가($/ton)']:,.2f} / ton</div></div>
                    <div class="mbox"><div class="sp">한화 (@{XR:,.0f})</div>
                      <div class="ph">₩{r['단가(원/kg)']:,.2f} / kg</div>
                      <div class="sp">₩{r['단가(원/ton)']:,.0f} / ton</div></div>
                    <div class="mbox"><div class="sp">Ni ({b['ni_content']}% × {b['ni_payable']})</div>
                      <div>${r['Ni Value($/ton)']:,.2f}/ton</div>
                      <div class="sp">Co ({b['co_content']}% × {b['co_payable']})</div>
                      <div>${r['Co Value($/ton)']:,.2f}/ton</div></div>""",unsafe_allow_html=True)
            st.divider()
            df=pd.DataFrame(rows)
            st.dataframe(df.style.format({"Ni 함유량":"{:.2f}%","Co 함유량":"{:.2f}%",
                "Ni 지불율":"{:.2f}","Co 지불율":"{:.2f}",
                "Ni Value($/ton)":"${:,.2f}","Co Value($/ton)":"${:,.2f}",
                "단가($/ton)":"${:,.2f}","단가($/kg)":"${:.5f}",
                "단가(원/ton)":"₩{:,.0f}","단가(원/kg)":"₩{:,.2f}"}),
                use_container_width=True,hide_index=True)
            st.download_button("📥 CSV",df.to_csv(index=False,encoding="utf-8-sig"),
                f"BP_BM_{date.today():%Y%m%d}.csv","text/csv")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — 민감도 분석
# ══════════════════════════════════════════════════════════════════════════════
with t_sens:
    st.subheader("민감도 분석")
    if not active_buyers: st.warning("매입사가 없습니다.")
    else:
        sb1,sb2=st.columns(2)
        with sb1: sens_b=st.selectbox("매입사",[f"{b['name']} ({b['product']})" for b in active_buyers],key="sens_b")
        with sb2: sens_t=st.selectbox("변동 대상",["Ni INDEX","Co INDEX","Ni + Co 동시"],key="sens_t")
        sel=active_buyers[[f"{b['name']} ({b['product']})" for b in active_buyers].index(sens_b)]
        st.caption(f"기준 — Ni: ${NI:,.2f} / Co: ${CO:,.2f}  (최신 INDEX 자동 적용)")
        rng=st.slider("변동 범위 (%)",min_value=-30,max_value=30,value=(-20,20),step=5)
        steps=list(range(rng[0],rng[1]+1,5))
        srows=[]
        for pct in steps:
            f=1+pct/100
            ni_v=NI*f if sens_t in ["Ni INDEX","Ni + Co 동시"] else NI
            co_v=CO*f if sens_t in ["Co INDEX","Ni + Co 동시"] else CO
            _,_,tot,pkg=bp_price(ni_v,co_v,sel["ni_content"],sel["co_content"],sel["ni_payable"],sel["co_payable"])
            srows.append({"변동률":f"{pct:+d}%","Ni INDEX($/ton)":round(ni_v,2),
                "Co INDEX($/ton)":round(co_v,2),"단가($/kg)":round(pkg,5),"단가(원/kg)":round(pkg*XR,2)})
        df_s=pd.DataFrame(srows)

        def hl_base(row):
            if row["변동률"]=="+0%":
                return ["background-color:#1f4e79;color:white" for _ in row]
            return [""]*len(row)

        st.dataframe(df_s.style.apply(hl_base,axis=1).format({
            "Ni INDEX($/ton)":"${:,.2f}","Co INDEX($/ton)":"${:,.2f}",
            "단가($/kg)":"${:.5f}","단가(원/kg)":"₩{:,.2f}"}),
            use_container_width=True,hide_index=True)
        st.line_chart(df_s.set_index("변동률")[["단가($/kg)"]])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — 선적 정산 추적  (Provisional/Final 비교 통합)
# ══════════════════════════════════════════════════════════════════════════════
with t_ship:
    st.subheader("선적 정산 추적")
    with st.expander("ℹ️ 정산 프로세스 안내", expanded=False):
        st.markdown("""
**Provisional 정산 (Provisional, M-1)**
선적 후 바이어 도착 전, 전월 INDEX 기준으로 임시 Invoice를 발행합니다.
Ni/Co 지불율과 M-1 INDEX를 적용한 단가로 먼저 대금을 수취합니다.

**최종정산 (Final, M+0)**
바이어가 실제 인수 검사한 함량·중량 확정 후, 선적월 INDEX로 재정산합니다.
Provisional 정산액과의 차액을 추가 수취 또는 반환합니다.

**입금완료 (Paid)**
최종정산 금액까지 모두 수취 완료된 상태입니다.
""")
    shipments = cfg.get("shipments",[])
    buyer_map  = {b["id"]:b for b in cfg["buyers"]}
    buyer_opts = {f"{b['name']} ({b['product']})":b["id"] for b in cfg["buyers"]}
    hm_all = {h["month"]:h for h in cfg.get("index_history",[])}

    # ── 요약 메트릭 ──
    if shipments:
        total_wkg   = sum(s.get("weight_kg",0) for s in shipments)
        total_inv   = sum(s.get("invoice_usd",0) for s in shipments)
        prov_cnt    = sum(1 for s in shipments if s.get("status")=="provisional")
        final_cnt   = sum(1 for s in shipments if s.get("status")=="final")
        paid_cnt    = sum(1 for s in shipments if s.get("status")=="paid")
        m1,m2,m3,m4,m5=st.columns(5)
        m1.metric("총 선적건",f"{len(shipments)}건")
        m2.metric("총 중량",f"{total_wkg/1000:,.1f} ton")
        m3.metric("Provisional 정산",f"{prov_cnt}건")
        m4.metric("최종정산",f"{final_cnt}건")
        m5.metric("입금완료",f"{paid_cnt}건")
        # ── 미완료 알림 ──
        _pend = [s for s in shipments if s.get("status") in ("provisional","final")]
        if _pend:
            _pend_msgs = []
            for _ps in _pend:
                _pb = buyer_map.get(_ps.get("buyer_id",""),{})
                _st_lbl = "🟡 Provisional" if _ps.get("status")=="provisional" else "🟢 최종(미입금)"
                _pend_msgs.append(f"{_st_lbl} **{_ps.get('hbl','—')}** ({_pb.get('name','?')}, {_ps.get('loading_date','?')})")
            st.warning("⚠️ **정산 미완료 건이 있습니다.**\n\n" + "  \n".join(_pend_msgs))
        st.divider()

    # ── 필터 ──
    sf1,sf2=st.columns(2)
    with sf1: flt_stat=st.multiselect("상태 필터",["provisional","final","paid"],
                                       default=["provisional","final","paid"],key="ship_stat")
    with sf2: flt_buy=st.multiselect("매입사 필터",list(buyer_opts.keys()),
                                      default=list(buyer_opts.keys()),key="ship_buy")

    flt_ids={v for k,v in buyer_opts.items() if k in flt_buy}
    show_ships=[s for s in shipments if s.get("status","provisional") in flt_stat
                and (not s.get("buyer_id") or s.get("buyer_id") in flt_ids)]
    show_ships=sorted(show_ships, key=lambda x: x.get("loading_date","9999"))

    if not show_ships and shipments:
        st.info("필터 조건에 맞는 선적건이 없습니다.")
    elif not shipments:
        st.info("등록된 선적건이 없습니다. 아래에서 추가하세요.")
    else:
        # ── 📅 항차 일정 타임라인 ──────────────────────────────────────────────
        st.markdown("#### 📅 항차 일정")
        _tl = []
        for _si, _s in enumerate(show_ships, 1):
            _b2  = buyer_map.get(_s.get("buyer_id",""), {})
            _ld2 = _s.get("loading_date","") or "미정"
            _eta2= _s.get("eta","") or "TBD"
            _tl.append({
                "No.":    _si,
                "HBL":    _s.get("hbl","—"),
                "매입사": f"{_b2.get('name','?')} ({_b2.get('product','?')})",
                "선적일": _ld2,
                "ETA":    _eta2,
                "중량(kg)": _s.get("weight_kg",0),
                "상태":   {"provisional":"🟡 Provisional","final":"🟢 최종","paid":"🔵 입금"}.get(_s.get("status","provisional"),"—"),
            })
        st.dataframe(pd.DataFrame(_tl).style.format({"중량(kg)":"{:,.0f}"}),
                     use_container_width=True, hide_index=True)

        # Gantt 차트
        try:
            import plotly.express as px
            _gantt = []
            for _r in _tl:
                if _r["선적일"] == "미정": continue
                try:
                    _s_dt = datetime.strptime(_r["선적일"], "%Y-%m-%d")
                    _e_dt = (datetime.strptime(_r["ETA"], "%Y-%m-%d")
                             if _r["ETA"] != "TBD" else _s_dt + timedelta(days=60))
                    _gantt.append({
                        "항차":   f"#{_r['No.']}  {_r['HBL']}",
                        "매입사": _r["매입사"],
                        "ETD":    _s_dt,
                        "ETA":    _e_dt,
                        "상태":   _r["상태"],
                    })
                except: pass
            if _gantt:
                _df_g = pd.DataFrame(_gantt)
                # 막대 안 라벨: 매입사 약칭 + 항해일수
                _df_g["라벨"] = _df_g.apply(
                    lambda r: f"{r['매입사'].split('(')[0].strip()}  "
                              f"({(r['ETA']-r['ETD']).days}일)", axis=1)
                # 상태별 색상 — 진하게
                _cmap = {"🟡 Provisional":"#1F4E79","🟢 최종":"#1E8449","🔵 입금":"#117A65"}
                _fig  = px.timeline(
                    _df_g, x_start="ETD", x_end="ETA", y="항차",
                    color="상태", text="라벨",
                    hover_name="항차",
                    hover_data={"매입사":True,"ETD":True,"ETA":True,"라벨":False,"상태":False},
                    color_discrete_map=_cmap,
                )
                _fig.update_traces(
                    textposition="inside",
                    insidetextanchor="middle",
                    textfont=dict(color="white", size=10, family="Arial"),
                    marker_line=dict(color="rgba(255,255,255,0.4)", width=1),
                )
                # 오늘 날짜 기준선
                _today_dt = datetime.today()
                _fig.add_vline(
                    x=_today_dt.timestamp()*1000,
                    line_dash="dot", line_color="#E74C3C", line_width=2,
                    annotation_text=f"오늘 ({_today_dt.strftime('%m/%d')})",
                    annotation_font=dict(color="#E74C3C", size=10),
                    annotation_position="top left",
                )
                _fig.update_yaxes(
                    autorange="reversed",
                    tickfont=dict(size=11, color="#D0D0E8"),
                    gridcolor="rgba(255,255,255,0.08)",
                )
                _fig.update_xaxes(
                    showgrid=True, gridcolor="rgba(255,255,255,0.08)",
                    tickformat="%m/%d",
                    tickfont=dict(size=10, color="#D0D0E8"),
                )
                _fig.update_layout(
                    height=max(280, len(_gantt)*54+110),
                    xaxis_title="", yaxis_title="",
                    margin=dict(l=150, r=20, t=50, b=20),
                    plot_bgcolor="#1E1E2E",
                    paper_bgcolor="#16213E",
                    font=dict(color="#D0D0E8"),
                    legend=dict(
                        title="",
                        orientation="h",
                        yanchor="bottom", y=1.02,
                        xanchor="right", x=1,
                        font=dict(size=11, color="#D0D0E8"),
                        bgcolor="rgba(30,30,50,0.85)",
                    ),
                )
                st.plotly_chart(_fig, use_container_width=True)
        except ImportError:
            st.caption("plotly 설치 시 Gantt 차트 표시 — `pip install plotly`")

        st.divider()

        # ── 개별 선적건 expander ────────────────────────────────────────────────
        _ship_id_idx = {sh.get("id"): _ri for _ri, sh in enumerate(shipments)}
        for i,s in enumerate(show_ships):
            real_i=_ship_id_idx[s["id"]]
            b=buyer_map.get(s.get("buyer_id"),{})
            buyer_lbl=f"{b.get('name','?')} ({b.get('product','?')})"
            # 상태 텍스트 (expander는 HTML 미지원 → 이모지 사용)
            stat_txt={"provisional":"🟡 Provisional 정산","final":"🟢 최종정산","paid":"🔵 입금완료"}.get(s.get("status","provisional"),"—")
            # 정산 미리보기 (헤더에 표시)
            settle_preview=""
            if (b and s.get("prov_month","—")!="—" and s.get("final_month","—")!="—"
                    and s.get("prov_month") in hm_all and s.get("final_month") in hm_all):
                pm_h=hm_all[s["prov_month"]]; fm_h=hm_all[s["final_month"]]
                _,_,_,ppk=bp_price(pm_h["ni_index"],pm_h["co_index"],b.get("ni_content",0),b.get("co_content",0),b.get("ni_payable",0),b.get("co_payable",0))
                bni=s.get("buyer_ni_content") or b.get("ni_content",0)
                bco=s.get("buyer_co_content") or b.get("co_content",0)
                moist=s.get("moisture_pct") or 0
                fw=s.get("weight_kg",0)*(1-moist/100)
                _,_,_,fpk=bp_price(fm_h["ni_index"],fm_h["co_index"],bni,bco,b.get("ni_payable",0),b.get("co_payable",0))
                net_s=(fpk*fw)-s.get("invoice_usd",0)+(s.get("other_adj_usd") or 0)
                settle_preview=f"  |  추가정산: ${net_s:+,.2f}"
            ld_disp  = s.get("loading_date","").strip() or "선적일 미정"
            eta_disp = s.get("eta","").strip() or "TBD"
            hdr=f"#{i+1}  {stat_txt}  {s.get('hbl','—')}  |  {buyer_lbl}  |  {ld_disp}  →  ETA {eta_disp}  |  {s.get('weight_kg',0):,.0f} kg{settle_preview}"
            with st.expander(hdr,expanded=False):
                # ── 기본 정보 입력 ──
                e1,e2,e3,e4=st.columns(4)
                with e1:
                    new_hbl=st.text_input("HBL",s.get("hbl",""),key=f"sh_hbl_{real_i}")
                    new_inv=st.text_input("Invoice No.",s.get("invoice_no",""),key=f"sh_inv_{real_i}")
                    new_ld =st.text_input("선적일 (YYYY-MM-DD)",s.get("loading_date",""),key=f"sh_ld_{real_i}")
                with e2:
                    cur_b_lbl=[k for k,v in buyer_opts.items() if v==s.get("buyer_id","")]
                    new_b=st.selectbox("매입사",list(buyer_opts.keys()),
                        index=list(buyer_opts.keys()).index(cur_b_lbl[0]) if cur_b_lbl else 0,key=f"sh_buy_{real_i}")
                    new_wkg=st.number_input("선적 중량 (kg)",value=float(s.get("weight_kg",0)),step=1.0,format="%.0f",key=f"sh_wkg_{real_i}")
                    new_iusd=st.number_input("Invoice (USD, Provisional 정산액)",value=float(s.get("invoice_usd",0)),step=1.0,format="%.2f",key=f"sh_iusd_{real_i}")
                    new_eusd=st.number_input("수출비 (USD)",value=float(s.get("export_cost_usd") or 0),step=1.0,format="%.2f",key=f"sh_eusd_{real_i}",
                        help="Ocean Freight, THC 등 이 선적건 전체 수출비 합계. 배치별 생산량 비례로 자동 배분됩니다.")
                with e3:
                    new_pm=st.selectbox("Provisional 월",["—"]+hist_opts,
                        index=(["—"]+hist_opts).index(s.get("prov_month","—")) if s.get("prov_month","—") in ["—"]+hist_opts else 0,
                        key=f"sh_pm_{real_i}")
                    new_fm=st.selectbox("Final 월",["—"]+hist_opts,
                        index=(["—"]+hist_opts).index(s.get("final_month","—")) if s.get("final_month","—") in ["—"]+hist_opts else 0,
                        key=f"sh_fm_{real_i}")
                    new_stat=st.selectbox("상태",["provisional","final","paid"],
                        index=["provisional","final","paid"].index(s.get("status","provisional")),key=f"sh_stat_{real_i}")
                with e4:
                    new_etd=st.text_input("ETD",s.get("etd",""),key=f"sh_etd_{real_i}")
                    new_eta=st.text_input("ETA",s.get("eta",""),key=f"sh_eta_{real_i}")
                    new_note=st.text_input("비고",s.get("notes",""),key=f"sh_note_{real_i}")

                # ── 확정산 세부 정보 ──
                st.markdown("---")
                st.markdown("**📊 확정산 상세 (수분·분석값·기타 조정)**")
                sa1,sa2,sa3=st.columns(3)
                with sa1:
                    st.markdown("**수분 공제**")
                    new_moisture=st.number_input("수분 공제율 (%)",
                        value=float(s.get("moisture_pct") or 0),
                        min_value=0.0,max_value=20.0,step=0.1,format="%.2f",
                        key=f"sh_moist_{real_i}")
                    final_weight_disp=new_wkg*(1-new_moisture/100)
                    st.caption(f"정산 중량: **{final_weight_disp:,.1f} kg** ({new_wkg:,.0f} → {final_weight_disp:,.1f})")
                with sa2:
                    st.markdown("**매입사 샘플 분석값**")
                    default_ni=float(s.get("buyer_ni_content") or (b.get("ni_content",0) if b else 0))
                    default_co=float(s.get("buyer_co_content") or (b.get("co_content",0) if b else 0))
                    new_buyer_ni=st.number_input("Ni 분석값 (%)",
                        value=default_ni,step=0.01,format="%.2f",
                        key=f"sh_bni_{real_i}")
                    new_buyer_co=st.number_input("Co 분석값 (%)",
                        value=default_co,step=0.01,format="%.2f",
                        key=f"sh_bco_{real_i}")
                    if b:
                        ni_diff=new_buyer_ni-b.get("ni_content",0)
                        co_diff=new_buyer_co-b.get("co_content",0)
                        st.caption(f"초기값 대비: Ni {ni_diff:+.2f}%p / Co {co_diff:+.2f}%p")
                with sa3:
                    st.markdown("**기타 조정**")
                    new_other_adj=st.number_input("기타 조정 (USD)",
                        value=float(s.get("other_adj_usd") or 0),
                        step=1.0,format="%.2f",
                        help="+: 추가 수령 / -: 추가 지급",
                        key=f"sh_adj_{real_i}")
                    new_other_desc=st.text_input("조정 사유",
                        s.get("other_adj_desc",""),
                        key=f"sh_adjd_{real_i}")

                # ── 정산 요약 계산 ──
                if b and new_pm!="—" and new_pm in hm_all:
                    pm_data=hm_all[new_pm]
                    _,_,_,prov_pkg=bp_price(pm_data["ni_index"],pm_data["co_index"],
                        b.get("ni_content",0),b.get("co_content",0),
                        b.get("ni_payable",0),b.get("co_payable",0))
                    st.markdown("---")
                    if new_fm!="—" and new_fm in hm_all:
                        # 확정산 계산
                        fm_data=hm_all[new_fm]
                        _,_,_,final_pkg=bp_price(fm_data["ni_index"],fm_data["co_index"],
                            new_buyer_ni,new_buyer_co,
                            b.get("ni_payable",0),b.get("co_payable",0))
                        final_w=new_wkg*(1-new_moisture/100)
                        prov_amt=new_iusd
                        final_amt=final_pkg*final_w
                        index_diff=(final_pkg-prov_pkg)
                        net_settle=final_amt-prov_amt+new_other_adj

                        st.markdown("**📋 정산 요약**")
                        rs1,rs2,rs3,rs4=st.columns(4)
                        with rs1:
                            st.metric("Provisional 단가",f"${prov_pkg:.5f}/kg",
                                      help=f"INDEX {new_pm}: Ni ${pm_data['ni_index']:,.2f} / Co ${pm_data['co_index']:,.2f}")
                            st.metric("Provisional 정산액",f"${prov_amt:,.2f}")
                        with rs2:
                            st.metric("Final 단가",f"${final_pkg:.5f}/kg",
                                      delta=f"{index_diff:+.5f}",
                                      help=f"INDEX {new_fm}: Ni ${fm_data['ni_index']:,.2f} / Co ${fm_data['co_index']:,.2f}")
                            st.metric("정산 중량",f"{final_w:,.1f} kg",
                                      delta=f"{final_w-new_wkg:+,.1f} kg (수분 {new_moisture:.1f}%)")
                        with rs3:
                            st.metric("최종정산액",f"${final_amt:,.2f}")
                            if new_other_adj!=0:
                                st.metric("기타 조정",f"${new_other_adj:+,.2f}",
                                          help=new_other_desc if new_other_desc else "기타")
                        with rs4:
                            net_color="normal" if net_settle>=0 else "inverse"
                            st.metric("🔁 추가정산 (Final − Prov + 기타)",
                                      f"${net_settle:+,.2f}",
                                      delta="추가 수령" if net_settle>=0 else "추가 지급",
                                      delta_color=net_color)
                            st.metric("KRW",f"₩{net_settle*XR:+,.0f}")

                        # 상세 항목 표
                        st.markdown(f"""| 항목 | Provisional | Final | 비고 |
|------|------------|-------|------|
|기준 월|{new_pm}|{new_fm}||
|Ni INDEX|${pm_data['ni_index']:,.2f}|${fm_data['ni_index']:,.2f}|${fm_data['ni_index']-pm_data['ni_index']:+,.2f}|
|Co INDEX|${pm_data['co_index']:,.2f}|${fm_data['co_index']:,.2f}|${fm_data['co_index']-pm_data['co_index']:+,.2f}|
|Ni 함유량|{b.get('ni_content',0):.2f}%|{new_buyer_ni:.2f}%|{new_buyer_ni-b.get('ni_content',0):+.2f}%p|
|Co 함유량|{b.get('co_content',0):.2f}%|{new_buyer_co:.2f}%|{new_buyer_co-b.get('co_content',0):+.2f}%p|
|중량|{new_wkg:,.0f} kg|{final_w:,.1f} kg|수분 {new_moisture:.1f}% 공제|
|단가 ($/kg)|${prov_pkg:.5f}|${final_pkg:.5f}|${index_diff:+.5f}|
|정산금액|${prov_amt:,.2f}|${final_amt:,.2f}||
|기타 조정|—|${new_other_adj:+,.2f}|{new_other_desc}|
|**추가정산**|—|**${net_settle:+,.2f}**|{"🟢 수령" if net_settle>=0 else "🔴 지급"}|""")
                    else:
                        # Provisional만 있는 경우
                        st.info(f"Provisional ({new_pm}): **${prov_pkg:.5f}/kg**  |  Provisional 정산액: **${new_iusd:,.2f}**  — Final 월을 선택하면 추가정산 계산이 가능합니다.")

                ca,cb=st.columns(2)
                with ca:
                    if st.button("💾 저장",key=f"sh_save_{real_i}"):
                        _save_err = []
                        # 선적일 형식 검사
                        if new_ld:
                            try: datetime.strptime(new_ld, "%Y-%m-%d")
                            except ValueError: _save_err.append("선적일 형식이 잘못됐습니다 (YYYY-MM-DD)")
                        # HBL 중복 체크 (자기 자신 제외)
                        if new_hbl:
                            _dup = [s2 for j2,s2 in enumerate(shipments) if j2!=real_i and s2.get("hbl","").strip()==new_hbl.strip()]
                            if _dup: _save_err.append(f"HBL '{new_hbl}' 이(가) 이미 존재합니다")
                        if _save_err:
                            for _e in _save_err: st.error(_e)
                        else:
                            cfg["shipments"][real_i].update({
                                "hbl":new_hbl,"invoice_no":new_inv,
                                "loading_date":new_ld,"buyer_id":buyer_opts[new_b],
                                "weight_kg":new_wkg,"invoice_usd":new_iusd,
                                "export_cost_usd":new_eusd if new_eusd > 0 else None,
                                "prov_month":new_pm,"final_month":new_fm,
                                "status":new_stat,"etd":new_etd,"eta":new_eta,"notes":new_note,
                                "moisture_pct":new_moisture if new_moisture > 0 else None,
                                "buyer_ni_content":new_buyer_ni if new_buyer_ni!=default_ni or s.get("buyer_ni_content") else None,
                                "buyer_co_content":new_buyer_co if new_buyer_co!=default_co or s.get("buyer_co_content") else None,
                                "other_adj_usd":new_other_adj if new_other_adj else None,
                                "other_adj_desc":new_other_desc})
                            save_cfg(cfg); st.toast("✅ 저장 완료"); st.rerun()
                with cb:
                    with st.popover("🗑️ 삭제", use_container_width=True):
                        st.warning(f"**{s.get('hbl','?')}** 선적건을 삭제합니다.")
                        if st.button("삭제 확인", key=f"sh_del_cfm_{real_i}", type="primary", use_container_width=True):
                            cfg["shipments"].pop(real_i); save_cfg(cfg); st.rerun()

        # ── 전체 요약 테이블 ──
        if show_ships:
            st.divider()
            tbl_rows=[]
            for s in show_ships:
                bx=buyer_map.get(s.get("buyer_id",""),{})
                # 추가정산 계산 (가능한 경우)
                net_disp="—"
                pm2=s.get("prov_month","—"); fm2=s.get("final_month","—")
                if (bx and pm2!="—" and fm2!="—" and pm2 in hm_all and fm2 in hm_all):
                    _,_,_,fpkg2=bp_price(hm_all[fm2]["ni_index"],hm_all[fm2]["co_index"],
                        s.get("buyer_ni_content") or bx.get("ni_content",0),
                        s.get("buyer_co_content") or bx.get("co_content",0),
                        bx.get("ni_payable",0),bx.get("co_payable",0))
                    mst=s.get("moisture_pct") or 0
                    fw2=s.get("weight_kg",0)*(1-mst/100)
                    net_v=fpkg2*fw2 - s.get("invoice_usd",0) + (s.get("other_adj_usd") or 0)
                    net_disp=f"${net_v:+,.2f}"
                tbl_rows.append({
                    "HBL":s.get("hbl","—"),
                    "매입사":f"{bx.get('name','?')} ({bx.get('product','?')})",
                    "선적일":s.get("loading_date",""),
                    "중량(kg)":s.get("weight_kg",0),
                    "Invoice(USD)":s.get("invoice_usd",0),
                    "Prov월":s.get("prov_month","—"),
                    "Final월":s.get("final_month","—"),
                    "추가정산(USD)":net_disp,
                    "상태":s.get("status","provisional")})
            st.dataframe(pd.DataFrame(tbl_rows).style.format({"중량(kg)":"{:,.0f}","Invoice(USD)":"${:,.2f}"}),
                         use_container_width=True,hide_index=True)
            st.download_button("📥 CSV",pd.DataFrame(tbl_rows).to_csv(index=False,encoding="utf-8-sig"),
                f"선적정산_{date.today():%Y%m%d}.csv","text/csv")

    st.divider()
    st.subheader("새 선적건 추가")
    with st.form("add_ship"):
        a1,a2,a3,a4=st.columns(4)
        with a1: a_hbl=st.text_input("HBL"); a_inv=st.text_input("Invoice No.")
        with a2: a_buy=st.selectbox("매입사",list(buyer_opts.keys())); a_ld=st.text_input("선적일 (YYYY-MM-DD)",placeholder="2026-04-01")
        with a3: a_etd=st.text_input("ETD (YYYY-MM-DD)",placeholder="2026-04-05"); a_wkg=st.number_input("중량 (kg)",value=0.0,step=1.0,format="%.0f")
        with a4: a_iusd=st.number_input("Invoice (USD)",value=0.0,step=1.0,format="%.2f"); a_pm=st.selectbox("Provisional 월",["—"]+hist_opts); a_fm=st.selectbox("Final 월",["—"]+hist_opts)
        if st.form_submit_button("➕ 추가"):
            cfg["shipments"].append({"id":str(uuid.uuid4())[:8],"hbl":a_hbl,"invoice_no":a_inv,
                "loading_date":a_ld,"etd":a_etd.strip(),"buyer_id":buyer_opts[a_buy],"weight_kg":a_wkg,"invoice_usd":a_iusd,
                "prov_month":a_pm,"final_month":a_fm,"status":"provisional","eta":"","notes":"",
                "moisture_pct":None,"buyer_ni_content":None,"buyer_co_content":None,
                "other_adj_usd":None,"other_adj_desc":""})
            save_cfg(cfg); st.toast("✅ 추가 완료!"); st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — 포워더 운임 관리
# ══════════════════════════════════════════════════════════════════════════════
with t_freight:
    st.subheader("포워더 운임 관리")
    st.caption("포워더별 운임 견적을 관리합니다. 임가공사 관리 탭의 수출비 설정에 활용됩니다.")

    # EUR/USD 환율 설정
    eur_usd_cur = cfg.get("freight", {}).get("eur_usd", 1.08)
    with st.expander("⚙️ EUR/USD 환율 설정 (운임 계산용 고정값)", expanded=False):
        st.caption("⚠️ 이 EUR/USD는 **운임 견적 단가 변환 전용** 고정값입니다.  \n"
                   "scrap보관비 자동계산에 사용되는 **월별 EUR/USD**는 → **INDEX 이력 탭**에서 관리합니다.")
        with st.form("eur_usd_form"):
            new_eur_usd = st.number_input("EUR/USD 환율 (운임 전용)", value=eur_usd_cur, step=0.0001, format="%.4f")
            if st.form_submit_button("💾 저장"):
                cfg.setdefault("freight", {})["eur_usd"] = new_eur_usd
                save_cfg(cfg); st.success("EUR/USD 저장됨"); st.rerun()

    eur_usd_val = cfg.get("freight", {}).get("eur_usd", 1.08)
    fwd_list_tab = cfg.get("forwarders", [])
    st.divider()

    if not fwd_list_tab:
        st.info("등록된 포워더가 없습니다. 아래에서 추가하세요.")

    for fi, fwd in enumerate(fwd_list_tab):
        q_count = len(fwd.get("quotes", []))
        with st.expander(
            f"{'✅' if fwd.get('active', True) else '⛔'}  {fwd['name']}  ({q_count}개 견적)",
            expanded=True
        ):
            fhc1, fhc2 = st.columns([3, 1])
            with fhc1: fnm = st.text_input("포워더명", fwd["name"], key=f"fwd_nm_{fi}")
            with fhc2: fact = st.checkbox("활성", fwd.get("active", True), key=f"fwd_act_{fi}")

            quotes = fwd.get("quotes", [])
            if not quotes:
                st.caption("등록된 견적이 없습니다.")

            for qi, q in enumerate(quotes):
                curr    = q.get("currency", "USD")
                dest    = q.get("destination","").strip() or "목적지 미지정"
                t_items = (q.get("pickup_rail",0)+q.get("ocean_freight",0)+q.get("dg_surcharge",0)
                           +q.get("fuel_surcharge",0)+q.get("documentation",0)+q.get("terminal_handling",0))
                t_usd   = t_items * eur_usd_val if curr == "EUR" else t_items
                cap     = q.get("capacity_kg", 1) or 1
                pkg_usd = t_usd / cap

                q_title = (f"견적 {qi+1}  ·  [{dest}]  {q.get('label','—')}  ·  "
                           f"{q.get('container_type','?').upper()}  ·  "
                           f"{curr} {t_items:,.0f}  →  ${t_usd:,.2f}  (${pkg_usd:.4f}/kg)")
                with st.expander(q_title, expanded=False):
                    qr1, qr2, qr3, qr4 = st.columns(4)
                    with qr1:
                        q_dest  = st.text_input("목적지", q.get("destination",""),
                                                placeholder="예: CIF BUSAN", key=f"fwd_qdest_{fi}_{qi}")
                        q_label = st.text_input("견적명", q.get("label",""), key=f"fwd_ql_{fi}_{qi}")
                        q_cntr  = st.selectbox("컨테이너", ["20ft","40ft"],
                                               ["20ft","40ft"].index(q.get("container_type","20ft")),
                                               key=f"fwd_qcntr_{fi}_{qi}")
                        q_cap   = st.number_input("용량 (kg)", value=float(q.get("capacity_kg",20000)),
                                                  step=100.0, format="%.0f", key=f"fwd_qcap_{fi}_{qi}")
                        q_curr  = st.selectbox("통화", ["USD","EUR"],
                                               ["USD","EUR"].index(q.get("currency","USD")),
                                               key=f"fwd_qcurr_{fi}_{qi}")
                    with qr2:
                        st.markdown(f"**비용 항목 ({q.get('currency','USD')})**")
                        q_pr = st.number_input("픽업+Rail",        value=float(q.get("pickup_rail",0)),    step=1.0, format="%.0f", key=f"fwd_pr_{fi}_{qi}")
                        q_of = st.number_input("Ocean Freight",    value=float(q.get("ocean_freight",0)),  step=1.0, format="%.0f", key=f"fwd_of_{fi}_{qi}")
                        q_dg = st.number_input("DG Surcharge",     value=float(q.get("dg_surcharge",0)),   step=1.0, format="%.0f", key=f"fwd_dg_{fi}_{qi}")
                    with qr3:
                        st.markdown("‎")
                        q_fs = st.number_input("Fuel Surcharge",   value=float(q.get("fuel_surcharge",0)), step=1.0, format="%.0f", key=f"fwd_fs_{fi}_{qi}")
                        q_dc = st.number_input("Documentation",    value=float(q.get("documentation",0)),  step=1.0, format="%.0f", key=f"fwd_dc_{fi}_{qi}")
                        q_th = st.number_input("Terminal Handling", value=float(q.get("terminal_handling",0)), step=1.0, format="%.0f", key=f"fwd_th_{fi}_{qi}")
                    with qr4:
                        t_loc = q_pr+q_of+q_dg+q_fs+q_dc+q_th
                        t_u   = t_loc * eur_usd_val if q_curr=="EUR" else t_loc
                        t_e   = t_loc / eur_usd_val if q_curr=="USD" else t_loc
                        p_u   = t_u / q_cap if q_cap > 0 else 0
                        st.markdown("**합계**")
                        st.metric(f"합계 ({q_curr})", f"{q_curr} {t_loc:,.0f}")
                        if q_curr=="EUR":
                            st.metric("합계 (USD)", f"${t_u:,.2f}", help=f"× {eur_usd_val} EUR/USD")
                        else:
                            st.metric("합계 (EUR)", f"€{t_e:,.2f}", help=f"÷ {eur_usd_val} EUR/USD")
                        st.metric("단가 ($/kg)", f"${p_u:.4f}")
                        q_notes = st.text_input("비고", q.get("notes",""), key=f"fwd_qnotes_{fi}_{qi}")

                    qs1, qs2 = st.columns(2)
                    with qs1:
                        if st.button("💾 견적 저장", key=f"fwd_qsave_{fi}_{qi}", use_container_width=True):
                            cfg["forwarders"][fi]["quotes"][qi].update({
                                "destination":q_dest, "label":q_label, "container_type":q_cntr, "capacity_kg":q_cap,
                                "currency":q_curr, "pickup_rail":q_pr, "ocean_freight":q_of,
                                "dg_surcharge":q_dg, "fuel_surcharge":q_fs,
                                "documentation":q_dc, "terminal_handling":q_th, "notes":q_notes})
                            save_cfg(cfg); st.success("견적 저장됨"); st.rerun()
                    with qs2:
                        with st.popover("🗑️", use_container_width=True):
                            st.warning(f"견적 **{q.get('label','?')}** 삭제")
                            if st.button("삭제 확인", key=f"fwd_qdel_cfm_{fi}_{qi}", type="primary", use_container_width=True):
                                cfg["forwarders"][fi]["quotes"].pop(qi)
                                save_cfg(cfg); st.rerun()

            st.markdown("---")
            fb1, fb2, fb3 = st.columns(3)
            with fb1:
                if st.button("➕ 견적 추가", key=f"fwd_qadd_{fi}", use_container_width=True):
                    cfg["forwarders"][fi].setdefault("quotes", []).append({
                        "id":str(uuid.uuid4())[:8], "destination":"", "label":"새 견적",
                        "container_type":"20ft", "capacity_kg":20000, "currency":"USD",
                        "pickup_rail":0, "ocean_freight":0, "dg_surcharge":0,
                        "fuel_surcharge":0, "documentation":0, "terminal_handling":0, "notes":""})
                    save_cfg(cfg); st.rerun()
            with fb2:
                if st.button("💾 포워더 저장", key=f"fwd_save_{fi}", use_container_width=True):
                    cfg["forwarders"][fi].update({"name":fnm, "active":fact})
                    save_cfg(cfg); st.toast("✅ 저장됨"); st.rerun()
            with fb3:
                with st.popover("🗑️ 포워더 삭제", use_container_width=True):
                    st.warning(f"포워더 **{fwd.get('name','?')}** 전체 삭제 (견적 포함)")
                    if st.button("삭제 확인", key=f"fwd_del_cfm_{fi}", type="primary", use_container_width=True):
                        cfg["forwarders"].pop(fi)
                        save_cfg(cfg); st.rerun()

    st.divider()
    st.subheader("새 포워더 추가")
    with st.form("add_forwarder"):
        fnew = st.text_input("포워더명")
        if st.form_submit_button("➕ 추가"):
            if not fnew: st.error("포워더명을 입력하세요.")
            else:
                cfg.setdefault("forwarders", []).append({
                    "id":str(uuid.uuid4())[:8], "name":fnew, "active":True, "quotes":[]})
                save_cfg(cfg); st.success(f"{fnew} 추가!"); st.rerun()

    # ── 목적지별 수출비 비교표 ──────────────────────────────────────────────────
    st.divider()
    st.markdown("#### 🗺️ 목적지별 수출비 비교")
    _eur = cfg.get("freight", {}).get("eur_usd", 1.08)
    _cmp = []
    for _fwd in cfg.get("forwarders", []):
        if not _fwd.get("active", True): continue
        for _q in _fwd.get("quotes", []):
            _curr = _q.get("currency", "USD")
            _tot  = (_q.get("pickup_rail",0)+_q.get("ocean_freight",0)+_q.get("dg_surcharge",0)
                     +_q.get("fuel_surcharge",0)+_q.get("documentation",0)+_q.get("terminal_handling",0))
            _tusd = _tot * _eur if _curr=="EUR" else _tot
            _cap  = _q.get("capacity_kg", 1) or 1
            _cmp.append({
                "목적지":   _q.get("destination","").strip() or "(미지정)",
                "포워더":   _fwd["name"],
                "견적명":   _q.get("label","—"),
                "컨테이너": _q.get("container_type","?").upper(),
                "용량(kg)": int(_cap),
                "통화":     _curr,
                "수출비합계": round(_tot, 0),
                "USD 환산": round(_tusd, 2),
                "$/kg":     round(_tusd / _cap, 4),
            })

    if _cmp:
        df_cmp_all = pd.DataFrame(_cmp).sort_values(["목적지","컨테이너","$/kg"]).reset_index(drop=True)

        # 컨테이너 필터
        _cntr_types = ["전체"] + sorted(df_cmp_all["컨테이너"].unique().tolist())
        _sel_cntr = st.radio("컨테이너 필터", _cntr_types, horizontal=True, key="cmp_cntr_flt")
        df_cmp = df_cmp_all if _sel_cntr=="전체" else df_cmp_all[df_cmp_all["컨테이너"]==_sel_cntr].reset_index(drop=True)

        _min_dest = df_cmp.groupby("목적지")["$/kg"].min()

        def _hl_best(row):
            if row["$/kg"] == _min_dest.get(row["목적지"]):
                return ["background-color:#1E8449;color:white;font-weight:600"]*len(row)
            return [""]*len(row)

        st.dataframe(
            df_cmp.style.apply(_hl_best, axis=1).format({
                "용량(kg)":"{:,.0f}", "수출비합계":"{:,.0f}",
                "USD 환산":"${:,.2f}", "$/kg":"${:.4f}"}),
            use_container_width=True, hide_index=True)
        st.caption("🟢 목적지별 최저 수출비")

        # 목적지 × 포워더 피벗 매트릭스
        st.markdown("##### 목적지 × 포워더  $/kg 매트릭스")
        try:
            _pivot = df_cmp.pivot_table(
                index="목적지", columns="포워더", values="$/kg", aggfunc="min").round(4)

            def _hl_row_min(row):
                mn = row.min()
                return ["background-color:#1E8449;color:white;font-weight:600" if v==mn else "" for v in row]

            st.dataframe(
                _pivot.style.apply(_hl_row_min, axis=1).format("${:.4f}", na_rep="—"),
                use_container_width=True)
            st.caption("셀 값: 해당 포워더·목적지의 최저 $/kg  |  🟢 행 최저값")
        except Exception:
            st.info("목적지 및 포워더가 2개 이상 등록되면 매트릭스가 표시됩니다.")
    else:
        st.info("견적을 등록하면 비교표가 표시됩니다.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — 손익 분석
# ══════════════════════════════════════════════════════════════════════════════
with t_pnl:
    st.subheader("손익 분석")

    # ③ EUR/USD 공백 경고: 출고·처리 이력 월 중 최신 등록 환율보다 이후 월 검출
    _eur_rates_sorted = sorted(cfg.get("eur_usd_rates", []), key=lambda x: x["month"])
    if _eur_rates_sorted:
        _eur_latest = _eur_rates_sorted[-1]["month"]
        _dispatch_months = {
            (dr.get("date") or "")[:7]
            for dr in cfg.get("dispatch_records", [])
            if (dr.get("date") or "")[:7] > _eur_latest
        }
        _ph_months = {
            _rec_ref_date(r, cfg)[:7]
            for r in cfg.get("processing_history", [])
            if _rec_ref_date(r, cfg) and _rec_ref_date(r, cfg)[:7] > _eur_latest
        }
        _gap_months = sorted(_dispatch_months | _ph_months)
        if _gap_months:
            st.warning(
                f"⚠️ **EUR/USD 환율 미등록 월**: {', '.join(_gap_months)}  \n"
                f"해당 월의 창고비(보관비) 계산에 {_eur_latest} 환율({_eur_rates_sorted[-1]['rate']:.4f})이 적용됩니다.  \n"
                "정확한 계산을 위해 **INDEX 이력 탭**에서 해당 월 EUR/USD를 등록하세요."
            )

    _ph_all   = cfg.get("processing_history", [])
    _bm_pnl   = {b["id"]: b for b in cfg["buyers"]}
    _sm_pnl   = {s["id"]: s for s in cfg.get("shipments", [])}
    _pm_pnl   = {p["id"]: p for p in cfg.get("processors", [])}
    _scm_pnl  = {s["id"]: s for s in cfg.get("scrap_types", [])}

    # ── FIFO 원가 + 창고비 사전 계산 (섹션 전체 공용) ────────────────────────
    # (scrap_type_id, shipment_id) → FIFO 가중평균 단가 / FIFO 자동 창고비
    # 조건: 기초재고 있고 dispatch_records 있는 스크랩 유형만
    _fifo_rmc_map     = {}   # (sc_id, ship_id) → $/kg 스크랩
    _fifo_storage_map = {}   # (sc_id, ship_id) → USD 창고비
    _fifo_sc_ids_avail = {
        r.get("scrap_type_id","") for r in _ph_all
        if r.get("scrap_type_id")
        and cfg.get("raw_material_inventory",{}).get(r.get("scrap_type_id",""),{}).get("opening")
        and any(dr.get("scrap_type_id")==r.get("scrap_type_id","")
                for dr in cfg.get("dispatch_records",[]))
    }
    for _fsc in _fifo_sc_ids_avail:
        _f_bl, _, _ = _fifo_lot_trace(cfg, _fsc)
        for _f_sid, _f_data in _f_bl.items():
            _f_lots = _f_data.get("lots", {})
            _f_known_qty = sum(v["qty"] for v in _f_lots.values()
                               if v.get("unit_cost") is not None)
            _f_known_amt = sum(v["amount"] for v in _f_lots.values()
                               if v.get("unit_cost") is not None)
            if _f_known_qty > 0:
                _fifo_rmc_map[(_fsc, _f_sid)] = _f_known_amt / _f_known_qty
            _fifo_storage_map[(_fsc, _f_sid)] = _f_data.get("storage_cost", 0.0)

    # 배치 입고량 합계 (동일 scrap+shipment 내 비례 배분용)
    _batch_inp_total = defaultdict(float)
    for _br in _ph_all:
        _batch_inp_total[(_br.get("scrap_type_id",""), _br.get("shipment_id",""))] += _ph_input_kg(_br)

    def _auto_storage_for_batch(rec):
        """FIFO 자동 창고비 — 수동 storage_days 미입력 배치에만 fallback으로 사용.
        동일 (scrap_type, shipment) 내 배치별 입고량 비례 배분.
        """
        _sc   = rec.get("scrap_type_id","")
        _sid  = rec.get("shipment_id","")
        total = _fifo_storage_map.get((_sc, _sid), 0.0)
        if total <= 0:
            return 0.0
        inp_total = _batch_inp_total.get((_sc, _sid), 0.0)
        inp_r     = _ph_input_kg(rec)
        return round(total * (inp_r / inp_total), 4) if inp_total > 0 else 0.0

    def _get_rmc_fifo(rec, default_rmc=0.0):
        """FIFO 우선 원료단가 결정. 반환: (unit_cost, source_label)"""
        _sc  = rec.get("scrap_type_id","")
        _sid = rec.get("shipment_id","") or f"__no_ship__{rec.get('id','')}"
        fifo = _fifo_rmc_map.get((_sc, _sid))
        if fifo is not None:
            return fifo, "FIFO"
        stored = rec.get("raw_material_cost_per_kg")
        if stored is not None:
            return float(stored), "수동"
        _as_date = _rec_ref_date(rec, cfg)
        _avg, _ = _inv_moving_avg(cfg, _sc, _as_date)
        if _avg is not None:
            return _avg, "이동평균"
        if default_rmc > 0:
            return default_rmc, "기본값"
        return 0.0, "—"

    def _get_rmc_mavg(rec, default_rmc=0.0):
        """이동평균 우선 원료단가 결정. 반환: (unit_cost, source_label)"""
        _sc  = rec.get("scrap_type_id","")
        _as_date = _rec_ref_date(rec, cfg)
        _avg, _ = _inv_moving_avg(cfg, _sc, _as_date)
        if _avg is not None:
            return _avg, "이동평균"
        stored = rec.get("raw_material_cost_per_kg")
        if stored is not None:
            return float(stored), "수동"
        _sid = rec.get("shipment_id","") or f"__no_ship__{rec.get('id','')}"
        fifo = _fifo_rmc_map.get((_sc, _sid))
        if fifo is not None:
            return fifo, "FIFO"
        if default_rmc > 0:
            return default_rmc, "기본값"
        return 0.0, "—"

    # ══════════════════════════════════════════════════════════════════════════
    # 섹션 1 — 핵심 실적 지표
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### 💰 핵심 실적 지표")
    if not _ph_all:
        st.info("처리 이력(임가공사 관리 > 세부 내역)을 입력하면 실적 기반 손익이 표시됩니다.")
    else:
        _tot_out  = sum(r.get("output_kg",0) or 0 for r in _ph_all)
        _tot_inp  = sum(_ph_input_kg(r) for r in _ph_all)
        # 전체 흐름 집계 (스크랩 매각수익 · BP 재매입원가 모두 기재)
        _tot_sc_rev  = sum((r.get("scrap_sale_per_kg",0) or 0) * _ph_input_kg(r) for r in _ph_all)  # 스크랩 매각수익
        _tot_pf      = sum((r.get("processing_fee_per_kg",0) or 0) * _ph_input_kg(r) for r in _ph_all)  # 임가공비
        _tot_repr    = _tot_sc_rev + _tot_pf   # BP 재매입 원가 합계 (= 스크랩 + 임가공비)
        _tot_eu      = sum(_ph_export_usd(r, cfg) for r in _ph_all)
        _tot_bp      = sum((r.get("bp_sale_per_kg",0) or 0) * (r.get("output_kg",0) or 0) for r in _ph_all)
        # 순이익 = BP매각 + 스크랩매각 - BP재매입 - 수출비  (스크랩 상계 → 임가공비+수출비만 남음)
        _tot_net  = _tot_bp + _tot_sc_rev - _tot_repr - _tot_eu
        _avg_mg   = _tot_net / _tot_bp * 100 if _tot_bp > 0 else 0
        _repr_per_kg = _tot_repr / _tot_out if _tot_out > 0 else 0
        _pf_per_kg   = _tot_pf   / _tot_out if _tot_out > 0 else 0
        _epk_all     = _tot_eu   / _tot_out if _tot_out > 0 else 0

        _km1,_km2,_km3,_km4 = st.columns(4)
        _km1.metric("총 BP 생산",     f"{_tot_out/1000:,.2f} ton")
        _km2.metric("BP 매각 수익",   f"${_tot_bp:,.2f}")
        _km3.metric("BP 재매입 원가", f"${_tot_repr:,.2f}",
                    help="스크랩 매각단가 × 투입량 + 임가공비 × 투입량")
        _km4.metric("거래 마진",       f"${_tot_net:+,.2f}",
                    delta_color="normal" if _tot_net >= 0 else "inverse")

        # 원가 구성 상세
        st.markdown("---")
        st.caption("💡 스크랩 매각수익과 BP 재매입원가의 스크랩 부분은 상계 — 순 차감원가 = 임가공비 + 수출비")
        _kd1,_kd2,_kd3,_kd4 = st.columns(4)
        _kd1.metric("스크랩 매각수익",  f"${_tot_sc_rev:,.2f}",
                    delta=f"${_tot_sc_rev/_tot_out:.4f}/kg BP" if _tot_out>0 else None)
        _kd2.metric("BP 재매입 원가",   f"${_tot_repr:,.2f}",
                    delta=f"${_repr_per_kg:.4f}/kg BP" if _tot_out>0 else None)
        _kd3.metric("임가공비 (순)",    f"${_tot_pf:,.2f}",
                    delta=f"${_pf_per_kg:.4f}/kg BP" if _tot_out>0 else None,
                    help="BP 재매입 원가에서 스크랩 매각수익 차감한 순 임가공 비용")
        _kd4.metric("수출비",           f"${_tot_eu:,.2f}",
                    delta=f"${_epk_all:.4f}/kg BP" if _tot_out>0 else None)

        # P&L Waterfall 차트 (전체 흐름 표시)
        try:
            import plotly.graph_objects as go
            _net_color = "#1E8449" if _tot_net >= 0 else "#922b21"
            _wf = go.Figure(go.Waterfall(
                orientation="v",
                measure=["absolute", "relative", "relative", "relative", "total"],
                x=["BP 매각수익", "스크랩 매각수익", "BP 재매입원가", "수출비", "거래 마진"],
                y=[_tot_bp, _tot_sc_rev, -_tot_repr, -_tot_eu, 0],
                text=[f"${_tot_bp:,.0f}", f"+${_tot_sc_rev:,.0f}", f"-${_tot_repr:,.0f}",
                      f"-${_tot_eu:,.0f}", f"${_tot_net:+,.0f}"],
                textposition="outside",
                increasing=dict(marker=dict(color="#2E75B6")),
                decreasing=dict(marker=dict(color="#E74C3C")),
                totals=dict(marker=dict(color=_net_color)),
                connector=dict(line=dict(color="#555577", width=1, dash="dot")),
                hovertemplate="%{x}<br>$%{y:+,.2f}<extra></extra>",
            ))
            _wf.update_layout(
                title=dict(text=f"매출 ${_tot_bp:,.0f}  →  거래 마진 ${_tot_net:+,.0f}",
                           font=dict(size=13)),
                height=340, margin=dict(l=10, r=10, t=50, b=10),
                yaxis=dict(tickformat="$,.0f", gridcolor="rgba(255,255,255,0.08)"),
                plot_bgcolor="#1E1E2E", paper_bgcolor="#16213E",
                font=dict(color="#D0D0E8"),
                showlegend=False,
            )
            st.plotly_chart(_wf, use_container_width=True)
        except ImportError:
            pass

        # ── BP 1kg당 단가 분해 ────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("**📐 BP 1kg당 단가 분해**")
        st.caption("전체 처리 이력 합산 기준. 원료 취득원가는 FIFO·이동평균 두 가지 기준 비교.")
        _default_rmc_s1 = float(st.session_state.get("pnl_rmc_default", 0.0))
        if _tot_out > 0:
            _raw_fifo_s1 = sum(_get_rmc_fifo(r, _default_rmc_s1)[0] * _ph_input_kg(r) for r in _ph_all)
            _raw_mavg_s1 = sum(_get_rmc_mavg(r, _default_rmc_s1)[0] * _ph_input_kg(r) for r in _ph_all)
            _stor_s1     = sum((_ph_storage_cost(r, cfg) or _auto_storage_for_batch(r)) for r in _ph_all)

            _bp_pk   = _tot_bp  / _tot_out
            _pf_pk   = _tot_pf  / _tot_out
            _eu_pk   = _tot_eu  / _tot_out
            _gm_pk   = _bp_pk - _pf_pk - _eu_pk
            _rf_pk   = _raw_fifo_s1 / _tot_out
            _rm_pk   = _raw_mavg_s1 / _tot_out
            _st_pk   = _stor_s1     / _tot_out
            _rlf_pk  = _gm_pk - _rf_pk - _st_pk   # 실질손익 FIFO
            _rlm_pk  = _gm_pk - _rm_pk - _st_pk   # 실질손익 이동평균

            _bpkg_rows = [
                ("BP 매각가",           _bp_pk,  _bp_pk,  0.0),
                ("  (−) 임가공비",      _pf_pk,  _pf_pk,  0.0),
                ("  (−) 수출비",        _eu_pk,  _eu_pk,  0.0),
                ("= 거래 마진",         _gm_pk,  _gm_pk,  0.0),
                ("  (−) 원료 취득원가", _rf_pk,  _rm_pk,  _rf_pk - _rm_pk),
                ("  (−) 보관비",        _st_pk,  _st_pk,  0.0),
                ("= 실질 손익",         _rlf_pk, _rlm_pk, _rlf_pk - _rlm_pk),
            ]
            _df_bpkg = pd.DataFrame(
                _bpkg_rows,
                columns=["항목", "FIFO 기준 ($/kg BP)", "이동평균 기준 ($/kg BP)", "차이"],
            )

            def _hl_bpkg(row):
                lbl  = row["항목"]
                fv   = row["FIFO 기준 ($/kg BP)"]
                mv   = row["이동평균 기준 ($/kg BP)"]
                diff = row["차이"]
                if "실질 손익" in lbl:
                    bg = "#1E8449" if fv >= 0 else "#922b21"
                    s  = f"background-color:{bg};color:white;font-weight:700"
                    return ["font-weight:700", s, s, "font-weight:700"]
                if lbl.startswith("="):
                    return ["font-weight:700", "font-weight:700", "font-weight:700", "font-weight:700"]
                if "BP 매각가" in lbl:
                    return ["font-weight:700",
                            "color:#1565c0;font-weight:600",
                            "color:#1565c0;font-weight:600", ""]
                if "원료" in lbl and abs(diff) > 0.0001:
                    # 낮은 값(= 유리한 쪽)을 녹색으로 표시
                    fc = "color:#1E8449;font-weight:600" if fv <= mv else "color:#E74C3C;font-weight:600"
                    mc = "color:#1E8449;font-weight:600" if mv <= fv else "color:#E74C3C;font-weight:600"
                    return ["color:#555", fc, mc, "font-weight:600"]
                return ["color:#555", "", "", ""]

            def _fmt_pk(v):
                return f"${v:+.4f}" if v is not None else "—"

            def _fmt_diff(v):
                return f"${v:+.4f}" if v is not None and abs(v) > 0.00001 else "—"

            st.dataframe(
                _df_bpkg.style.apply(_hl_bpkg, axis=1).format({
                    "FIFO 기준 ($/kg BP)":    _fmt_pk,
                    "이동평균 기준 ($/kg BP)": _fmt_pk,
                    "차이":                    _fmt_diff,
                }),
                use_container_width=True, hide_index=True,
            )
            _conv_avg = _tot_out / _tot_inp * 100 if _tot_inp > 0 else 0
            _rmc_gap  = abs(_rf_pk - _rm_pk)
            _gap_note = ""
            if _rmc_gap > 0.0001:
                _cheaper = "FIFO" if _rf_pk < _rm_pk else "이동평균"
                _gap_note = f" | 원료단가 차이 **${_rmc_gap:.4f}/kg BP** → {_cheaper}가 유리"
            st.caption(
                f"전환율 가중평균 **{_conv_avg:.1f}%** (스크랩 투입 → BP 생산)"
                + _gap_note
                + "  \n원료단가 기본값 변경·FIFO↔이동평균 전환은 아래 HBL별 손익 요약 섹션에서 가능합니다."
            )

        # ── 월별 손익 집계 ────────────────────────────────────────────────────
        _mon_agg2 = defaultdict(lambda: {"bp":0,"sc":0,"repr":0,"eu":0,"raw":0,"stor":0,"cnt":0,"out":0})
        for _mr in _ph_all:
            _msid  = _mr.get("shipment_id","")
            _ms    = _sm_pnl.get(_msid, {})
            _mld   = (_ms.get("loading_date") or "")[:7] or "미상"
            _mo    = _mr.get("output_kg",0) or 0
            _mi    = _ph_input_kg(_mr)
            _msc   = (_mr.get("scrap_sale_per_kg",0) or 0) * _mi
            _mpf   = (_mr.get("processing_fee_per_kg",0) or 0) * _mi
            _meu   = _ph_export_usd(_mr, cfg)
            _mbp   = (_mr.get("bp_sale_per_kg",0) or 0) * _mo
            _mrc, _ = _get_rmc_fifo(_mr, _default_rmc_s1)
            # 보관비: 수동 storage_days 우선, 없으면 FIFO 자동 fallback
            _mstor = _ph_storage_cost(_mr, cfg) or _auto_storage_for_batch(_mr)
            _mon_agg2[_mld]["bp"]   += _mbp
            _mon_agg2[_mld]["sc"]   += _msc
            _mon_agg2[_mld]["repr"] += _msc + _mpf
            _mon_agg2[_mld]["eu"]   += _meu
            _mon_agg2[_mld]["raw"]  += _mrc * _mi
            _mon_agg2[_mld]["stor"] += _mstor
            _mon_agg2[_mld]["cnt"]  += 1
            _mon_agg2[_mld]["out"]  += _mo
        if _mon_agg2:
            with st.expander("📅 월별 손익 집계", expanded=False):
                _mon_rows2 = []
                for _mkey in sorted(_mon_agg2.keys()):
                    _mv2  = _mon_agg2[_mkey]
                    _mn2  = _mv2["bp"] + _mv2["sc"] - _mv2["repr"] - _mv2["eu"]
                    _mr2  = _mn2 - _mv2["raw"] - _mv2["stor"]
                    _mon_rows2.append({
                        "월":       _mkey,
                        "HBL수":    _mv2["cnt"],
                        "생산(kg)": round(_mv2["out"], 0),
                        "BP 매각":  round(_mv2["bp"], 2),
                        "BP 재매입": round(_mv2["repr"], 2),
                        "수출비":   round(_mv2["eu"], 2),
                        "거래 마진": round(_mn2, 2),
                        "실질 손익": round(_mr2, 2),
                    })
                _df_mon2 = pd.DataFrame(_mon_rows2)
                def _hl_mon2(row):
                    styles = [""] * len(row)
                    _cols2 = list(row.index)
                    v_net  = row.get("거래 마진", 0) or 0
                    v_real = row.get("실질 손익", 0) or 0
                    c_net  = ("color:#155724;font-weight:600" if v_net  >= 0 else "color:#721c24;font-weight:600")
                    c_real = ("color:#155724;font-weight:600" if v_real >= 0 else "color:#721c24;font-weight:600")
                    if "거래 마진"  in _cols2: styles[_cols2.index("거래 마진")]  = c_net
                    if "실질 손익" in _cols2: styles[_cols2.index("실질 손익")] = c_real
                    return styles
                st.dataframe(_df_mon2.style.apply(_hl_mon2, axis=1).format({
                    "생산(kg)":  "{:,.0f}",
                    "BP 매각":   "${:,.2f}",
                    "BP 재매입": "${:,.2f}",
                    "수출비":    "${:,.2f}",
                    "거래 마진":  "${:+,.2f}",
                    "실질 손익": "${:+,.2f}",
                }), use_container_width=True, hide_index=True)
                st.caption("실질 손익 = 거래 마진 − 원료비(FIFO 우선) − 보관비  "
                           "| 원료단가 수동 기본값은 아래 '실질 손익 분석' 섹션에서 조정 가능합니다.")

    st.divider()

    # 보관비 헬퍼 (섹션 2·3 공용)
    def _eff_storage(rec):
        manual = _ph_storage_cost(rec, cfg)
        return manual if manual else _auto_storage_for_batch(rec)

    # ══════════════════════════════════════════════════════════════════════════
    # 섹션 2 — HBL별 손익 요약
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### 🚢 HBL별 손익 요약")

    # 기본값 — linked 배치 없을 때 섹션 3에서 참조
    _rmc_mode    = "FIFO 우선"
    _default_rmc = 0.0

    _linked_pnl = [r for r in _ph_all if r.get("shipment_id","")]
    if not _linked_pnl:
        st.info("세부 내역(임가공사 관리)에서 배치를 HBL과 연결하면 선적 단위 손익이 표시됩니다.")
    else:
        # ── 컨트롤 ──────────────────────────────────────────────────────────
        _hc1, _hc2, _hc3 = st.columns([3, 2, 3])
        with _hc1:
            _rmc_mode = st.radio(
                "원료단가 기준",
                ["FIFO 우선", "이동평균 우선"],
                horizontal=True,
                key="rmc_mode_toggle",
                help="FIFO: 출고기록 Lot 원가 → 수동 → 이동평균  |  이동평균: 선적일 기준 → 수동 → FIFO",
            )
        with _hc2:
            _default_rmc = st.number_input(
                "원료단가 기본값 ($/kg)",
                value=0.0, step=0.001, format="%.4f", key="pnl_rmc_default",
                help="이동평균·FIFO 없는 배치에 적용. 0 = 미반영.",
            )

        # ── HBL별 집계 ──────────────────────────────────────────────────────
        _hbl_agg = {}
        for _rr in _linked_pnl:
            _hid    = _rr["shipment_id"]
            _hs     = _sm_pnl.get(_hid, {})
            _hb     = _bm_pnl.get(_hs.get("buyer_id",""), {})
            _op_r   = float(_rr.get("output_kg") or 0)
            _ip_r   = _ph_input_kg(_rr)
            _pf_r   = float(_rr.get("processing_fee_per_kg") or 0) * _ip_r
            _sc_r   = float(_rr.get("scrap_sale_per_kg") or 0) * _ip_r
            _eu_r   = _ph_export_usd(_rr, cfg)
            _bp_r   = float(_rr.get("bp_sale_per_kg") or 0) * _op_r
            # 거래 마진: BP매각 + 스크랩매각수익 − 재매입원가(sc+pf) − 수출비
            # sc 상계 후 → bp − pf − eu (임가공비와 수출비만 남음)
            _trade_r  = _bp_r - _pf_r - _eu_r
            if _rmc_mode == "이동평균 우선":
                _ermc_r, _rmc_src = _get_rmc_mavg(_rr, _default_rmc)
            else:
                _ermc_r, _rmc_src = _get_rmc_fifo(_rr, _default_rmc)
            _raw_r    = _ermc_r * _ip_r
            _stor_r   = _eff_storage(_rr)
            _stor_src = "수동" if _ph_storage_cost(_rr, cfg) else ("FIFO 자동" if _stor_r else "—")
            _real_r   = _trade_r - _raw_r - _stor_r
            _proc_nm  = _pm_pnl.get(_rr.get("processor_id",""), {}).get("name","—")
            _sc_nm    = _scm_pnl.get(_rr.get("scrap_type_id",""), {}).get("name","—")

            if _hid not in _hbl_agg:
                _hbl_agg[_hid] = {
                    "hbl":       _hs.get("hbl","—"),
                    "load_date": _hs.get("loading_date","—"),
                    "매입사":    f"{_hb.get('name','?')} ({_hb.get('product','?')})",
                    "procs":     set(),
                    "scrap_ids": set(),   # FIFO Lot 조회용
                    "out":  0.0, "bp_rev": 0.0, "pf": 0.0,
                    "eu":   0.0, "raw":    0.0,  "stor": 0.0,
                    "batches": [],
                }
            _h = _hbl_agg[_hid]
            _h["out"]    += _op_r
            _h["bp_rev"] += _bp_r;  _h["pf"]  += _pf_r
            _h["eu"]     += _eu_r;  _h["raw"] += _raw_r
            _h["stor"]   += _stor_r
            if _proc_nm != "—": _h["procs"].add(_proc_nm)
            if _rr.get("scrap_type_id"): _h["scrap_ids"].add(_rr["scrap_type_id"])
            _h["batches"].append({
                "임가공사":   _proc_nm,
                "스크랩":     _sc_nm,
                "BP(kg)":     round(_op_r, 0),
                "BP매각":     round(_bp_r, 2),
                "임가공비":   round(_pf_r, 2),
                "수출비":     round(_eu_r, 2),
                "원료비":     round(_raw_r, 2),
                "원가기준":   _rmc_src,
                "보관비":     round(_stor_r, 2) if _stor_r else None,
                "실질손익":   round(_real_r, 2),
            })

        # ── 요약 테이블 ──────────────────────────────────────────────────────
        _sum_rows = []
        for _hid, _h in sorted(_hbl_agg.items(),
                                key=lambda x: x[1]["load_date"], reverse=True):
            _htrade = _h["bp_rev"] - _h["pf"] - _h["eu"]
            _hreal  = _htrade - _h["raw"] - _h["stor"]
            _hmgr   = _hreal / _h["bp_rev"] * 100 if _h["bp_rev"] > 0 else None
            _sum_rows.append({
                "HBL":        _h["hbl"],
                "선적일":     _h["load_date"],
                "매입사":     _h["매입사"],
                "임가공사":   "·".join(sorted(_h["procs"])) or "—",
                "BP생산(kg)": round(_h["out"], 0),
                "거래 마진":  round(_htrade, 2),
                "원료비":     round(_h["raw"], 2),
                "보관비":     round(_h["stor"], 2) if _h["stor"] else None,
                "실질 손익":  round(_hreal, 2),
                "마진율(%)":  round(_hmgr, 2) if _hmgr is not None else None,
            })

        def _hl_sum_tbl(row):
            styles = [""] * len(row)
            _ci = list(row.index)
            v = row.get("실질 손익", 0) or 0
            if "실질 손익" in _ci:
                styles[_ci.index("실질 손익")] = (
                    "background-color:#1E8449;color:white;font-weight:600" if v >= 0
                    else "background-color:#922b21;color:white;font-weight:600")
            return styles

        st.dataframe(
            pd.DataFrame(_sum_rows).style.apply(_hl_sum_tbl, axis=1).format(na_rep="—", formatter={
                "BP생산(kg)": "{:,.0f}",
                "거래 마진":  "${:+,.2f}",
                "원료비":     "${:,.2f}",
                "보관비":     lambda v: f"${v:,.2f}" if v else "—",
                "실질 손익":  "${:+,.2f}",
                "마진율(%)":  lambda v: f"{v:+.2f}%" if v is not None else "—",
            }),
            use_container_width=True, hide_index=True,
        )

        # Excel 다운로드
        try:
            _xl_buf = BytesIO()
            with pd.ExcelWriter(_xl_buf, engine="openpyxl") as _xew:
                pd.DataFrame(_sum_rows).to_excel(_xew, index=False, sheet_name="HBL손익요약")
            _xl_buf.seek(0)
            st.download_button(
                "📥 Excel 다운로드",
                data=_xl_buf,
                file_name=f"HBL_손익요약_{date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception:
            pass

        # ── HBL별 카드 ──────────────────────────────────────────────────────
        st.markdown("---")
        for _hid, _h in sorted(_hbl_agg.items(),
                                key=lambda x: x[1]["load_date"], reverse=True):
            _htrade = _h["bp_rev"] - _h["pf"] - _h["eu"]
            _hreal  = _htrade - _h["raw"] - _h["stor"]
            _hmgr   = _hreal / _h["bp_rev"] * 100 if _h["bp_rev"] > 0 else 0
            _icon   = "🟢" if _hreal >= 0 else "🔴"
            with st.expander(
                f"{_icon}  {_h['hbl']}  ·  {_h['매입사']}  ·  {_h['load_date']}"
                f"  |  실질손익 ${_hreal:+,.0f}  ({_hmgr:+.1f}%)",
                expanded=False,
            ):
                _cl, _cr = st.columns([5, 7])
                with _cl:
                    st.markdown("**손익 분해**")
                    _wf_df = pd.DataFrame([
                        ("BP 매각",              _h["bp_rev"]),
                        ("  (−) 임가공비",       -_h["pf"]),
                        ("  (−) 수출비",         -_h["eu"]),
                        ("= 거래 마진",           _htrade),
                        ("  (−) 원료 취득원가",  -_h["raw"]),
                        ("  (−) 보관비",         -_h["stor"]),
                        ("= 실질 손익",           _hreal),
                    ], columns=["항목", "금액 (USD)"])

                    def _hl_wf(row):
                        lbl = row["항목"]
                        v   = row["금액 (USD)"]
                        if "실질 손익" in lbl:
                            bg = "#1E8449" if v >= 0 else "#922b21"
                            return ["font-weight:700",
                                    f"background-color:{bg};color:white;font-weight:700"]
                        if lbl.startswith("="):
                            return ["font-weight:700", "font-weight:700"]
                        return ["color:#666", ""]

                    st.dataframe(
                        _wf_df.style.apply(_hl_wf, axis=1)
                              .format({"금액 (USD)": "${:+,.2f}"}),
                        use_container_width=True, hide_index=True, height=280,
                    )
                    _mk1, _mk2 = st.columns(2)
                    _mk1.metric("BP 생산", f"{_h['out']/1000:.2f} ton")
                    _mk2.metric("수익률",  f"{_hmgr:+.1f}%",
                                delta_color="normal" if _hreal >= 0 else "inverse")
                with _cr:
                    st.markdown("**배치 상세**")
                    _df_bat = pd.DataFrame(_h["batches"])

                    def _hl_bat(row):
                        styles = [""] * len(row)
                        _bc = list(row.index)
                        v   = row.get("실질손익", 0) or 0
                        src = row.get("원가기준", "")
                        if "실질손익" in _bc:
                            styles[_bc.index("실질손익")] = (
                                "color:#1E8449;font-weight:600" if v >= 0
                                else "color:#922b21;font-weight:600")
                        if "원가기준" in _bc:
                            styles[_bc.index("원가기준")] = (
                                "color:#1565c0;font-weight:600" if src == "FIFO"
                                else "color:#6c757d")
                        return styles

                    st.dataframe(
                        _df_bat.style.apply(_hl_bat, axis=1).format(na_rep="—", formatter={
                            "BP(kg)":   "{:,.0f}",
                            "BP매각":   "${:,.2f}",
                            "임가공비": "${:,.2f}",
                            "수출비":   "${:,.2f}",
                            "원료비":   "${:,.2f}",
                            "보관비":   lambda v: f"${v:,.2f}" if v else "—",
                            "실질손익": "${:+,.2f}",
                        }),
                        use_container_width=True, hide_index=True,
                    )

                    # ── 원료 Lot 구성 ──────────────────────────────────────
                    st.markdown("**📦 원료 Lot 구성 (FIFO)**")
                    _lot_any = False
                    for _lot_scid in sorted(_h.get("scrap_ids", set())):
                        _lot_sc_nm = _scm_pnl.get(_lot_scid, {}).get("name", "?")
                        try:
                            _lot_bl, _, _ = _fifo_lot_trace(cfg, _lot_scid)
                            _lot_data = _lot_bl.get(_hid, {}).get("lots", {})
                            if not _lot_data:
                                continue
                            _lot_total_qty = sum(v["qty"] for v in _lot_data.values())
                            _lot_rows = []
                            for _lot_lbl, _lv in sorted(
                                _lot_data.items(),
                                key=lambda x: x[1].get("lot_date") or ""
                            ):
                                _lot_qty  = _lv["qty"]
                                _lot_pct  = _lot_qty / _lot_total_qty * 100 if _lot_total_qty else 0
                                _lot_date = (_lv.get("lot_date") or "")[:7] or ("기초재고" if "기초재고" in _lot_lbl else "—")
                                _lot_uc   = _lv.get("unit_cost")
                                _lot_rows.append({
                                    "입고월":     _lot_date,
                                    "Lot":        _lot_lbl,
                                    "소진량(kg)": round(_lot_qty, 0),
                                    "비중(%)":    round(_lot_pct, 1),
                                    "원료단가":   _lot_uc,
                                })
                            if _lot_rows:
                                _lot_any = True
                                if len(_h.get("scrap_ids", set())) > 1:
                                    st.caption(f"**{_lot_sc_nm}**")
                                st.dataframe(
                                    pd.DataFrame(_lot_rows).style.format(na_rep="—", formatter={
                                        "소진량(kg)": "{:,.0f}",
                                        "비중(%)":    "{:.1f}%",
                                        "원료단가":   lambda v: f"${v:.5f}" if v is not None else "—",
                                    }),
                                    use_container_width=True, hide_index=True,
                                )
                        except Exception:
                            pass
                    if not _lot_any:
                        st.caption("FIFO Lot 정보 없음 (입출고 기록 탭에서 출고 기록 입력 필요)")

        _tot_real_hbl = sum(
            (_h["bp_rev"] - _h["pf"] - _h["eu"]) - _h["raw"] - _h["stor"]
            for _h in _hbl_agg.values()
        )
        _fm1, _fm2 = st.columns(2)
        _fm1.metric("전체 HBL 합산 실질 손익", f"${_tot_real_hbl:+,.2f}")
        _fm2.metric("미연결 배치", f"{len(_ph_all)-len(_linked_pnl)}건",
                    help="임가공사 관리 > 세부 내역에서 HBL을 연결하세요.")

    st.divider()

    # ══════════════════════════════════════════════════════════════════════════
    # 섹션 3 — 실질 손익 분석 (전체)
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### 🔬 실질 손익 분석 (전체)")
    st.caption("전체 배치의 거래 마진에서 원료 취득원가·판관비를 차감한 실질 수익성입니다. "
               "원료단가 기준은 위 HBL 요약과 동기화됩니다.")

    if not _ph_all:
        st.info("처리 이력을 입력하면 실질 손익 분석이 가능합니다.")
    else:
        # ── 원료 원가 기준 안내 ──
        _inv_avail_pnl = {}
        for _isc_p in cfg.get("scrap_types", []):
            _avg_p, _ = _inv_moving_avg(cfg, _isc_p["id"])
            if _avg_p is not None:
                _inv_avail_pnl[_isc_p["name"]] = (_isc_p["id"], _avg_p)
        _has_dispatch_any = bool(cfg.get("dispatch_records", []))
        if _has_dispatch_any:
            st.info("📦 원료단가 적용 우선순위: **FIFO** (출고 기록 탭) → 수동 입력 → 이동평균 → 수동 기본값  \n"
                    "출고 기록이 완전히 입력된 B/L은 FIFO 원가가 자동 적용됩니다.")
        elif _inv_avail_pnl:
            _avg_parts = [f"**{nm}** ${avg:.5f}/kg" for nm, (_, avg) in _inv_avail_pnl.items()]
            st.info("📊 이동평균 자동 적용 중 — " + " | ".join(_avg_parts)
                    + "  _(출고 기록 탭에서 임가공 출고를 입력하면 FIFO 원가로 자동 전환됩니다)_")

        _rl2, _rl4 = st.columns(2)
        with _rl2:
            _sga_total = st.number_input(
                "간접 판관비 (USD, 기간 합계)",
                value=0.0, step=100.0, format="%.2f", key="pnl_sga",
                help="인건비·임차료·감가상각 등 배치에 직접 귀속하기 어려운 고정비 총액",
            )
        with _rl4:
            _other_cost = st.number_input(
                "기타 원가 (USD)",
                value=0.0, step=100.0, format="%.2f", key="pnl_other",
                help="위 항목 외 추가 반영할 비용",
            )

        # 배치별 원료비 집계 (_rmc_mode·_default_rmc 는 섹션 2 컨트롤과 동기화)
        if _rmc_mode == "이동평균 우선":
            _tot_raw = sum(_get_rmc_mavg(r, _default_rmc)[0] * _ph_input_kg(r) for r in _ph_all)
        else:
            _tot_raw = sum(_get_rmc_fifo(r, _default_rmc)[0] * _ph_input_kg(r) for r in _ph_all)
        _tot_storage = sum(_eff_storage(r) for r in _ph_all)
        _auto_stor_cnt = sum(1 for r in _ph_all
                             if not _ph_storage_cost(r, cfg)
                             and _auto_storage_for_batch(r) > 0)

        # 적용 기준별 카운트
        _rmc_fifo_cnt   = sum(1 for r in _ph_all
                              if _fifo_rmc_map.get((r.get("scrap_type_id",""),
                                  r.get("shipment_id","") or f"__no_ship__{r.get('id','')}"
                              )) is not None)
        _rmc_stored_cnt = sum(1 for r in _ph_all
                              if r.get("raw_material_cost_per_kg") is not None
                              and _fifo_rmc_map.get((r.get("scrap_type_id",""),
                                  r.get("shipment_id","") or f"__no_ship__{r.get('id','')}"
                              )) is None)
        _rmc_auto_cnt   = sum(1 for r in _ph_all
                              if r.get("raw_material_cost_per_kg") is None
                              and _fifo_rmc_map.get((r.get("scrap_type_id",""),
                                  r.get("shipment_id","") or f"__no_ship__{r.get('id','')}"
                              )) is None
                              and _inv_moving_avg(
                                  cfg, r.get("scrap_type_id",""),
                                  _sm_pnl.get(r.get("shipment_id",""),{}).get("loading_date")
                              )[0] is not None)
        _rmc_dflt_cnt   = len(_ph_all) - _rmc_fifo_cnt - _rmc_stored_cnt - _rmc_auto_cnt

        # 거래 마진 (섹션1과 동일 기준)
        _trade_net_r = _tot_bp + _tot_sc_rev - _tot_repr - _tot_eu
        _real_net_r  = _trade_net_r - _tot_raw - _tot_storage - _sga_total - _other_cost
        _real_pct_r  = _real_net_r / _tot_bp * 100 if _tot_bp > 0 else 0

        # 원료비 적용 현황 안내
        _rmc_parts = []
        if _rmc_fifo_cnt   > 0: _rmc_parts.append(f"**{_rmc_fifo_cnt}건 FIFO**")
        if _rmc_stored_cnt > 0: _rmc_parts.append(f"{_rmc_stored_cnt}건 수동 입력")
        if _rmc_auto_cnt   > 0: _rmc_parts.append(f"{_rmc_auto_cnt}건 이동평균 자동")
        if _rmc_dflt_cnt   > 0 and _default_rmc > 0:
            _rmc_parts.append(f"{_rmc_dflt_cnt}건 수동 기본값(${_default_rmc:.4f})")
        if _rmc_parts:
            st.caption("📌 원료비 적용 기준: " + " / ".join(_rmc_parts)
                       + "  _(FIFO = 출고 기록 탭 dispatch_records 기반)_")
        if _rmc_dflt_cnt > 0 and _default_rmc == 0 and not _inv_avail_pnl and not _fifo_sc_ids_avail:
            st.warning(f"⚠️ {_rmc_dflt_cnt}건의 배치에 원료 원료단가가 없습니다. "
                       "출고 기록 탭에서 임가공 출고를 입력하거나, 스크랩 유형 관리 탭에서 기초재고를 설정하세요.")

        st.markdown("---")
        _ra1, _ra2, _ra3, _ra4, _ra5, _ra6 = st.columns(6)
        _ra1.metric("거래 마진",          f"${_trade_net_r:+,.2f}",
                    delta=f"{_trade_net_r/_tot_bp*100:+.1f}%" if _tot_bp > 0 else None)
        _ra2.metric("원료 취득원가",      f"−${_tot_raw:,.2f}",
                    delta=f"${_tot_raw/_tot_inp:.4f}/kg" if _tot_inp > 0 else None)
        _stor_delta = f"{_auto_stor_cnt}건 FIFO 자동" if _auto_stor_cnt else "미입력"
        _ra3.metric("직접 판관비 (보관)", f"−${_tot_storage:,.2f}",
                    delta=_stor_delta, delta_color="off")
        _ra4.metric("간접 판관비",        f"−${_sga_total:,.2f}")
        _ra5.metric("기타 원가",          f"−${_other_cost:,.2f}")
        _nc_r2 = "normal" if _real_net_r >= 0 else "inverse"
        _ra6.metric("실질 손익",        f"${_real_net_r:+,.2f}",
                    delta=f"실질 수익률 {_real_pct_r:+.1f}%", delta_color=_nc_r2)

    st.divider()

    # ══════════════════════════════════════════════════════════════════════════
    # 섹션 4 — 이론 마진
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("#### 📊 이론 마진")

    # INDEX 월 선택
    _th_idx_opts = [h["month"] for h in sorted(
        cfg.get("index_history", []), key=lambda x: x["month"], reverse=True)]
    _th_c1, _th_c2, _th_c3 = st.columns([2, 2, 4])
    with _th_c1:
        _th_ref = st.selectbox("INDEX 기준월", _th_idx_opts if _th_idx_opts else ["—"], key="pnl_th_ref")
    _hm = {h["month"]: h for h in cfg.get("index_history", [])}
    if _th_ref in _hm:
        _th_NI = _hm[_th_ref]["ni_index"]
        _th_CO = _hm[_th_ref]["co_index"]
    else:
        _th_NI, _th_CO = NI, CO
    with _th_c2:
        st.metric("Ni INDEX", f"${_th_NI:,.2f}")
    with _th_c3:
        st.metric("Co INDEX", f"${_th_CO:,.2f}")

    st.caption(
        "💡 스크랩 매각 ↔ BP 재매입이 상계되므로  "
        "**순 재매입 원가 = 임가공비 ÷ 전환율** (스크랩 단가 무관)  →  "
        "**이론 마진 = 매입사 BP 매각단가 − 임가공비/전환율**  "
        "⚠️ 이론 마진에는 **수출비·원료비 미포함**. 이론 vs 실적 비교는 동일 기준(수출비 제외)으로 계산됩니다."
    )

    if not active_buyers or not active_procs:
        st.info("매입사와 임가공사를 등록하면 이론 마진이 표시됩니다.")
    else:
        # 스크랩 → 출력 제품 결정 (양극 → BP, 나머지 → BM)
        def _sc_product(sc):
            return "BP" if ("양극" in sc.get("name","") or sc.get("id","") == "cathode") else "BM"

        # 매입사별 현재 단가 사전 계산 (product별 분리, 선택 INDEX 적용)
        _buyer_sale_bp = {}   # BP 매입사
        _buyer_sale_bm = {}   # BM 매입사
        for _b in active_buyers:
            _, _, _, _bskg = bp_price(_th_NI, _th_CO, _b["ni_content"], _b["co_content"],
                                      _b["ni_payable"], _b["co_payable"])
            _entry = (_b["name"], _b["product"], round(_bskg, 5))
            if _b["product"] == "BP":
                _buyer_sale_bp[_b["id"]] = _entry
            else:
                _buyer_sale_bm[_b["id"]] = _entry

        def _color_margin(val):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return ""
            return ("background-color:#1E5C35;color:#A8F0C0;font-weight:600" if val > 0
                    else "background-color:#5C1E1E;color:#F0A8A8;font-weight:600")

        def _build_th_table(scraps, buyer_map, label):
            """임가공사×스크랩 조합 + 매입사별 마진 테이블 생성"""
            if not scraps or not buyer_map:
                return
            _buyer_cols_local = [f"{n} [{p}]" for _, (n, p, _) in buyer_map.items()]
            _rows = []
            for _proc in active_procs:
                for _sc in scraps:
                    _cnd = _proc.get("conditions", {}).get(_sc["id"], {})
                    _pf  = _cnd.get("processing_fee"); _cv = _cnd.get("conversion_rate")
                    # 순 재매입 원가 = 임가공비 ÷ 전환율 (스크랩 상계)
                    _bmc = _pf / (_cv/100) if (_pf is not None and _cv and _cv>0) else None
                    row  = {
                        "임가공사":           _proc["name"],
                        "스크랩 유형":        _sc["name"],
                        "전환율(%)":          _cv,
                        "임가공비($/kg 투입)": _pf,
                        "순 재매입원가/kg BP": round(_bmc, 4) if _bmc is not None else None,
                    }
                    for _bid, (_bname, _bprod, _bskg) in buyer_map.items():
                        _col = f"{_bname} [{_bprod}]"
                        row[_col] = round(_bskg - _bmc, 4) if _bmc is not None else None
                    _rows.append(row)
            if not _rows:
                return
            _df = pd.DataFrame(_rows)
            _fmt = {
                "전환율(%)":           lambda v: f"{v}%" if v is not None else "—",
                "임가공비($/kg 투입)": lambda v: f"${v}" if v is not None else "—",
                "순 재매입원가/kg BP": lambda v: f"${v:.4f}" if v is not None else "—",
            }
            for _bc in _buyer_cols_local:
                _fmt[_bc] = lambda v: f"${v:+.4f}" if v is not None else "—"
            st.markdown(f"**{label}**")
            st.dataframe(
                _df.style.map(_color_margin, subset=_buyer_cols_local)
                         .format(na_rep="—", formatter=_fmt),
                use_container_width=True, hide_index=True
            )
            return _rows   # 이론 vs 실적 비교에서 재사용

        st.divider()
        _bp_scraps = [s for s in active_scraps if _sc_product(s) == "BP"]
        _bm_scraps = [s for s in active_scraps if _sc_product(s) == "BM"]
        _th_rows_bp = _build_th_table(_bp_scraps, _buyer_sale_bp, "🔵 BP 계열 (양극)")
        _th_rows_bm = _build_th_table(_bm_scraps, _buyer_sale_bm, "🟢 BM 계열 (젤리롤 · 셀 · 모듈)")
        _th_rows_all = (_th_rows_bp or []) + (_th_rows_bm or [])

        # ── 이론 vs 실적 비교 ───────────────────────────────────────────────
        if _linked_pnl:
            st.divider()
            st.markdown("##### 📈 이론 vs 실적 비교")
            st.caption("실제 선적 이력 기준 — HBL에 연결된 매입사의 현재 단가를 이론값으로 사용")
            _cmp_rows = []
            for _rp2 in _linked_pnl:
                _pp2   = _pm_pnl.get(_rp2.get("processor_id",""), {})
                _ss2   = _scm_pnl.get(_rp2.get("scrap_type_id",""), {})
                _ship2 = _sm_pnl.get(_rp2.get("shipment_id",""), {})
                _actual_bid   = _rp2.get("buyer_id") or _ship2.get("buyer_id","")
                _actual_buyer = _bm_pnl.get(_actual_bid, {})
                # 이론 BP/BM 단가 (실제 매입사 기준, 선택 INDEX 적용)
                _th_bskg = None
                if _actual_buyer.get("ni_payable"):
                    _, _, _, _th_bskg = bp_price(
                        _th_NI, _th_CO,
                        _actual_buyer.get("ni_content", 0),
                        _actual_buyer.get("co_content", 0),
                        _actual_buyer["ni_payable"],
                        _actual_buyer["co_payable"],
                    )
                # 이론 재매입 원가 = 임가공비 ÷ 전환율 (스크랩 상계)
                _cnd2 = _pp2.get("conditions",{}).get(_ss2.get("id",""),{}) if _pp2 else {}
                _pf_c = _cnd2.get("processing_fee"); _cv_c = _cnd2.get("conversion_rate")
                _bmc2 = _pf_c / (_cv_c/100) if (_pf_c is not None and _cv_c and _cv_c>0) else None
                _th_mg2 = round(_th_bskg - _bmc2, 5) if (_th_bskg and _bmc2) else None
                # 실적 — 수출비 포함/미포함 마진 분리 계산
                _op2  = _rp2.get("output_kg",0) or 0
                _ip2  = _ph_input_kg(_rp2)
                _sc2  = (_rp2.get("scrap_sale_per_kg",0) or 0) * _ip2
                _pf2  = (_rp2.get("processing_fee_per_kg",0) or 0) * _ip2
                _repr2= _sc2 + _pf2   # BP 재매입원가
                _eu2_total = _ph_export_usd(_rp2, cfg)
                _epk2 = _eu2_total / _op2 if _op2>0 else 0
                _bp2  = _rp2.get("bp_sale_per_kg",0) or 0
                # 실적 마진(수출비 제외) — 이론 마진과 동일 기준으로 비교
                _repr2_per_kg = _repr2 / _op2 if _op2>0 else 0
                _sc2_per_kg   = _sc2 / _op2 if _op2>0 else 0
                _act_mg2_noex = _bp2 + _sc2_per_kg - _repr2_per_kg   # 수출비 미포함
                _act_mg2      = _act_mg2_noex - _epk2                 # 수출비 포함
                _diff2 = round(_act_mg2_noex - _th_mg2, 5) if _th_mg2 is not None else None
                _cmp_rows.append({
                    "HBL":                  _ship2.get("hbl","—"),
                    "매입사":               f"{_actual_buyer.get('name','?')} ({_actual_buyer.get('product','?')})",
                    "임가공사":             _pp2.get("name","—"),
                    "스크랩":               _ss2.get("name","—"),
                    "이론 BP단가":          round(_th_bskg,5) if _th_bskg else None,
                    "이론 재매입원가/kg":   round(_bmc2,5) if _bmc2 else None,
                    "이론 마진($/kg)":      _th_mg2,
                    "실적 마진/kg(수출 제외)": round(_act_mg2_noex,5),
                    "수출비/kg":            round(_epk2,5),
                    "실적 마진/kg(수출 포함)": round(_act_mg2,5),
                    "이론대비 차이":        _diff2,
                })
            _df_cmp = pd.DataFrame(_cmp_rows)
            st.caption("💡 **이론 마진** = 매입사 단가 − 임가공비/전환율 (수출비 미포함)  "
                       "| **차이** = 실적 마진(수출 제외) − 이론 마진  (양수 = 이론보다 좋음)")
            def _hl_cmp(row):
                v = row.get("이론대비 차이")
                if v is None: return [""]*len(row)
                c = ("background-color:#1E5C35;color:#A8F0C0;font-weight:600" if v>=0
                     else "background-color:#5C1E1E;color:#F0A8A8;font-weight:600")
                return [""]*(len(row)-1) + [c]
            st.dataframe(_df_cmp.style.apply(_hl_cmp, axis=1).format(na_rep="—", formatter={
                "이론 BP단가":             lambda v: f"${v:.5f}" if v else "—",
                "이론 재매입원가/kg":      lambda v: f"${v:.5f}" if v else "—",
                "이론 마진($/kg)":         lambda v: f"${v:+.5f}" if v else "—",
                "실적 마진/kg(수출 제외)": "${:+.5f}",
                "수출비/kg":               "${:.5f}",
                "실적 마진/kg(수출 포함)": "${:+.5f}",
                "이론대비 차이":           lambda v: f"${v:+.5f}" if v else "—",
            }), use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════════════════
    # 섹션 5 — 직접 판매 손익
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 🏷️ 직접 판매 손익")
    st.caption("직접 판매 스크랩의 매각 수익 및 FIFO 원가 대비 손익을 분석합니다.")

    _ds_all = cfg.get("direct_sales", [])
    if not _ds_all:
        st.info("직접 판매 이력이 없습니다. 입출고 기록 탭에서 추가하세요.")
    else:
        _ds_with_price = [ds for ds in _ds_all if ds.get("sale_price_per_kg") is not None]
        if not _ds_with_price:
            st.info("직접 판매 단가(sale_price_per_kg)가 등록된 건이 없습니다. "
                    "입출고 기록 탭의 직접 판매 섹션에서 단가를 입력하세요.")
        else:
            # FIFO 원가 계산 (기초재고 있는 유형만)
            _ds_fifo_sc_ids = {
                ds.get("scrap_type_id","") for ds in _ds_with_price
                if cfg.get("raw_material_inventory",{}).get(ds.get("scrap_type_id",""),{}).get("opening")
            }
            _ds_fifo_cost_map = {}   # (sc_id, ds_id) → FIFO $/kg (from bl_result)
            for _dsc in _ds_fifo_sc_ids:
                _, _d_events, _ = _fifo_lot_trace(cfg, _dsc)
                for _dev in _d_events:
                    if _dev["type"] == "직접판매":
                        # sum weighted FIFO cost for this direct_sale event
                        _d_total_qty = sum(a["qty"] for a in _dev["attributions"])
                        _d_total_amt = sum(a["qty"] * a.get("unit_cost", 0) for a in _dev["attributions"])
                        if _d_total_qty > 0:
                            _ds_fifo_cost_map[(_dsc, _dev.get("id",""))] = _d_total_amt / _d_total_qty

            _ds_rows_pnl = []
            _ds_scm = {s["id"]: s["name"] for s in cfg.get("scrap_types", [])}
            for ds in sorted(_ds_with_price, key=lambda x: x.get("date","")):
                _dqty   = float(ds.get("quantity_kg", 0))
                _dspkg  = float(ds.get("sale_price_per_kg", 0) or 0)
                _drev   = _dspkg * _dqty
                _sc_id  = ds.get("scrap_type_id","")
                # FIFO 원가 — 이벤트 ref_id 매칭
                _fifo_cpkg = _ds_fifo_cost_map.get((_sc_id, ds.get("id","")))
                # 이동평균 단가 fallback
                _mavg_cpkg, _ = _inv_moving_avg(cfg, _sc_id)
                _cost_cpkg = _fifo_cpkg if _fifo_cpkg is not None else (_mavg_cpkg or 0)
                _cost_src  = "FIFO" if _fifo_cpkg is not None else ("이동평균" if _mavg_cpkg else "—")
                _dcost  = _cost_cpkg * _dqty
                _dnet   = _drev - _dcost
                _dmg    = _dnet / _drev * 100 if _drev > 0 else None
                _ds_rows_pnl.append({
                    "판매일":        ds.get("date",""),
                    "스크랩 유형":   _ds_scm.get(_sc_id, "—"),
                    "판매량 (kg)":   _dqty,
                    "단가 ($/kg)":   _dspkg,
                    "매출액 (USD)":  round(_drev, 2),
                    "원가 ($/kg)":   round(_cost_cpkg, 4) if _cost_cpkg else None,
                    "원가 산출":     _cost_src,
                    "원가 합계 (USD)": round(_dcost, 2) if _dcost else None,
                    "거래 마진 (USD)": round(_dnet, 2),
                    "거래 마진률 (%)": round(_dmg, 2) if _dmg is not None else None,
                })

            _df_ds_pnl = pd.DataFrame(_ds_rows_pnl)
            def _hl_ds(row):
                v = row.get("거래 마진 (USD)", 0) or 0
                c = ("background-color:#1E8449;color:white;font-weight:600" if v >= 0
                     else "background-color:#922b21;color:white;font-weight:600")
                return [""] * (len(row) - 2) + [c, ""]
            st.dataframe(
                _df_ds_pnl.style.apply(_hl_ds, axis=1).format(na_rep="—", formatter={
                    "판매량 (kg)":      "{:,.0f}",
                    "단가 ($/kg)":      "${:.4f}",
                    "매출액 (USD)":     "${:,.2f}",
                    "원가 ($/kg)":      lambda v: f"${v:.4f}" if v else "—",
                    "원가 합계 (USD)":  lambda v: f"${v:,.2f}" if v else "—",
                    "거래 마진 (USD)":   "${:+,.2f}",
                    "거래 마진률 (%)":   lambda v: f"{v:+.2f}%" if v is not None else "—",
                }),
                use_container_width=True, hide_index=True,
            )
            _ds_tot_rev  = sum(r["매출액 (USD)"]      for r in _ds_rows_pnl)
            _ds_tot_net  = sum(r["거래 마진 (USD)"]     for r in _ds_rows_pnl)
            _ds_tot_cost = sum(r["원가 합계 (USD)"] or 0 for r in _ds_rows_pnl)
            _ds_mg_pct   = _ds_tot_net / _ds_tot_rev * 100 if _ds_tot_rev > 0 else 0
            _dsk1, _dsk2, _dsk3 = st.columns(3)
            _dsk1.metric("직접 판매 매출액",  f"${_ds_tot_rev:,.2f}")
            _dsk2.metric("직접 판매 원가",    f"${_ds_tot_cost:,.2f}")
            _dsk3.metric("직접 판매 매출이익", f"${_ds_tot_net:+,.2f}",
                         delta=f"{_ds_mg_pct:+.2f}%",
                         delta_color="normal" if _ds_tot_net >= 0 else "inverse")
            if any(r["원가 산출"] == "이동평균" for r in _ds_rows_pnl):
                st.caption("⚠️ 일부 항목은 FIFO 추적 데이터 부족으로 **이동평균** 원가를 사용했습니다. "
                           "입출고 기록 탭의 임가공 출고 이력을 입력하면 FIFO 원가로 전환됩니다.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 8 — 매입사 관리
# ══════════════════════════════════════════════════════════════════════════════
with t_buy:
    st.subheader("BP/BM 매입사 목록")
    with st.expander("ℹ️ 지불율(Payable) 개념 안내", expanded=False):
        st.markdown("""
**지불율(Payable)**은 매입사가 LME/MB INDEX 기준 금속 가치 중 실제로 지불하는 비율입니다.

- **Ni 지불율 1.00** = LME Ni INDEX × Ni 함유량 × 1.00 (100% 지불)
- **Co 지불율 0.95** = MB Co INDEX × Co 함유량 × 0.95 (95% 지불)

**매각단가($/kg) = (NI × Ni함유량% × Ni지불율 + CO × Co함유량% × Co지불율) ÷ 1,000**

지불율이 낮을수록 매입사에 유리합니다. 계약 협상 시 이 값이 핵심 조건입니다.
""")
    for i,b in enumerate(cfg["buyers"]):
        with st.expander(f"{'✅' if b.get('active',True) else '⛔'}  {b['name']} — {b['product']}"):
            c1,c2,c3,c4,c5,c6=st.columns(6)
            with c1: nn =st.text_input("매입사명",b["name"],key=f"bn_{i}")
            with c2: np_=st.selectbox("품목",["BP","BM"],["BP","BM"].index(b["product"]),key=f"bpr_{i}")
            with c3: nnp=st.number_input("Ni 지불율",value=b["ni_payable"],step=0.01,format="%.2f",key=f"bnp_{i}")
            with c4: ncp=st.number_input("Co 지불율",value=b["co_payable"],step=0.01,format="%.2f",key=f"bcp_{i}")
            with c5: nnc=st.number_input("Ni 함유량(%)",value=b["ni_content"],step=0.01,format="%.2f",key=f"bnc_{i}")
            with c6: ncc=st.number_input("Co 함유량(%)",value=b["co_content"],step=0.01,format="%.2f",key=f"bcc_{i}")
            ba,bb=st.columns(2)
            with ba: na=st.checkbox("활성",b.get("active",True),key=f"bact_{i}")
            with bb:
                s1,s2,s3,s4=st.columns(4)
                with s1:
                    if st.button("▲",key=f"b_up_{i}",disabled=(i==0),use_container_width=True):
                        cfg["buyers"][i-1],cfg["buyers"][i]=cfg["buyers"][i],cfg["buyers"][i-1]
                        save_cfg(cfg); st.rerun()
                with s2:
                    if st.button("▽",key=f"b_dn_{i}",disabled=(i==len(cfg["buyers"])-1),use_container_width=True):
                        cfg["buyers"][i+1],cfg["buyers"][i]=cfg["buyers"][i],cfg["buyers"][i+1]
                        save_cfg(cfg); st.rerun()
                with s3:
                    if st.button("💾 저장",key=f"bsave_{i}",use_container_width=True):
                        cfg["buyers"][i].update({"name":nn,"product":np_,"ni_payable":nnp,"co_payable":ncp,"ni_content":nnc,"co_content":ncc,"active":na})
                        save_cfg(cfg); st.toast("✅ 저장 완료"); st.rerun()
                with s4:
                    with st.popover("🗑️", use_container_width=True):
                        st.warning(f"매입사 **{b['name']}** 삭제")
                        if st.button("삭제 확인", key=f"bdel_cfm_{i}", type="primary", use_container_width=True):
                            cfg["buyers"].pop(i); save_cfg(cfg); st.rerun()
    st.divider()
    st.subheader("새 매입사 추가")
    with st.form("add_buyer"):
        a1,a2,a3,a4,a5,a6=st.columns(6)
        with a1: anm=st.text_input("매입사명")
        with a2: apr=st.selectbox("품목",["BP","BM"])
        with a3: anp=st.number_input("Ni 지불율",value=1.0,step=0.01,format="%.2f")
        with a4: acp=st.number_input("Co 지불율",value=1.0,step=0.01,format="%.2f")
        with a5: anc=st.number_input("Ni 함유량(%)",value=41.93,step=0.01,format="%.2f")
        with a6: acc=st.number_input("Co 함유량(%)",value=7.23,step=0.01,format="%.2f")
        if st.form_submit_button("➕ 추가"):
            if not anm: st.error("매입사명을 입력하세요.")
            else:
                cfg["buyers"].append({"id":str(uuid.uuid4())[:8],"name":anm,"product":apr,"ni_payable":anp,"co_payable":acp,"ni_content":anc,"co_content":acc,"active":True})
                save_cfg(cfg); st.success(f"{anm} ({apr}) 추가!"); st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 — 임가공사 관리
# ══════════════════════════════════════════════════════════════════════════════
with t_proc:
    st.subheader("임가공사 관리")
    scrap_list   = cfg.get("scrap_types", [])
    ph_list      = cfg.get("processing_history", [])
    proc_map_ph  = {p["id"]: p for p in cfg.get("processors", [])}
    scrap_map_ph = {s["id"]: s for s in scrap_list}
    buyer_map_ph = {b["id"]: b for b in cfg.get("buyers", [])}
    ship_list_ph = cfg.get("shipments", [])
    ship_map_ph  = {s["id"]: s for s in ship_list_ph}

    # HBL 연결 옵션 목록
    ship_opts_ph = {"(미연결)": ""}
    for _soi, _sos in enumerate(ship_list_ph):
        _sob    = buyer_map_ph.get(_sos.get("buyer_id",""), {})
        _sohbl  = _sos.get("hbl","").strip() or f"HBL미정 {_sos.get('weight_kg',0):,.0f}kg"
        ship_opts_ph[f"#{_soi+1}  {_sohbl}  ({_sos.get('loading_date','?')} · {_sob.get('name','?')})"] = _sos["id"]

    # ── 서브탭 ──────────────────────────────────────────────────────────────
    proc_tab1, proc_tab2 = st.tabs(["📋 전체 내역", "📦 세부 내역 (HBL)"])

    # ══════════════════════════════════════════════════════════════════════════
    # 서브탭 1 — 전체 내역 : 계약 조건 + 누적 실적
    # ══════════════════════════════════════════════════════════════════════════
    with proc_tab1:
        # ── 배치 누락 경고 ───────────────────────────────────────────────────
        _dr_list_warn = cfg.get("dispatch_records", [])
        _ph_list_warn = cfg.get("processing_history", [])
        _warn_msgs = []
        for _sc_w in scrap_list:
            _scid_w  = _sc_w["id"]
            _dr_qty  = sum(float(r.get("quantity_kg",0)) for r in _dr_list_warn if r.get("scrap_type_id")==_scid_w)
            _ph_inp  = sum(_ph_input_kg(r) for r in _ph_list_warn if r.get("scrap_type_id")==_scid_w)
            _remain  = _dr_qty - _ph_inp
            if _remain > 100:   # 100kg 초과 미처리 시 경고
                _warn_msgs.append(f"**{_sc_w['name']}** — 출고 {_dr_qty:,.0f}kg 중 투입 미기록 {_remain:,.0f}kg (임가공사 보유 추정)")
        if _warn_msgs:
            with st.expander(f"🏭 임가공사 작업 중 (처리 결과 미입력) {len(_warn_msgs)}건", expanded=True):
                st.caption("출고 기록 대비 처리 결과(배치)가 아직 없는 스크랩입니다. 임가공이 완료되면 HBL별 탭에서 배치를 추가하세요.")
                for _wm in _warn_msgs:
                    st.markdown(f"- {_wm}")

        if not cfg.get("processors"):
            st.info("등록된 임가공사가 없습니다.")
        else:
            for _proc in cfg.get("processors", []):
                _pname = ('✅ ' if _proc.get('active', True) else '⛔ ') + _proc['name']
                st.markdown(f"#### {_pname}")
                _rows_sum = []
                for _sc in scrap_list:
                    _sid2 = _sc["id"]
                    _cond = _proc.get("conditions", {}).get(_sid2, {})
                    _recs = [p for p in ph_list
                             if p.get("processor_id") == _proc["id"]
                             and p.get("scrap_type_id") == _sid2]
                    _tot_out = sum(r.get("output_kg", 0) or 0 for r in _recs)
                    _tot_in  = sum(_ph_input_kg(r) for r in _recs)
                    _act_cv  = round(_tot_out / _tot_in * 100, 2) if _tot_in > 0 else None
                    _sc_cost = sum((r.get("scrap_sale_per_kg",0) or 0) * _ph_input_kg(r) for r in _recs)
                    _pf_cost = sum((r.get("processing_fee_per_kg",0) or 0) * _ph_input_kg(r) for r in _recs)
                    _bpr_kg  = (_sc_cost + _pf_cost) / _tot_out if _tot_out > 0 else None
                    _rows_sum.append({
                        "스크랩 유형":      _sc["name"],
                        "계약 임임가공비($/kg)": _cond.get("processing_fee"),
                        "계약 전환율(%)":    _cond.get("conversion_rate"),
                        "계약 불순물율(%)":  _cond.get("impurity_rate"),
                        "배치수":           len(_recs),
                        "누적 투입(kg)":    round(_tot_in, 0)  if _tot_in  > 0 else None,
                        "누적 생산(kg)":    round(_tot_out, 0) if _tot_out > 0 else None,
                        "실제 전환율(%)":   _act_cv,
                        "BP 재매입가($/kg)":round(_bpr_kg, 4)  if _bpr_kg  else None,
                    })
                st.dataframe(pd.DataFrame(_rows_sum).style.format(na_rep="—", formatter={
                    "계약 임임가공비($/kg)": lambda v: f"${v}" if v is not None else "—",
                    "계약 전환율(%)":      lambda v: f"{v}%" if v is not None else "—",
                    "계약 불순물율(%)":    lambda v: f"{v}%" if v is not None else "—",
                    "누적 투입(kg)":       lambda v: f"{v:,.0f}" if v is not None else "—",
                    "누적 생산(kg)":       lambda v: f"{v:,.0f}" if v is not None else "—",
                    "실제 전환율(%)":      lambda v: f"{v:.2f}%" if v is not None else "—",
                    "BP 재매입가($/kg)":   lambda v: f"${v:.4f}" if v is not None else "—",
                }), use_container_width=True, hide_index=True)
                st.divider()

        # ── 계약 조건 편집 (접힘) ────────────────────────────────────────
        with st.expander("⚙️ 계약 조건 편집 / 임가공사 추가·삭제", expanded=False):
            for pi, proc in enumerate(cfg.get("processors", [])):
                st.markdown(f"**{proc['name']}**")
                hc1, hc2 = st.columns([3, 1])
                with hc1: pnm  = st.text_input("임가공사명", proc["name"], key=f"pnm_{pi}")
                with hc2: pact = st.checkbox("활성", proc.get("active", True), key=f"pact_{pi}")
                gh = st.columns([1.5, 1.5, 1.5, 1.5])
                for _col, _lbl in zip(gh, ["스크랩 유형","임임가공비($/kg)","전환율(%)","불순물율(%)"]):
                    _col.markdown(f"**{_lbl}**")
                new_conds = {sid: dict(v) for sid, v in proc.get("conditions", {}).items()}
                for scrap in scrap_list:
                    sid  = scrap["id"]
                    cond = new_conds.setdefault(sid, {})
                    rc   = st.columns([1.5, 1.5, 1.5, 1.5])
                    rc[0].markdown(f'<span class="b-sc">{scrap["name"]}</span>', unsafe_allow_html=True)
                    def _ni(col, val, key):
                        raw = col.text_input("_", value="" if val is None else str(val),
                                             key=key, label_visibility="collapsed")
                        try: return float(raw) if raw.strip() else None
                        except: return val
                    cond["processing_fee"]  = _ni(rc[1], cond.get("processing_fee"),  f"ppf_{pi}_{sid}")
                    cond["conversion_rate"] = _ni(rc[2], cond.get("conversion_rate"), f"pcv_{pi}_{sid}")
                    cond["impurity_rate"]   = _ni(rc[3], cond.get("impurity_rate"),   f"pim_{pi}_{sid}")
                pc1, pc2, pc3, pc4 = st.columns(4)
                with pc1:
                    if st.button("▲", key=f"p_up_{pi}", disabled=(pi==0), use_container_width=True):
                        cfg["processors"][pi-1], cfg["processors"][pi] = cfg["processors"][pi], cfg["processors"][pi-1]
                        save_cfg(cfg); st.rerun()
                with pc2:
                    if st.button("▽", key=f"p_dn_{pi}", disabled=(pi==len(cfg["processors"])-1), use_container_width=True):
                        cfg["processors"][pi+1], cfg["processors"][pi] = cfg["processors"][pi], cfg["processors"][pi+1]
                        save_cfg(cfg); st.rerun()
                with pc3:
                    if st.button("💾 저장", key=f"psave_{pi}", use_container_width=True):
                        cfg["processors"][pi].update({"name": pnm, "active": pact, "conditions": new_conds})
                        save_cfg(cfg); st.toast("✅ 저장 완료"); st.rerun()
                with pc4:
                    with st.popover("🗑️", use_container_width=True):
                        st.warning(f"임가공사 **{proc['name']}** 삭제")
                        if st.button("삭제 확인", key=f"pdel_cfm_{pi}", type="primary", use_container_width=True):
                            cfg["processors"].pop(pi); save_cfg(cfg); st.rerun()
                st.markdown("---")

            st.markdown("**➕ 새 임가공사 추가**")
            with st.form("add_proc"):
                pnew = st.text_input("임가공사명")
                if st.form_submit_button("추가"):
                    if not pnew: st.error("임가공사명을 입력하세요.")
                    else:
                        blank = {s["id"]: {"processing_fee": None, "conversion_rate": None, "impurity_rate": None}
                                 for s in scrap_list}
                        cfg["processors"].append({"id": str(uuid.uuid4())[:8], "name": pnew,
                                                  "active": True, "conditions": blank})
                        save_cfg(cfg); st.success(f"{pnew} 추가!"); st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # 서브탭 2 — HBL 중심 배치 관리
    # ══════════════════════════════════════════════════════════════════════════
    with proc_tab2:
        st.caption("HBL을 선택해 연결 배치를 관리합니다. 배치 추가 시 임가공비는 계약 표준값이 자동 적용됩니다.")

        # ── 공통 옵션 ────────────────────────────────────────────────────────
        _pp_opts_t2 = {p["name"]: p["id"] for p in cfg.get("processors",[])}
        _ps_opts_t2 = {s["name"]: s["id"] for s in scrap_list}
        # 선적건 HBL 선택용 (미연결 배치에서 연결할 때 사용)
        _ship_opts_t2 = {"(미연결)": ""}
        for _so in sorted(ship_list_ph, key=lambda x: x.get("loading_date",""), reverse=True):
            _so_hbl = _so.get("hbl","").strip() or f"HBL미정 {_so.get('weight_kg',0):,.0f}kg"
            _ship_opts_t2[f"{_so_hbl} [{_so.get('loading_date','?')[:7]}]"] = _so["id"]

        # ── 새 선적건 인라인 등록 ────────────────────────────────────────────
        with st.expander("➕ 새 선적건 등록", expanded=not ship_list_ph):
            with st.form("t2_new_ship"):
                _ns1, _ns2, _ns3 = st.columns(3)
                with _ns1:
                    _ns_hbl = st.text_input("HBL")
                    _ns_ld  = st.text_input("선적일 (YYYY-MM-DD)", placeholder="2026-04-01")
                with _ns2:
                    _ns_buy = st.selectbox("매입사", list(buyer_opts.keys()))
                    _ns_wkg = st.number_input("선적 중량 (kg)", value=0.0, step=1.0, format="%.0f")
                with _ns3:
                    _ns_inv  = st.number_input("Invoice USD (잠정)", value=0.0, step=1.0, format="%.2f")
                    _ns_pm   = st.selectbox("Provisional 월", ["—"]+hist_opts)
                    _ns_etd  = st.text_input("ETD (선택)", placeholder="2026-04-05")
                if st.form_submit_button("➕ 선적건 추가"):
                    _ns_errs = []
                    if not _ns_hbl: _ns_errs.append("HBL을 입력하세요.")
                    if not _ns_ld:  _ns_errs.append("선적일을 입력하세요.")
                    else:
                        try: datetime.strptime(_ns_ld, "%Y-%m-%d")
                        except ValueError: _ns_errs.append("선적일 형식이 잘못됐습니다 (YYYY-MM-DD)")
                    if _ns_hbl and any(s.get("hbl","").strip()==_ns_hbl.strip() for s in ship_list_ph):
                        _ns_errs.append(f"HBL '{_ns_hbl}' 이(가) 이미 존재합니다.")
                    if _ns_errs:
                        for _e in _ns_errs: st.error(_e)
                    else:
                        cfg["shipments"].append({
                            "id": str(uuid.uuid4())[:8],
                            "hbl": _ns_hbl, "invoice_no": "",
                            "loading_date": _ns_ld, "etd": _ns_etd.strip(), "eta": "",
                            "buyer_id": buyer_opts[_ns_buy],
                            "weight_kg": _ns_wkg, "invoice_usd": _ns_inv,
                            "prov_month": _ns_pm, "final_month": "—",
                            "status": "provisional", "notes": "",
                            "moisture_pct": None, "buyer_ni_content": None,
                            "buyer_co_content": None,
                            "other_adj_usd": None, "other_adj_desc": "",
                        })
                        save_cfg(cfg); st.success("선적건 추가 완료 — 아래에서 선택하세요."); st.rerun()

        st.markdown("---")

        # ── HBL 선택 ─────────────────────────────────────────────────────────
        # 배치 수 사전 계산
        _t2_batch_cnt = {}
        for _ph_c in ph_list:
            _sid_c = _ph_c.get("shipment_id","")
            if _sid_c:
                _t2_batch_cnt[_sid_c] = _t2_batch_cnt.get(_sid_c, 0) + 1
        _unlinked_cnt = sum(1 for _ph_c in ph_list if not _ph_c.get("shipment_id",""))
        # 고아 배치: shipment_id가 있지만 해당 선적건이 삭제되어 존재하지 않는 배치
        _valid_sids   = {s["id"] for s in ship_list_ph}
        _orphan_cnt   = sum(1 for _ph_c in ph_list
                             if _ph_c.get("shipment_id","") and _ph_c.get("shipment_id","") not in _valid_sids)

        _t2_hbl_d = {"─ HBL 선택 ─": None}
        for _s2 in sorted(ship_list_ph, key=lambda x: x.get("loading_date",""), reverse=True):
            _b2     = buyer_map_ph.get(_s2.get("buyer_id",""), {})
            _icon2  = {"provisional":"🟡","final":"🟢","paid":"🔵"}.get(_s2.get("status",""),"⚪")
            _bcnt2  = _t2_batch_cnt.get(_s2["id"], 0)
            _bcnt_lbl = f"  [{_bcnt2}건]" if _bcnt2 else "  [배치없음]"
            _hbl_lbl = _s2.get("hbl","").strip() or f"HBL미정 {_s2.get('weight_kg',0):,.0f}kg"
            _t2_hbl_d[
                f"{_icon2}  {_hbl_lbl}  |  "
                f"{_s2.get('loading_date','?')[:7]}  "
                f"{_b2.get('name','?')} ({_b2.get('product','?')})"
                f"{_bcnt_lbl}"
            ] = _s2["id"]
        _unlinked_lbl = f"🔖 미연결 배치" + (f"  [{_unlinked_cnt}건]" if _unlinked_cnt else "  [없음]")
        _t2_hbl_d[_unlinked_lbl] = "__unlinked__"
        if _orphan_cnt:
            _t2_hbl_d[f"🔗💥 깨진 연결 (선적건 삭제됨)  [{_orphan_cnt}건]"] = "__orphan__"

        _t2_sel = st.selectbox("HBL 선택", list(_t2_hbl_d.keys()), key="t2_hbl_sel",
                               label_visibility="collapsed")
        _t2_sid = _t2_hbl_d[_t2_sel]

        # ── slim 배치 편집 헬퍼 ──────────────────────────────────────────────
        def _slim_batch_expander(ph_list_ref, idx, rec, fixed_hbl_sid, ship_obj):
            """slim 배치 expander.
            fixed_hbl_sid: 상위 HBL 선택값. None이면 드롭다운으로 HBL 선택 가능 (미연결 배치용).
            """
            _bpo  = proc_map_ph.get(rec.get("processor_id",""), {})
            _bso  = scrap_map_ph.get(rec.get("scrap_type_id",""), {})
            _bout = float(rec.get("output_kg",0) or 0)
            _bbps = float(rec.get("bp_sale_per_kg") or 0)
            _bcv  = float(rec.get("conversion_rate_pct") or rec.get("conversion_rate") or 0)
            _rk   = rec.get("id", str(idx))  # stable key: batch id 사용
            _lbl  = (f"{'🔗' if fixed_hbl_sid else '🔖'}  "
                     f"{_bpo.get('name','?')} × {_bso.get('name','?')}  |  "
                     f"{_bout:,.0f} kg BP  |  ${_bbps:.4f}/kg")
            with st.expander(_lbl, expanded=False):
                # 미연결 배치: HBL 연결 드롭다운 표시
                if fixed_hbl_sid is None:
                    _cur_ship_lbl = next(
                        (k for k,v in _ship_opts_t2.items() if v == rec.get("shipment_id","")),
                        "(미연결)")
                    _e_hbl_link = st.selectbox("HBL 연결",
                        list(_ship_opts_t2.keys()),
                        index=list(_ship_opts_t2.keys()).index(_cur_ship_lbl)
                              if _cur_ship_lbl in _ship_opts_t2 else 0,
                        key=f"slim_hbl_{_rk}")
                    _save_hbl_sid = _ship_opts_t2[_e_hbl_link]
                else:
                    _save_hbl_sid = fixed_hbl_sid

                _ec1, _ec2 = st.columns(2)
                with _ec1:
                    _e_pp = st.selectbox("임가공사", list(_pp_opts_t2.keys()),
                        index=list(_pp_opts_t2.values()).index(rec.get("processor_id",""))
                              if rec.get("processor_id","") in _pp_opts_t2.values() else 0,
                        key=f"slim_proc_{_rk}")
                    _e_ps = st.selectbox("스크랩", list(_ps_opts_t2.keys()),
                        index=list(_ps_opts_t2.values()).index(rec.get("scrap_type_id",""))
                              if rec.get("scrap_type_id","") in _ps_opts_t2.values() else 0,
                        key=f"slim_scrap_{_rk}")
                    _econd = proc_map_ph.get(_pp_opts_t2[_e_pp],{}).get("conditions",{}).get(_ps_opts_t2[_e_ps],{})
                    _e_cv = st.number_input("전환율 (%)",
                        value=float(_bcv or _econd.get("conversion_rate") or 0),
                        min_value=0.0, max_value=100.0, step=0.1, format="%.2f",
                        key=f"slim_conv_{_rk}",
                        help=f"계약값: {_econd.get('conversion_rate')}%" if _econd.get("conversion_rate") else "계약 전환율 미설정")
                with _ec2:
                    _auto_w = float(ship_obj.get("weight_kg",0)) if ship_obj else 0.0
                    _e_out = st.number_input("생산량 kg",
                        value=float(rec.get("output_kg") or _auto_w),
                        step=1.0, format="%.0f", key=f"slim_out_{_rk}")
                    _e_inp = _e_out / (_e_cv/100) if _e_cv > 0 else 0
                    st.metric("투입량 (스크랩)", f"{_e_inp:,.0f} kg")
                    _e_bps = st.number_input("BP 매각단가 ($/kg)",
                        value=float(rec.get("bp_sale_per_kg") or 0),
                        step=0.0001, format="%.4f", key=f"slim_bps_{_rk}")
                    _e_note = st.text_input("비고", rec.get("notes",""), key=f"slim_note_{_rk}")

                _es1, _es2 = st.columns(2)
                with _es1:
                    if st.button("💾 저장", key=f"slim_save_{_rk}", use_container_width=True):
                        ph_list_ref[idx].update({
                            "shipment_id":         _save_hbl_sid,
                            "processor_id":        _pp_opts_t2[_e_pp],
                            "scrap_type_id":       _ps_opts_t2[_e_ps],
                            "conversion_rate_pct": _e_cv if _e_cv > 0 else None,
                            "input_kg":            _e_inp,
                            "output_kg":           _e_out,
                            "bp_sale_per_kg":      _e_bps,
                            "notes":               _e_note,
                        })
                        save_cfg(cfg); st.toast("✅ 저장"); st.rerun()
                with _es2:
                    with st.popover("🗑️", use_container_width=True):
                        st.warning(f"배치 **{_bpo.get('name','?')} × {_bso.get('name','?')}** 삭제")
                        if st.button("삭제 확인", key=f"slim_del_cfm_{_rk}", type="primary", use_container_width=True):
                            ph_list_ref.pop(idx); save_cfg(cfg); st.rerun()

        # ── 뷰 분기 ──────────────────────────────────────────────────────────
        if _t2_sid is None:
            st.info("위에서 HBL을 선택하거나, 선적건이 없으면 위 **새 선적건 등록**을 먼저 펼치세요.")

        elif _t2_sid == "__unlinked__":
            _unlinked = [(i,p) for i,p in enumerate(ph_list) if not p.get("shipment_id","")]
            if not _unlinked:
                st.success("HBL 미연결 배치 없음 ✅  (모든 배치가 선적건에 연결되어 있습니다)")
            else:
                st.warning(f"⚠️ 미연결 배치 {len(_unlinked)}건 — 배치를 열어 HBL을 연결하세요.")
                for _uri, _up in _unlinked:
                    _slim_batch_expander(ph_list, _uri, _up, None, {})

        elif _t2_sid == "__orphan__":
            _orphans = [(i,p) for i,p in enumerate(ph_list)
                        if p.get("shipment_id","") and p.get("shipment_id","") not in _valid_sids]
            if not _orphans:
                st.success("깨진 연결 없음 ✅")
            else:
                st.error(f"💥 연결된 선적건이 삭제되어 고아가 된 배치 {len(_orphans)}건 — "
                         f"손익 분석 탭에 'HBL —' 카드로 표시됩니다. 삭제하거나 다른 HBL에 재연결하세요.")
                for _ori, _op in _orphans:
                    _sc_o = scrap_map_ph.get(_op.get("scrap_type_id",""), {}).get("name","?")
                    _proc_o = proc_map_ph.get(_op.get("processor_id",""), {}).get("name","?")
                    with st.expander(
                        f"💥 {_proc_o} · {_sc_o} · output {_op.get('output_kg',0):,.0f}kg "
                        f"(존재하지 않는 shipment_id: {_op.get('shipment_id','')})",
                        expanded=False,
                    ):
                        st.caption(f"배치 ID: {_op.get('id','')}")
                        _oc1, _oc2 = st.columns(2)
                        with _oc1:
                            _o_relink_sel = st.selectbox(
                                "다른 HBL로 재연결", list(_ship_opts_t2.keys()),
                                key=f"orphan_relink_{_op.get('id','')}",
                            )
                            if st.button("재연결", key=f"orphan_relink_btn_{_op.get('id','')}"):
                                ph_list[_ori]["shipment_id"] = _ship_opts_t2[_o_relink_sel]
                                save_cfg(cfg); st.toast("✅ 재연결 완료"); st.rerun()
                        with _oc2:
                            with st.popover("🗑️ 배치 삭제", use_container_width=True):
                                st.warning("이 고아 배치를 삭제합니다. 되돌릴 수 없습니다.")
                                if st.button("삭제 확인", key=f"orphan_del_cfm_{_op.get('id','')}",
                                             type="primary", use_container_width=True):
                                    ph_list.pop(_ori); save_cfg(cfg); st.rerun()

        else:
            _t2_ship  = ship_map_ph.get(_t2_sid, {})
            _t2_buyer = buyer_map_ph.get(_t2_ship.get("buyer_id",""), {})
            _stat2_lbl= {"provisional":"🟡 Provisional","final":"🟢 최종","paid":"🔵 입금"}.get(
                         _t2_ship.get("status",""),"—")

            # HBL 정보 요약 바
            _ti1,_ti2,_ti3,_ti4,_ti5,_ti6 = st.columns(6)
            _ti1.metric("선적일",    _t2_ship.get("loading_date","—"))
            _ti2.metric("매입사",    f"{_t2_buyer.get('name','?')} ({_t2_buyer.get('product','?')})")
            _ti3.metric("선적 중량", f"{_t2_ship.get('weight_kg',0):,.0f} kg")
            _ti4.metric("Invoice",   f"${_t2_ship.get('invoice_usd',0):,.0f}")
            _ti6.metric("상태",      _stat2_lbl)
            _eu_disp = _t2_ship.get("export_cost_usd")
            _ti5.metric("수출비",
                        f"${_eu_disp:,.0f}" if _eu_disp else "—",
                        help="선적 정산 탭에서 입력")

            # 연결된 배치 목록
            _t2_batches = [(i,p) for i,p in enumerate(ph_list)
                           if p.get("shipment_id","") == _t2_sid]
            st.markdown(f"**📦 연결 배치 ({len(_t2_batches)}건)**")

            if _t2_batches:
                for _bri2, _bp2 in _t2_batches:
                    _slim_batch_expander(ph_list, _bri2, _bp2, _t2_sid, _t2_ship)
            else:
                st.info("연결된 배치가 없습니다. 아래에서 추가하세요.")

            # ── 배치 추가 폼 ──────────────────────────────────────────────────
            st.markdown("---")
            _fk = _t2_sid[:8]  # form key prefix
            with st.form(f"add_ph2_{_fk}"):
                st.markdown("**➕ 배치 추가**")
                _fa1, _fa2, _fa3 = st.columns(3)
                with _fa1:
                    _fnp = st.selectbox("임가공사",    list(_pp_opts_t2.keys()), key=f"fa_proc_{_fk}")
                    _fns = st.selectbox("스크랩 유형", list(_ps_opts_t2.keys()), key=f"fa_scrap_{_fk}")
                with _fa2:
                    _fn_cond = proc_map_ph.get(_pp_opts_t2.get(_fnp,""),{}).get(
                                   "conditions",{}).get(_ps_opts_t2.get(_fns,""),{})
                    _fn_cv_d = _fn_cond.get("conversion_rate")
                    _fno = st.number_input("생산량 kg",
                        value=float(_t2_ship.get("weight_kg",0)),
                        step=1.0, format="%.0f", key=f"fa_out_{_fk}")
                    _fnv = st.number_input("전환율 (%)",
                        value=float(_fn_cv_d or 0),
                        min_value=0.0, max_value=100.0, step=0.1, format="%.2f",
                        key=f"fa_cv_{_fk}",
                        help=f"계약값 {_fn_cv_d}%" if _fn_cv_d else "계약 전환율 미설정")
                with _fa3:
                    _auto_bps_f = 0.0
                    if _t2_buyer:
                        _, _, _, _auto_bps_f = bp_price(
                            NI, CO,
                            _t2_buyer.get("ni_content",0), _t2_buyer.get("co_content",0),
                            _t2_buyer.get("ni_payable",0), _t2_buyer.get("co_payable",0))
                    _fnbps = st.number_input("BP 매각단가 ($/kg)",
                        value=_auto_bps_f, step=0.0001, format="%.4f", key=f"fa_bps_{_fk}",
                        help=f"현재 INDEX → {_t2_buyer.get('name','?')}: ${_auto_bps_f:.5f}" if _t2_buyer else "")
                    _fnnotes = st.text_input("비고", key=f"fa_notes_{_fk}")
                if st.form_submit_button("➕ 추가"):
                    _fn_inp = _fno / (_fnv/100) if _fnv > 0 else 0
                    cfg.setdefault("processing_history",[]).append({
                        "id":                    str(uuid.uuid4())[:8],
                        "shipment_id":           _t2_sid,
                        "processor_id":          _pp_opts_t2.get(_fnp,""),
                        "scrap_type_id":         _ps_opts_t2.get(_fns,""),
                        "output_kg":             _fno,
                        "conversion_rate_pct":   _fnv if _fnv > 0 else None,
                        "input_kg":              _fn_inp,
                        "bp_sale_per_kg":        _fnbps,
                        "processing_fee_per_kg": _fn_cond.get("processing_fee"),  # 계약값 자동
                        "buyer_id":              _t2_buyer.get("id","") if _t2_buyer else "",
                        "notes":                 _fnnotes,
                    })
                    save_cfg(cfg); st.toast("✅ 추가 완료!"); st.rerun()

            # ── HBL 손익 요약 ─────────────────────────────────────────────────
            if _t2_batches:
                st.markdown("---")
                _h2_bp=0.0; _h2_pf=0.0
                for _, _bph2 in _t2_batches:
                    _h2out = float(_bph2.get("output_kg",0) or 0)
                    _h2inp = _ph_input_kg(_bph2)
                    _h2_bp += float(_bph2.get("bp_sale_per_kg",0) or 0) * _h2out
                    _h2_pf += float(_bph2.get("processing_fee_per_kg",0) or 0) * _h2inp
                # 수출비: HBL 레벨 값 직접 사용
                _h2_eu = float(_t2_ship.get("export_cost_usd") or 0)
                _h2_net = _h2_bp - _h2_pf - _h2_eu
                _h2_mg  = _h2_net / _h2_bp * 100 if _h2_bp > 0 else 0
                _hm1,_hm2,_hm3,_hm4 = st.columns(4)
                _hm1.metric("BP 매각",   f"${_h2_bp:,.0f}")
                _hm2.metric("임가공비",  f"${_h2_pf:,.0f}", help="계약값 자동 적용")
                _hm3.metric("수출비",    f"${_h2_eu:,.0f}")
                _hm4.metric("거래 마진",  f"${_h2_net:+,.0f}",
                            delta=f"{_h2_mg:+.1f}%",
                            delta_color="normal" if _h2_net >= 0 else "inverse")



# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 — 스크랩 유형 관리  (지불율 없음)
# ══════════════════════════════════════════════════════════════════════════════
with t_stype:
    st.subheader("스크랩 유형 마스터")
    st.caption("샘플 분석 결과 업데이트 시 함유량을 수정하세요.")
    for si,sc in enumerate(cfg.get("scrap_types",[])):
        with st.expander(f"{'✅' if sc.get('active',True) else '⛔'}  {sc['name']}",expanded=True):
            d1,d2,d3,d4=st.columns(4)
            with d1: snm =st.text_input("유형명",sc["name"],key=f"st_nm_{si}")
            with d2: sni =st.number_input("Ni 함유량(%)",sc["ni_content"],step=0.01,format="%.2f",key=f"st_ni_{si}")
            with d3: sco =st.number_input("Co 함유량(%)",sc["co_content"],step=0.01,format="%.2f",key=f"st_co_{si}")
            with d4: sact=st.checkbox("활성",sc.get("active",True),key=f"st_act_{si}")
            d5,d6,d7,d8=st.columns(4)
            with d5:
                srate=st.number_input("창고비 (EUR/톤백/day)",
                    value=float(sc.get("storage_rate_eur") or 1.5),
                    step=0.1, format="%.2f", key=f"st_rate_{si}",
                    help="FIFO 자동 창고비 계산에 사용. 양극=1.3, 젤리롤=1.0 등 유형별로 설정")
            sa,sb,sc_btn,sd=d6,d7,d8,st.empty()
            with sa:
                if st.button("▲",key=f"st_up_{si}",disabled=(si==0),use_container_width=True):
                    cfg["scrap_types"][si-1],cfg["scrap_types"][si]=cfg["scrap_types"][si],cfg["scrap_types"][si-1]
                    save_cfg(cfg); st.rerun()
            with sb:
                if st.button("▽",key=f"st_dn_{si}",disabled=(si==len(cfg["scrap_types"])-1),use_container_width=True):
                    cfg["scrap_types"][si+1],cfg["scrap_types"][si]=cfg["scrap_types"][si],cfg["scrap_types"][si+1]
                    save_cfg(cfg); st.rerun()
            with sc_btn:
                if st.button("💾 저장",key=f"st_save_{si}",use_container_width=True):
                    cfg["scrap_types"][si].update({"name":snm,"ni_content":sni,"co_content":sco,
                                                   "active":sact,"storage_rate_eur":float(srate)})
                    save_cfg(cfg); st.toast("✅ 저장 완료"); st.rerun()
            _sdd1,_sdd2=st.columns([1,3])
            with _sdd1:
                with st.popover("🗑️", use_container_width=True):
                    st.warning(f"스크랩 유형 **{sc['name']}** 삭제")
                    if st.button("삭제 확인", key=f"st_del_cfm_{si}", type="primary", use_container_width=True):
                        cfg["scrap_types"].pop(si); save_cfg(cfg); st.rerun()
    st.divider()
    st.subheader("새 스크랩 유형 추가")
    with st.form("add_scrap"):
        e1,e2,e3,e4=st.columns(4)
        with e1: enm  =st.text_input("유형명")
        with e2: eni  =st.number_input("Ni 함유량(%)",value=0.0,step=0.01,format="%.2f")
        with e3: eco  =st.number_input("Co 함유량(%)",value=0.0,step=0.01,format="%.2f")
        with e4: erate=st.number_input("창고비 (EUR/톤백/day)",value=1.5,step=0.1,format="%.2f")
        if st.form_submit_button("➕ 추가"):
            if not enm: st.error("유형명을 입력하세요.")
            else:
                sid=str(uuid.uuid4())[:8]
                cfg["scrap_types"].append({"id":sid,"name":enm,"ni_content":eni,"co_content":eco,
                                           "active":True,"storage_rate_eur":float(erate)})
                for p in cfg["processors"]:
                    p.setdefault("conditions",{})[sid]={"processing_fee":None,"conversion_rate":None,"impurity_rate":None}
                save_cfg(cfg); st.success(f"{enm} 추가!"); st.rerun()

    # (원료 재고 관리 섹션 → 입출고 기록 탭으로 이동)



# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 — 입출고 기록  (입고 이력 + 임가공 출고 + 직접 판매)
# ══════════════════════════════════════════════════════════════════════════════
with t_outflow:
    st.subheader("📥 입출고 기록")
    st.caption("원료 입고 이력과 임가공/직접판매 출고 이력을 관리합니다. "
               "FIFO Lot 추적 및 자동 창고비 계산의 기준 데이터입니다.")

    # ⑤ 창고비율 미설정 스크랩 유형 경고
    _missing_rate = [s["name"] for s in cfg.get("scrap_types", [])
                     if s.get("active", True) and not s.get("storage_rate_eur")]
    if _missing_rate:
        st.warning(f"⚠️ 창고비율(EUR/톤백/day) 미설정 스크랩 유형: **{', '.join(_missing_rate)}** — "
                   "FIFO 자동 창고비 계산 시 기본값(1.5 EUR)이 적용됩니다. "
                   "**스크랩 유형 관리** 탭에서 설정하세요.")

    _of_sc_opts = {s["name"]: s["id"] for s in cfg.get("scrap_types", [])}
    _of_sc_rev  = {v: k for k, v in _of_sc_opts.items()}
    _of_pr_opts = {p["name"]: p["id"] for p in cfg.get("processors", [])}
    _of_pr_rev  = {v: k for k, v in _of_pr_opts.items()}

    # ══════════════════════════════════════════════════════════════════════════
    # 원료 재고 관리 (이동평균법)
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### 📦 원료 재고 관리 (이동평균법)")
    st.caption("스크랩 유형별 기초재고와 입고 이력을 등록합니다.  \n"
               "이동평균단가 → 손익 분석 탭 실질 손익에 자동 반영.  \n"
               "입고일 **YYYY-MM-DD** + 톤백 수 입력 시 FIFO 창고비 자동 계산 정확도↑")

    if "raw_material_inventory" not in cfg:
        cfg["raw_material_inventory"] = {}
        # 저장은 하지 않음 — 실제 입력 시 save_cfg 호출

    # ── 스크랩 유형별 상세 ────────────────────────────────────────────────────
    _ph_all_inv = cfg.get("processing_history", [])
    for _isc in cfg.get("scrap_types", []):
        _isid = _isc["id"]; _isnm = _isc["name"]
        _stor_rate_disp = float(_isc.get("storage_rate_eur") or 1.5)
        if _isid not in cfg["raw_material_inventory"]:
            cfg["raw_material_inventory"][_isid] = {"opening": None, "purchases": []}
        _inv_i   = cfg["raw_material_inventory"][_isid]
        _avg_cur, _qty_cur = _inv_moving_avg(cfg, _isid)
        _bal_cur = _inv_balance(cfg, _isid, _ph_all_inv)
        if _avg_cur is not None:
            _exp_lbl = (f"📦 **{_isnm}** — 이동평균 ${_avg_cur:.5f}/kg | "
                        f"누적입고 {_qty_cur:,.0f} kg | 잔량 {_bal_cur:,.0f} kg | "
                        f"창고비 EUR {_stor_rate_disp:.2f}/톤백/day")
        else:
            _exp_lbl = f"📦 **{_isnm}** — 기초재고 미설정"
        with st.expander(_exp_lbl, expanded=(_avg_cur is None)):
            st.markdown("##### 기초재고")
            _op_i = _inv_i.get("opening") or {}
            _ic1,_ic2,_ic3,_ic4 = st.columns(4)
            with _ic1:
                _op_date_i = st.text_input("기준일 (YYYY-MM-DD)",
                    value=_op_i.get("date",""), key=f"inv_op_dt_{_isid}", placeholder="예: 2025-11-11")
            with _ic2:
                _op_qty_i = st.number_input("기초 재고량 (kg)",
                    value=float(_op_i.get("quantity_kg") or 0), step=1.0, format="%.0f", key=f"inv_op_qty_{_isid}")
            with _ic3:
                _op_cost_i = st.number_input("기초 평균단가 ($/kg)",
                    value=float(_op_i.get("unit_cost") or 0), step=0.00001, format="%.5f", key=f"inv_op_cost_{_isid}")
            with _ic4:
                st.markdown("&nbsp;", unsafe_allow_html=True)
                if st.button("💾 저장", key=f"inv_op_save_{_isid}", use_container_width=True):
                    if _op_date_i and _op_qty_i > 0 and _op_cost_i > 0:
                        cfg["raw_material_inventory"][_isid]["opening"] = {
                            "date": _op_date_i, "quantity_kg": float(_op_qty_i), "unit_cost": float(_op_cost_i)}
                        save_cfg(cfg); st.success("기초재고 저장 완료"); st.rerun()
                    else:
                        st.error("기준일·재고량·단가를 모두 입력하세요.")
            if _avg_cur is not None:
                st.markdown("---")
                _im1,_im2,_im3 = st.columns(3)
                _im1.metric("현재 이동평균단가", f"${_avg_cur:.5f}/kg")
                _im2.metric("누적 입고량",       f"{_qty_cur:,.0f} kg")
                _nc_bal = "inverse" if _bal_cur < 0 else "off"
                _im3.metric("잔량 (추정)", f"{_bal_cur:,.0f} kg",
                            delta="음수 재고 확인 필요" if _bal_cur < 0 else None, delta_color=_nc_bal)
            st.markdown("---")
            st.markdown("##### 입고 이력")
            _purs_i = sorted(_inv_i.get("purchases", []), key=lambda x: x.get("date",""))
            if _purs_i:
                _df_pur_i = pd.DataFrame([{
                    "날짜":           p.get("date",""),
                    "입고량 (kg)":    float(p.get("quantity_kg", 0)),
                    "톤백 (개)":      int(p.get("ton_bags",0) or 0),
                    "원료단가 ($/kg)": float(p.get("unit_cost", 0)),
                } for p in _purs_i])
                st.dataframe(_df_pur_i.style.format({
                    "입고량 (kg)": "{:,.0f}", "톤백 (개)": "{:,.0f}", "원료단가 ($/kg)": "${:.5f}",
                }), use_container_width=True, hide_index=True)
                _del_lbls_i = [
                    f"{p.get('date','')}  {float(p.get('quantity_kg',0)):,.0f} kg "
                    f"({int(p.get('ton_bags',0) or 0)}백) @ ${float(p.get('unit_cost',0)):.5f}"
                    for p in _purs_i]
                _deld1, _deld2 = st.columns([4,1])
                with _deld1:
                    # 인덱스를 값으로 사용 — 라벨 중복(동일 날짜·수량·단가) 시 오삭제 방지
                    _di_i = st.selectbox("삭제할 입고 건", range(len(_purs_i)),
                                         format_func=lambda i: _del_lbls_i[i], key=f"inv_del_sel_{_isid}")
                with _deld2:
                    st.markdown("&nbsp;", unsafe_allow_html=True)
                    with st.popover("🗑️", use_container_width=True):
                        st.warning(f"입고 건 삭제:\n{_del_lbls_i[_di_i]}")
                        if st.button("삭제 확인", key=f"inv_del_cfm_{_isid}", type="primary", use_container_width=True):
                            _pur_copy_i = list(_purs_i); _pur_copy_i.pop(_di_i)
                            cfg["raw_material_inventory"][_isid]["purchases"] = _pur_copy_i
                            save_cfg(cfg); st.rerun()
            else:
                st.info("등록된 입고 이력이 없습니다.")
            st.markdown("##### 입고 추가")
            _ac1,_ac2,_ac3,_ac4,_ac5 = st.columns(5)
            with _ac1:
                _new_dt_i = st.text_input("입고일 (YYYY-MM-DD)", key=f"inv_add_dt_{_isid}", placeholder="예: 2026-02-15")
            with _ac2:
                _new_qty_i = st.number_input("입고량 (kg)", value=0.0, step=1.0, format="%.0f", key=f"inv_add_qty_{_isid}")
            with _ac3:
                _new_tb_i  = st.number_input("톤백 (개)", value=0, step=1, format="%d", key=f"inv_add_tb_{_isid}",
                                             help="실제 톤백 수 — 비워도 되며 입고량÷510으로 역산")
            with _ac4:
                _new_cost_i = st.number_input("원료단가 ($/kg)", value=0.0, step=0.00001, format="%.5f", key=f"inv_add_cost_{_isid}")
            with _ac5:
                st.markdown("&nbsp;", unsafe_allow_html=True)
                if st.button("➕ 추가", key=f"inv_add_btn_{_isid}", use_container_width=True):
                    if _new_dt_i and _new_qty_i > 0 and _new_cost_i > 0:
                        cfg["raw_material_inventory"][_isid]["purchases"].append({
                            "date": _new_dt_i, "quantity_kg": float(_new_qty_i),
                            "ton_bags": int(_new_tb_i), "unit_cost": float(_new_cost_i)})
                        save_cfg(cfg); st.success("입고 추가 완료"); st.rerun()
                    else:
                        st.error("입고일·입고량·단가를 모두 입력하세요.")


        # ── 임가공 출고 이력 ──────────────────────────────────────────────────────
    st.markdown("### 🏭 임가공 출고 이력")
    st.caption("스크랩을 임가공사(톨링)로 출고한 날짜와 수량을 기록합니다. "
               "각 B/L 배치가 어느 입고 Lot에서 비롯되었는지 추적하는 기준이 됩니다.")

    if "dispatch_records" not in cfg:
        cfg["dispatch_records"] = []

    _dr_list = cfg["dispatch_records"]

    if _dr_list:
        _dr_rows = []
        for dr in sorted(_dr_list, key=lambda x: x.get("date", ""), reverse=True):
            _dr_rows.append({
                "출고일":      dr.get("date", ""),
                "임가공사":    _of_pr_rev.get(dr.get("processor_id", ""), "—"),
                "스크랩 유형": _of_sc_rev.get(dr.get("scrap_type_id", ""), "—"),
                "출고량 (kg)": float(dr.get("quantity_kg", 0)),
                "톤백 (개)":   int(dr.get("ton_bags", 0) or 0),
                "비고":        dr.get("notes", ""),
            })
        st.dataframe(
            pd.DataFrame(_dr_rows).style.format({"출고량 (kg)": "{:,.0f}", "톤백 (개)": "{:,.0f}"}),
            use_container_width=True, hide_index=True
        )

        _dr_del_opts = [
            f"{dr.get('date','')}  |  {_of_pr_rev.get(dr.get('processor_id',''),'—')}  |  "
            f"{_of_sc_rev.get(dr.get('scrap_type_id',''),'—')}  |  {float(dr.get('quantity_kg',0)):,.0f} kg"
            for dr in _dr_list
        ]
        _drd1, _drd2 = st.columns([4, 1])
        with _drd1:
            # 인덱스를 값으로 사용 — 라벨 중복 시 오삭제 방지
            _dri = st.selectbox("삭제할 출고 건", range(len(_dr_list)),
                                 format_func=lambda i: _dr_del_opts[i], key="dr_del_sel")
        with _drd2:
            st.markdown("&nbsp;", unsafe_allow_html=True)
            with st.popover("🗑️", use_container_width=True):
                st.warning(f"임가공 출고 삭제:\n{_dr_del_opts[_dri]}")
                if st.button("삭제 확인", key="dr_del_cfm", type="primary", use_container_width=True):
                    cfg["dispatch_records"].pop(_dri)
                    save_cfg(cfg); st.rerun()
    else:
        st.info("등록된 임가공 출고 기록이 없습니다.")

    st.markdown("##### ➕ 임가공 출고 추가")
    with st.form("add_dr"):
        _dr_a1, _dr_a2, _dr_a3, _dr_a4, _dr_a5 = st.columns(5)
        with _dr_a1: _dr_date  = st.text_input("출고일 (YYYY-MM-DD)", placeholder="예: 2025-11-15", key="dr_date")
        with _dr_a2: _dr_proc  = st.selectbox("임가공사", list(_of_pr_opts.keys()), key="dr_proc")
        with _dr_a3: _dr_sc    = st.selectbox("스크랩 유형", list(_of_sc_opts.keys()), key="dr_sc")
        with _dr_a4: _dr_qty   = st.number_input("출고량 (kg)", value=0.0, step=100.0, format="%.0f", key="dr_qty")
        with _dr_a5: _dr_notes = st.text_input("비고", placeholder="배치번호, 차량번호 등", key="dr_notes")
        if st.form_submit_button("➕ 출고 추가"):
            if not _dr_date or _dr_qty <= 0:
                st.error("출고일과 출고량을 입력하세요.")
            elif not _of_pr_opts:
                st.error("임가공사를 먼저 등록하세요.")
            else:
                cfg["dispatch_records"].append({
                    "id":            str(uuid.uuid4())[:8],
                    "date":          _dr_date.strip(),
                    "processor_id":  _of_pr_opts[_dr_proc],
                    "scrap_type_id": _of_sc_opts[_dr_sc],
                    "quantity_kg":   float(_dr_qty),
                    "notes":         _dr_notes,
                })
                save_cfg(cfg); st.success("임가공 출고 추가 완료"); st.rerun()

    # ── 직접 판매 출고 이력 ───────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🏷️ 직접 판매 출고 이력")
    st.caption("임가공(톨링) 없이 스크랩을 직접 판매한 경우 여기에 기록합니다. "
               "FIFO Lot 추적 계산 시 소진 이벤트로 반영됩니다.")

    if "direct_sales" not in cfg:
        cfg["direct_sales"] = []

    _ds_list    = cfg["direct_sales"]

    if _ds_list:
        _ds_rows = []
        for ds in sorted(_ds_list, key=lambda x: x.get("date", ""), reverse=True):
            _spkg = ds.get("sale_price_per_kg")
            _dqty = float(ds.get("quantity_kg", 0))
            _ds_rows.append({
                "판매일":        ds.get("date", ""),
                "스크랩 유형":   _of_sc_rev.get(ds.get("scrap_type_id", ""), "—"),
                "판매량 (kg)":   _dqty,
                "톤백 (개)":     int(ds.get("ton_bags", 0) or 0),
                "단가 ($/kg)":   float(_spkg) if _spkg is not None else None,
                "매출액 (USD)":  round(float(_spkg) * _dqty, 2) if _spkg is not None else None,
                "비고":          ds.get("notes", ""),
            })
        st.dataframe(
            pd.DataFrame(_ds_rows).style.format({
                "판매량 (kg)":  "{:,.0f}",
                "톤백 (개)":    "{:,.0f}",
                "단가 ($/kg)":  lambda v: f"${v:.4f}" if v is not None else "—",
                "매출액 (USD)": lambda v: f"${v:,.2f}" if v is not None else "—",
            }),
            use_container_width=True, hide_index=True
        )

        _ds_del_opts = [
            f"{ds.get('date','')}  |  {_of_sc_rev.get(ds.get('scrap_type_id',''),'—')}  |  "
            f"{float(ds.get('quantity_kg',0)):,.0f} kg"
            for ds in _ds_list
        ]
        _dsd1, _dsd2 = st.columns([4, 1])
        with _dsd1:
            # 인덱스를 값으로 사용 — 라벨 중복 시 오삭제 방지
            _dsi = st.selectbox("삭제할 판매 건", range(len(_ds_list)),
                                 format_func=lambda i: _ds_del_opts[i], key="ds_del_sel")
        with _dsd2:
            st.markdown("&nbsp;", unsafe_allow_html=True)
            with st.popover("🗑️", use_container_width=True):
                st.warning(f"직접 판매 삭제:\n{_ds_del_opts[_dsi]}")
                if st.button("삭제 확인", key="ds_del_cfm", type="primary", use_container_width=True):
                    cfg["direct_sales"].pop(_dsi)
                    save_cfg(cfg); st.rerun()
    else:
        st.info("등록된 직접 판매 출고 이력이 없습니다.")

    st.markdown("##### ➕ 직접 판매 추가")
    with st.form("add_ds"):
        _ds_a1, _ds_a2, _ds_a3, _ds_a4, _ds_a5 = st.columns(5)
        with _ds_a1: _ds_date  = st.text_input("판매일 (YYYY-MM-DD)", placeholder="예: 2026-02-10", key="ds_date")
        with _ds_a2: _ds_sc    = st.selectbox("스크랩 유형", list(_of_sc_opts.keys()), key="ds_sc")
        with _ds_a3: _ds_qty   = st.number_input("판매량 (kg)", value=0.0, step=1.0, format="%.0f", key="ds_qty")
        with _ds_a4: _ds_price = st.number_input("단가 ($/kg)", value=0.0, step=0.01, format="%.4f", key="ds_price",
                                                  help="직접 판매 단가 ($/kg 스크랩). 0 입력 시 미등록.")
        with _ds_a5: _ds_notes = st.text_input("비고", placeholder="거래처, 용도 등", key="ds_notes")
        if st.form_submit_button("➕ 추가"):
            if not _ds_date or _ds_qty <= 0:
                st.error("판매일과 수량을 입력하세요.")
            else:
                cfg["direct_sales"].append({
                    "id":               str(uuid.uuid4())[:8],
                    "date":             _ds_date.strip(),
                    "scrap_type_id":    _of_sc_opts[_ds_sc],
                    "quantity_kg":      float(_ds_qty),
                    "sale_price_per_kg": float(_ds_price) if _ds_price > 0 else None,
                    "notes":            _ds_notes,
                })
                save_cfg(cfg); st.success("추가 완료"); st.rerun()


    # ══════════════════════════════════════════════════════════════════════════
    # 원료 Lot 추적 (FIFO)
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 🔍 원료 Lot 추적 (FIFO)")
    st.caption("임가공 출고 기록을 기준으로 2단계 FIFO를 적용해 각 B/L이 어느 입고 Lot의 스크랩으로 "
               "생산되었는지 추적합니다.")

    _lt_inv_types = [s for s in cfg.get("scrap_types", [])
                     if cfg.get("raw_material_inventory", {}).get(s["id"], {}).get("opening")]
    if not _lt_inv_types:
        st.info("스크랩 유형 관리 탭에서 기초재고를 설정하면 Lot 추적이 가능합니다.")
    else:
        _ltc1, _ltc2 = st.columns([2, 3])
        with _ltc1:
            _lt_sc_opts = {s["name"]: s["id"] for s in _lt_inv_types}
            _lt_sc_sel  = st.selectbox("스크랩 유형 선택", list(_lt_sc_opts.keys()), key="lt_sc_sel")
            _lt_sc_id   = _lt_sc_opts[_lt_sc_sel]

        # 2단계 FIFO 계산
        _lt_bl_result, _lt_events, _lt_remaining = _fifo_lot_trace(cfg, _lt_sc_id)

        # dispatch_records 미입력 경고
        _has_dispatch = any(dr.get("scrap_type_id") == _lt_sc_id
                            for dr in cfg.get("dispatch_records", []))
        _has_ph = any(r.get("scrap_type_id") == _lt_sc_id
                      for r in cfg.get("processing_history", []))
        if _has_ph and not _has_dispatch:
            st.warning("⚠️ 임가공 출고 기록이 없습니다. 위 **임가공 출고 이력** 섹션에서 먼저 출고를 입력하세요.")

        with _ltc2:
            _lt_view_opts = {"📋 출고 이벤트 흐름": "__all__"}
            for _sid, _sdata in _lt_bl_result.items():
                if not _sid.startswith("__no_ship__") and _sdata.get("hbl", "미연결") != "미연결":
                    _k = f"🚢 {_sdata['hbl']}"
                    if _k not in _lt_view_opts:
                        _lt_view_opts[_k] = _sid
            if any(e["type"] == "직접판매" for e in _lt_events):
                _lt_view_opts["🏷️ 직접 판매"] = "__direct__"
            _lt_view_sel = st.selectbox("B/L 또는 보기 선택", list(_lt_view_opts.keys()), key="lt_view_sel")
            _lt_view_id  = _lt_view_opts[_lt_view_sel]

        if not _lt_events and not _lt_bl_result:
            st.info("이 스크랩 유형의 출고 이벤트가 없습니다.")

        # ── 출고 이벤트 흐름 (Level 1 결과) ─────────────────────────────────
        elif _lt_view_id == "__all__":
            _lt_all_rows = []
            for _lev in _lt_events:
                _ev_date = _lev["date"]
                for _attr in _lev["attributions"]:
                    # 보관일수: 출고일 - Lot 입고일
                    _lot_dt = _attr.get("lot_date", "")
                    _stor_d = None
                    if _lot_dt and _ev_date and _ev_date != "9999-12-31":
                        try:
                            _stor_d = max(0, (
                                datetime.strptime(_ev_date, "%Y-%m-%d") -
                                datetime.strptime(_lot_dt[:10], "%Y-%m-%d")
                            ).days)
                        except Exception:
                            _stor_d = None
                    _lt_all_rows.append({
                        "출고일":         _ev_date,
                        "유형":           _lev["type"],
                        "임가공사":       _lev["processor"],
                        "소진 Lot":       _attr["lot_label"],
                        "소진량 (kg)":    _attr["qty"],
                        "보관일수":       _stor_d,
                        "원료단가($/kg)": _attr["unit_cost"],
                        "원가 (USD)":     _attr["amount"],
                        "비고":           _lev.get("notes", ""),
                    })
            if _lt_all_rows:
                st.dataframe(pd.DataFrame(_lt_all_rows).style.format(na_rep="—", formatter={
                    "소진량 (kg)":    "{:,.1f}",
                    "보관일수":       lambda v: f"{int(v)}일" if v is not None else "—",
                    "원료단가($/kg)": lambda v: f"${v:.5f}" if v is not None else "—",
                    "원가 (USD)":     lambda v: f"${v:,.2f}" if v is not None else "—",
                }), use_container_width=True, hide_index=True)
            else:
                st.info("출고 이벤트가 없습니다.")
            _lt_rem_rows = [{"Lot": l["label"], "잔량 (kg)": round(l["remain"], 1),
                              "원료단가": l["unit_cost"]}
                             for l in _lt_remaining if l["remain"] > 0.001]
            if _lt_rem_rows:
                st.markdown("**📦 미소진 Lot 잔량** (모든 출고 차감 후)")
                st.dataframe(pd.DataFrame(_lt_rem_rows).style.format({
                    "잔량 (kg)": "{:,.1f}", "원료단가": "${:.5f}"}),
                    use_container_width=True, hide_index=True)

        # ── 직접 판매 ─────────────────────────────────────────────────────────
        elif _lt_view_id == "__direct__":
            _lt_ds_evs = [e for e in _lt_events if e["type"] == "직접판매"]
            _lt_ds_rows = []
            for _lev in _lt_ds_evs:
                for _attr in _lev["attributions"]:
                    _lt_ds_rows.append({
                        "판매일":         _lev["date"],
                        "소진 Lot":       _attr["lot_label"],
                        "소진량 (kg)":    _attr["qty"],
                        "원료단가($/kg)": _attr["unit_cost"],
                        "원가 (USD)":     _attr["amount"],
                        "비고":           _lev.get("notes", ""),
                    })
            if _lt_ds_rows:
                st.dataframe(pd.DataFrame(_lt_ds_rows).style.format(na_rep="—", formatter={
                    "소진량 (kg)":    "{:,.1f}",
                    "원료단가($/kg)": lambda v: f"${v:.5f}" if v is not None else "—",
                    "원가 (USD)":     lambda v: f"${v:,.2f}" if v is not None else "—",
                }), use_container_width=True, hide_index=True)
            else:
                st.info("직접 판매 이력이 없습니다.")

        # ── 특정 B/L 상세 ─────────────────────────────────────────────────────
        else:
            _lt_bl_data = _lt_bl_result.get(_lt_view_id)
            if not _lt_bl_data:
                st.info("해당 B/L에 연결된 처리 이력이 없습니다.")
            else:
                _lot_dict   = _lt_bl_data["lots"]
                _total_qty  = sum(v["qty"] for v in _lot_dict.values())
                _total_amt  = sum(v["amount"] for v in _lot_dict.values()
                                  if v.get("unit_cost") is not None)
                _total_stor = _lt_bl_data.get("storage_cost", 0.0)
                _wavg_cost  = _total_amt / _total_qty if _total_qty > 0 else 0
                _bm1, _bm2, _bm3, _bm4, _bm5 = st.columns(5)
                _bm1.metric("HBL",              _lt_bl_data["hbl"])
                _bm2.metric("총 투입 스크랩",   f"{_lt_bl_data['input_kg']:,.0f} kg")
                _bm3.metric("가중평균 원료단가", f"${_wavg_cost:.5f}/kg")
                _bm4.metric("총 원료비 (추정)",  f"${_total_amt:,.2f}")
                _bm5.metric("FIFO 자동 보관비",  f"${_total_stor:,.2f}",
                            help="Lot 입고일 → 임가공 출고일 기준 자동 계산")
                _lt_bl_rows = []
                for _lbl, _v in _lot_dict.items():
                    _qty = _v["qty"]
                    # 가중평균 보관일수: storage_days_wsum / qty
                    _wsum = _v.get("storage_days_wsum", 0.0)
                    _avg_days = round(_wsum / _qty, 1) if _qty > 0 else None  # wsum=0도 유효(당일 출고=0일)
                    _lt_bl_rows.append({
                        "입고 Lot":        _lbl,
                        "입고일":          _v.get("lot_date") or "—",
                        "원료단가($/kg)":  _v.get("unit_cost"),
                        "소진량 (kg)":     round(_qty, 1),
                        "비중 (%)":        round(_qty / _total_qty * 100, 1) if _total_qty else 0,
                        "원가 (USD)":      round(_v["amount"], 2) if _v.get("unit_cost") else None,
                        "보관일수 (평균)": _avg_days,
                        "보관비 (USD)":    round(_v["storage_cost"], 2) if _v.get("storage_cost") else None,
                    })
                st.dataframe(pd.DataFrame(_lt_bl_rows).style.format(na_rep="—", formatter={
                    "원료단가($/kg)":  lambda v: f"${v:.5f}" if v is not None else "—",
                    "소진량 (kg)":    "{:,.1f}",
                    "비중 (%)":       "{:.1f}%",
                    "원가 (USD)":     lambda v: f"${v:,.2f}" if v is not None else "—",
                    "보관일수 (평균)":lambda v: f"{v:.0f}일" if v is not None else "—",
                    "보관비 (USD)":   lambda v: f"${v:,.2f}" if v is not None else "—",
                }), use_container_width=True, hide_index=True)
                st.caption("보관일수 = Lot 입고일 → 임가공 출고일 (기초재고 Lot은 기준일부터 기산)  "
                           "| 복수 출고에 걸친 Lot은 가중평균으로 표시")

        # ── 임가공 출고 합계 (임가공사별) ─────────────────────────────────────
        st.divider()
        st.markdown("##### 📦 임가공 출고 합계 (임가공사별)")

        _proc_map_lt = {p["id"]: p["name"] for p in cfg.get("processors", [])}
        _dp_by_proc  = defaultdict(float)
        for _dr in cfg.get("dispatch_records", []):
            if _dr.get("scrap_type_id") == _lt_sc_id:
                _dp_by_proc[_dr.get("processor_id", "__없음__")] += float(_dr.get("quantity_kg", 0))
        _ph_by_proc = defaultdict(float)
        for _r in cfg.get("processing_history", []):
            if _r.get("scrap_type_id") == _lt_sc_id:
                _ph_by_proc[_r.get("processor_id", "__없음__")] += _ph_input_kg(_r)

        _lt_inv_total_in  = _inv_moving_avg(cfg, _lt_sc_id)[1]
        _lt_inv_remaining = sum(l["remain"] for l in _lt_remaining if l["remain"] > 0.001)
        _lt_total_disp    = sum(_dp_by_proc.values())
        _lt_total_proc    = sum(_ph_by_proc.values())

        _dm1, _dm2, _dm3, _dm4 = st.columns(4)
        _dm1.metric("누적 입고량",     f"{_lt_inv_total_in:,.0f} kg")
        _dm2.metric("총 임가공 출고",  f"{_lt_total_disp:,.0f} kg",
                    help="dispatch_records 합계")
        _dm3.metric("총 B/L 투입",     f"{_lt_total_proc:,.0f} kg",
                    help="processing_history input_kg 합계")
        _dm4.metric("창고 미출고 잔량", f"{_lt_inv_remaining:,.0f} kg",
                    help="Lot 큐에서 모든 출고 차감 후 남은 양 = 아직 임가공사로 보내지 않은 재고")

        _all_proc_ids = set(list(_dp_by_proc.keys()) + list(_ph_by_proc.keys()))
        if _all_proc_ids:
            _disp_rows = []
            for _pid in sorted(_all_proc_ids, key=lambda x: _proc_map_lt.get(x, "ㅎ")):
                _disp   = _dp_by_proc.get(_pid, 0)
                _proced = _ph_by_proc.get(_pid, 0)
                _disp_rows.append({
                    "임가공사":               _proc_map_lt.get(_pid, "미연결"),
                    "출고 누적 (kg)":         round(_disp, 0),
                    "B/L 투입 누적 (kg)":     round(_proced, 0),
                    "임가공사 보유 추정 (kg)": round(_disp - _proced, 0),
                })
            def _hl_disp(row):
                v = row.get("임가공사 보유 추정 (kg)", 0) or 0
                if v < -1:  return ["", "", "", "color:#b71c1c;font-weight:600"]
                if v > 0.5: return ["", "", "", "color:#1565c0;font-weight:600"]
                return [""] * 4
            st.dataframe(
                pd.DataFrame(_disp_rows).style
                .apply(_hl_disp, axis=1)
                .format({
                    "출고 누적 (kg)":          "{:,.0f}",
                    "B/L 투입 누적 (kg)":      "{:,.0f}",
                    "임가공사 보유 추정 (kg)":  "{:+,.0f}",
                }),
                use_container_width=True, hide_index=True
            )
            st.caption(
                "ℹ️ **창고 미출고 잔량** + **임가공사 보유 추정** = 원료 재고 관리의 추정 잔량  \n"
                "⚠️ 음수이면 출고 기록보다 B/L 투입량이 많음 → 출고 기록 누락 확인 필요"
            )

    # ══════════════════════════════════════════════════════════════════════════
    # 원배출자(SK) 지급 계산
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("### 💴 원배출자(SK) 지급 계산")
    st.caption("입고 기록의 원료단가를 기본값으로 자동 계산합니다. 실제 SK 납품단가가 다르면 직접 수정 후 저장하세요.")

    if "sk_prices" not in cfg:
        cfg["sk_prices"] = {}

    _sk_sc_list = [s for s in cfg.get("scrap_types", []) if s.get("active", True)]
    _sk_tabs    = st.tabs([s["name"] for s in _sk_sc_list]) if _sk_sc_list else []

    for _sk_tab, _sk_sc in zip(_sk_tabs, _sk_sc_list):
        with _sk_tab:
            _sk_scid = _sk_sc["id"]
            _sk_inv  = cfg.get("raw_material_inventory", {}).get(_sk_scid, {})
            _sk_purs = _sk_inv.get("purchases", [])

            # 월별 입고량 + 가중평균 단가 집계 (기존 unit_cost 활용)
            _sk_by_mo = {}   # month → {qty, amount}
            for _sp in _sk_purs:
                _sp_mo  = (_sp.get("date") or "")[:7]
                _sp_qty = float(_sp.get("quantity_kg") or 0)
                _sp_uc  = float(_sp.get("unit_cost") or 0)
                if _sp_mo and _sp_qty > 0:
                    if _sp_mo not in _sk_by_mo:
                        _sk_by_mo[_sp_mo] = {"qty": 0.0, "amount": 0.0}
                    _sk_by_mo[_sp_mo]["qty"]    += _sp_qty
                    _sk_by_mo[_sp_mo]["amount"] += _sp_qty * _sp_uc

            if not _sk_by_mo:
                st.info("입고 이력이 없습니다.")
                continue

            # 저장된 수동 단가 로드 (있으면 우선 적용)
            _sk_saved = cfg.get("sk_prices", {}).get(_sk_scid, {})

            # 편집용 DataFrame 구성 — 기본값: 입고 기록 가중평균 단가
            _sk_rows = []
            for _mo in sorted(_sk_by_mo.keys()):
                _qty     = _sk_by_mo[_mo]["qty"]
                _amt_pur = _sk_by_mo[_mo]["amount"]
                _avg_pur = _amt_pur / _qty if _qty else 0.0
                # 수동 저장 단가 우선, 없으면 입고 기록 가중평균
                _price = float(_sk_saved.get(_mo, _avg_pur) or _avg_pur)
                _sk_rows.append({
                    "월":             _mo,
                    "입고량 (kg)":    round(_qty, 0),
                    "입고 기록 단가": round(_avg_pur, 5),   # 참고용 (고정)
                    "SK 단가 ($/kg)": round(_price, 5),     # 편집 가능
                })

            _sk_df = pd.DataFrame(_sk_rows)

            st.caption("💡 **SK 단가** = 입고 기록 단가로 자동 채워집니다. 실제 납품단가가 다르면 해당 셀을 직접 수정하세요.")
            _sk_edited = st.data_editor(
                _sk_df,
                column_config={
                    "월":              st.column_config.TextColumn("월", disabled=True),
                    "입고량 (kg)":     st.column_config.NumberColumn("입고량 (kg)", disabled=True, format="%,.0f"),
                    "입고 기록 단가":  st.column_config.NumberColumn("입고 기록 단가 ($/kg)", disabled=True, format="$%.5f"),
                    "SK 단가 ($/kg)":  st.column_config.NumberColumn("SK 납품단가 ($/kg)", format="$%.5f",
                                                                      min_value=0.0, step=0.0001),
                },
                use_container_width=True,
                hide_index=True,
                key=f"sk_editor_{_sk_scid}",
                disabled=READ_ONLY,
            )

            # 지급액 계산
            _sk_edited["지급액 (USD)"] = (_sk_edited["입고량 (kg)"] * _sk_edited["SK 단가 ($/kg)"]).round(2)
            _sk_total_qty = _sk_edited["입고량 (kg)"].sum()
            _sk_total_pay = _sk_edited["지급액 (USD)"].sum()
            _sk_avg_price = _sk_total_pay / _sk_total_qty if _sk_total_qty > 0 else 0.0

            # 결과 테이블 (합계 행 포함)
            _sk_result_rows = []
            for _, _sr in _sk_edited.iterrows():
                _sk_result_rows.append({
                    "월":            _sr["월"],
                    "입고량 (kg)":   _sr["입고량 (kg)"],
                    "SK 단가($/kg)": _sr["SK 단가 ($/kg)"],
                    "지급액 (USD)":  _sr["지급액 (USD)"],
                })
            _sk_result_rows.append({
                "월":            "합계",
                "입고량 (kg)":   _sk_total_qty,
                "SK 단가($/kg)": _sk_avg_price,
                "지급액 (USD)":  _sk_total_pay,
            })
            st.dataframe(
                pd.DataFrame(_sk_result_rows).style.apply(
                    lambda row: ["font-weight:700"] * len(row) if row["월"] == "합계" else [""] * len(row),
                    axis=1
                ).format(na_rep="—", formatter={
                    "입고량 (kg)":   "{:,.0f}",
                    "SK 단가($/kg)": lambda v: f"${v:.5f}" if v else "—",
                    "지급액 (USD)":  lambda v: f"${v:,.2f}" if v else "—",
                }),
                use_container_width=True, hide_index=True,
            )

            _sk_m1, _sk_m2, _sk_m3 = st.columns(3)
            _sk_m1.metric("총 입고량", f"{_sk_total_qty:,.0f} kg")
            _sk_m2.metric("총 지급액", f"${_sk_total_pay:,.2f}")
            _sk_m3.metric("평균 단가", f"${_sk_avg_price:.5f}/kg")

            # 수동 수정분만 저장 (입고 기록 단가와 다른 경우)
            if st.button("💾 수정 단가 저장", key=f"sk_save_{_sk_scid}", type="primary",
                         help="입고 기록 단가와 다르게 수정한 경우에만 저장하세요. 동일하면 저장 불필요."):
                if "sk_prices" not in cfg:
                    cfg["sk_prices"] = {}
                if _sk_scid not in cfg["sk_prices"]:
                    cfg["sk_prices"][_sk_scid] = {}
                _saved_cnt = 0
                for _, _sr in _sk_edited.iterrows():
                    _mo_key = _sr["월"]
                    _new_p  = float(_sr["SK 단가 ($/kg)"])
                    _orig_p = round(float(_sk_by_mo[_mo_key]["amount"] / _sk_by_mo[_mo_key]["qty"]), 5) if _sk_by_mo[_mo_key]["qty"] else 0
                    if abs(_new_p - _orig_p) > 0.00001:   # 입고 기록과 다를 때만
                        cfg["sk_prices"][_sk_scid][_mo_key] = _new_p
                        _saved_cnt += 1
                    elif _mo_key in cfg.get("sk_prices", {}).get(_sk_scid, {}):
                        del cfg["sk_prices"][_sk_scid][_mo_key]  # 원래대로 돌아오면 삭제
                save_cfg(cfg)
                st.toast(f"✅ {_saved_cnt}개월 수정 단가 저장")


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# Google Sheets 동기화 헬퍼
# ══════════════════════════════════════════════════════════════════════════════
# _GSHEET_CREDS는 파일 상단에 정의됨
_GSHEET_NAME  = "bp_calculator_sync"
_GSHEET_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

@st.cache_resource(ttl=3600, show_spinner=False)
def _gsheet_connect():
    """gspread 클라이언트 반환 (세션 간 공유, 1시간 캐시). 실패 시 예외 발생."""
    import gspread
    creds = _get_gcp_creds(_GSHEET_SCOPES)
    return gspread.Client(auth=creds)

def _match_buyer_id(cell_val, buyers):
    """'ECOPRO (BP)' 형식 문자열 → buyer_id. 대소문자 무관."""
    v = cell_val.strip()
    # "NAME (PRODUCT)" 파싱
    if "(" in v and v.endswith(")"):
        name_part = v[:v.rfind("(")].strip().upper()
        prod_part = v[v.rfind("(")+1:-1].strip().upper()
        for b in buyers:
            if b["name"].upper() == name_part and b["product"].upper() == prod_part:
                return b["id"]
    # 이름만으로 fallback — 정확히 일치하는 것을 우선, 그래도 없으면 가장 긴(구체적인) 이름 매칭
    vu = v.upper()
    for b in buyers:
        if b["name"].upper() == vu:
            return b["id"]
    _sub_matches = [b for b in buyers if b["name"].upper() in vu]
    if _sub_matches:
        return max(_sub_matches, key=lambda b: len(b["name"]))["id"]
    return None

def _to_float(s):
    """시트 셀 문자열 → float. 천 단위 쉼표 허용. 빈 문자열/None → 0.0."""
    if not s:
        return 0.0
    return float(str(s).replace(",", "").strip())

def _sync_from_gsheets(cfg_ref):
    """Google Sheets 3개 탭 → config 동기화 (덮어쓰기).
    반환: (성공 여부, 메시지 문자열)
    """
    try:
        gc = _gsheet_connect()
        sh = gc.open(_GSHEET_NAME)
    except Exception as e:
        return False, f"연결 실패: {e}"

    buyers    = cfg_ref.get("buyers", [])
    scrap_map = {s["name"]: s["id"] for s in cfg_ref.get("scrap_types", [])}
    proc_map  = {p["name"].lower(): p["id"] for p in cfg_ref.get("processors", [])}
    log = []

    # ── ① 선적 탭 ────────────────────────────────────────────────────────────
    try:
        rows = sh.worksheet("선적").get_all_values()
        if len(rows) > 1:
            # BUG FIX: 빈 HBL 선적건이 같은 키("")로 충돌하지 않도록 제외
            hbl_idx = {s["hbl"].strip(): i for i, s in enumerate(cfg_ref.get("shipments", []))
                       if s.get("hbl","").strip()}
            added, updated = 0, 0
            for row in rows[1:]:
                # 열 수 보정
                row = [c.strip().replace("\r","") for c in row] + [""] * 10
                hbl, inv_no, ld, buyer_str, wkg, iusd, pm, fm, status, etd = row[:10]
                # HBL 공란(발급 전) 행도 처리 — 선적일+매입사+중량 복합키로 매핑
                if not ld:   # 선적일도 없으면 의미없는 빈 행
                    continue
                buyer_id = _match_buyer_id(buyer_str, buyers)
                entry = {
                    "hbl":         hbl,
                    "invoice_no":  inv_no,
                    "loading_date": ld,
                    "etd":         etd,
                    "buyer_id":    buyer_id or "",
                    "weight_kg":   _to_float(wkg),
                    "invoice_usd": _to_float(iusd),
                    "prov_month":  pm or "—",
                    "final_month": fm or "—",
                    "status":      status or "provisional",
                }
                if hbl and hbl in hbl_idx:
                    # HBL 있고 기존 항목 존재 → 업데이트
                    cfg_ref["shipments"][hbl_idx[hbl]].update(entry)
                    updated += 1
                else:
                    # HBL 공란이거나, HBL이 새로 채워졌는데 기존엔 공란이었던 경우
                    # → 선적일+buyer_id+중량 복합키로 기존 항목(공란 HBL) 탐색
                    _match_idx = None
                    _wkg_f = _to_float(wkg)
                    for _ci, _cs in enumerate(cfg_ref.get("shipments", [])):
                        if (not _cs.get("hbl","").strip()
                                and _cs.get("loading_date","") == ld
                                and _cs.get("buyer_id","") == (buyer_id or "")
                                and abs(float(_cs.get("weight_kg",0)) - _wkg_f) < 1):
                            _match_idx = _ci
                            break
                    if _match_idx is not None:
                        cfg_ref["shipments"][_match_idx].update(entry)
                        if hbl:
                            hbl_idx[hbl] = _match_idx
                        updated += 1
                    else:
                        entry.update({
                            "id": str(uuid.uuid4())[:8],
                            "eta": "", "notes": "",
                            "moisture_pct": None, "buyer_ni_content": None,
                            "buyer_co_content": None,
                            "other_adj_usd": None, "other_adj_desc": "",
                        })
                        cfg_ref.setdefault("shipments", []).append(entry)
                        if hbl:
                            hbl_idx[hbl] = len(cfg_ref["shipments"]) - 1
                        added += 1
            log.append(f"선적: 추가 {added}건 / 업데이트 {updated}건")
    except Exception as e:
        log.append(f"선적 탭 오류: {e}")

    # ── ② 입고 탭 ────────────────────────────────────────────────────────────
    try:
        rows = sh.worksheet("입고").get_all_values()
        if len(rows) > 1:
            new_purchases = {}  # scrap_id → [purchase list]
            _pur_skip = set()
            for row in rows[1:]:
                row = [c.strip() for c in row] + [""] * 6
                sc_nm, dt, qty, tb, price, notes = row[:6]
                if not sc_nm or not qty:
                    continue
                scid = scrap_map.get(sc_nm)
                if not scid:
                    _pur_skip.add(sc_nm)
                    continue
                new_purchases.setdefault(scid, []).append({
                    "date":        dt,
                    "quantity_kg": _to_float(qty),
                    "ton_bags":    int(_to_float(tb)),
                    "unit_cost":   _to_float(price),
                    "notes":       notes,
                })
            cnt = 0
            for scid, plist in new_purchases.items():
                if scid not in cfg_ref.get("raw_material_inventory", {}):
                    cfg_ref.setdefault("raw_material_inventory", {})[scid] = {
                        "opening": None, "purchases": []
                    }
                cfg_ref["raw_material_inventory"][scid]["purchases"] = plist
                cnt += len(plist)
            _pur_msg = f"입고: {cnt}건 동기화"
            if _pur_skip:
                _pur_msg += f" ⚠️ 스크랩명 미매핑 스킵: {', '.join(sorted(_pur_skip))}"
            log.append(_pur_msg)
    except Exception as e:
        log.append(f"입고 탭 오류: {e}")

    # ── ③ 출고 탭 ────────────────────────────────────────────────────────────
    try:
        rows = sh.worksheet("출고").get_all_values()
        if len(rows) > 1:
            new_dr, new_ds = [], []
            seen_dr_sc, seen_ds_sc = set(), set()
            _out_skip_sc, _out_skip_proc = set(), set()
            for row in rows[1:]:
                row = [c.strip() for c in row] + [""] * 8
                otype, dt, sc_nm, proc_nm, qty, tb, notes = row[:7]
                if not otype or not qty:
                    continue
                scid  = scrap_map.get(sc_nm)
                if not scid:
                    if sc_nm:
                        _out_skip_sc.add(sc_nm)
                    continue
                qty_f = _to_float(qty)
                tb_i  = int(_to_float(tb))
                if otype == "임가공출고":
                    pid = proc_map.get(proc_nm.lower())
                    if not pid:
                        if proc_nm:
                            _out_skip_proc.add(proc_nm)
                        continue
                    seen_dr_sc.add(scid)
                    new_dr.append({
                        "id": str(uuid.uuid4())[:8],
                        "date": dt, "processor_id": pid,
                        "scrap_type_id": scid,
                        "quantity_kg": qty_f, "ton_bags": tb_i, "notes": notes,
                    })
                elif otype == "직접판매":
                    seen_ds_sc.add(scid)
                    new_ds.append({
                        "id": str(uuid.uuid4())[:8],
                        "date": dt, "scrap_type_id": scid,
                        "quantity_kg": qty_f, "ton_bags": tb_i,
                        "sale_price_per_kg": None, "notes": notes,
                    })
            # 시트에 나온 조합만 교체, 나머지는 보존
            cfg_ref["dispatch_records"] = [
                r for r in cfg_ref.get("dispatch_records", [])
                if r.get("scrap_type_id") not in seen_dr_sc
            ] + new_dr
            cfg_ref["direct_sales"] = [
                r for r in cfg_ref.get("direct_sales", [])
                if r.get("scrap_type_id") not in seen_ds_sc
            ] + new_ds
            _out_msg = f"출고: 임가공 {len(new_dr)}건 / 직접판매 {len(new_ds)}건 동기화"
            if _out_skip_sc:
                _out_msg += f" ⚠️ 스크랩명 미매핑: {', '.join(sorted(_out_skip_sc))}"
            if _out_skip_proc:
                _out_msg += f" ⚠️ 임가공사명 미매핑: {', '.join(sorted(_out_skip_proc))}"
            log.append(_out_msg)
    except Exception as e:
        log.append(f"출고 탭 오류: {e}")

    return True, "\n".join(log)


# TAB 11 — INDEX 이력
# ══════════════════════════════════════════════════════════════════════════════
with t_idx:
    # ── Google Sheets 동기화 ─────────────────────────────────────────────────
    st.subheader("🔄 Google Sheets 동기화")
    if READ_ONLY:
        st.caption(f"시트: **{_GSHEET_NAME}**  |  탭: 선적 / 입고 / 출고")
    else:
        try:
            _svc_email = _get_gcp_creds(_GSHEET_SCOPES).service_account_email
        except Exception:
            _svc_email = "—"
        st.caption(
            f"시트: **{_GSHEET_NAME}**  |  탭: 선적 / 입고 / 출고  |  "
            f"서비스 계정: `{_svc_email}`"
        )

    _gs_c1, _gs_c2 = st.columns([2, 3])
    with _gs_c1:
        if st.button("🔄 지금 동기화", type="primary", use_container_width=True):
            with st.spinner("Google Sheets에서 데이터 가져오는 중..."):
                _ok, _msg = _sync_from_gsheets(cfg)
            if _ok:
                save_cfg(cfg)
                from datetime import datetime as _dt
                _now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state["last_sync_time"] = _now_str
                st.session_state["last_sync_log"] = _msg
                st.toast("✅ 동기화 완료")
                st.rerun()
            else:
                st.error(f"동기화 실패: {_msg}")
        _last_sync = st.session_state.get("last_sync_time")
        if _last_sync:
            st.caption(f"마지막 동기화: {_last_sync}")
        _last_log = st.session_state.get("last_sync_log")
        if _last_log:
            with st.expander("동기화 결과 상세", expanded=False):
                for _ll in _last_log.split("\n"):
                    st.caption(_ll)
    with _gs_c2:
        st.info(
            "**동기화 범위**  \n"
            "- 선적: HBL 기준 upsert (정산 상세·수분 등은 보존)  \n"
            "- 입고: 스크랩 유형별 구매 이력 전체 교체 (기초재고 보존)  \n"
            "- 출고: (출고유형, 스크랩유형) 조합 단위 교체"
        )

    st.divider()
    st.subheader("월별 INDEX 이력")
    history=cfg.get("index_history",[])
    if history:
        df_h=pd.DataFrame(sorted(history,key=lambda x:x["month"],reverse=True))
        df_h.columns=["기준월","Ni INDEX($/ton)","Co INDEX($/ton)"]
        st.dataframe(df_h.style.format({"Ni INDEX($/ton)":"${:,.2f}","Co INDEX($/ton)":"${:,.2f}"}),
                     use_container_width=True,hide_index=True)
        _idx_del_c1, _idx_del_c2 = st.columns([4, 1])
        with _idx_del_c1:
            dm=st.selectbox("삭제할 월",[h["month"] for h in sorted(history,key=lambda x:x["month"],reverse=True)])
        with _idx_del_c2:
            st.markdown("&nbsp;", unsafe_allow_html=True)
            with st.popover("🗑️", use_container_width=True):
                st.warning(f"INDEX **{dm}** 삭제")
                if st.button("삭제 확인", key="idx_del_cfm", type="primary", use_container_width=True):
                    cfg["index_history"]=[h for h in history if h["month"]!=dm]
                    save_cfg(cfg); st.toast(f"✅ {dm} 삭제"); st.rerun()
    else: st.info("저장된 INDEX 이력이 없습니다.")
    st.divider()
    st.subheader("새 INDEX 추가 / 수정")
    with st.form("add_idx"):
        i1,i2,i3=st.columns(3)
        with i1: im =st.text_input("기준월 (YYYY-MM)",placeholder="2026-04")
        with i2: ini=st.number_input("Ni INDEX($/ton)",value=17093.18,step=10.0,format="%.2f")
        with i3: ico=st.number_input("Co INDEX($/ton)",value=56598.72,step=10.0,format="%.2f")
        if st.form_submit_button("💾 저장"):
            try: datetime.strptime(im,"%Y-%m")
            except: st.error("YYYY-MM 형식으로 입력하세요.")
            else:
                rest=[h for h in cfg["index_history"] if h["month"]!=im]
                rest.append({"month":im,"ni_index":ini,"co_index":ico})
                cfg["index_history"]=sorted(rest,key=lambda x:x["month"])
                save_cfg(cfg); st.success(f"{im} 저장 — Ni ${ini:,.2f} / Co ${ico:,.2f}"); st.rerun()

    # ── EUR/USD 환율 관리 ─────────────────────────────────────────────────────
    st.divider()
    st.subheader("💱 월별 EUR/USD 환율")
    st.caption("scrap보관비 계산에 사용됩니다. (EUR 1.5/톤백/day → USD 자동환산)\n"
               "미등록 월은 직전 월 환율을 적용하며, 등록 환율이 없으면 기본값 1.10을 사용합니다.")

    eur_rates = cfg.get("eur_usd_rates", [])
    if eur_rates:
        _df_eur = pd.DataFrame(sorted(eur_rates, key=lambda x: x["month"], reverse=True))
        _df_eur.columns = ["기준월", "EUR/USD"]
        st.dataframe(
            _df_eur.style.format({"EUR/USD": "{:.4f}"}),
            use_container_width=True, hide_index=True
        )
        _eur_del_c1, _eur_del_c2 = st.columns([4, 1])
        with _eur_del_c1:
            _eur_del_m = st.selectbox(
                "삭제할 월",
                [r["month"] for r in sorted(eur_rates, key=lambda x: x["month"], reverse=True)],
                key="eur_del_sel"
            )
        with _eur_del_c2:
            st.markdown("&nbsp;", unsafe_allow_html=True)
            with st.popover("🗑️", use_container_width=True):
                st.warning(f"EUR/USD **{_eur_del_m}** 삭제")
                if st.button("삭제 확인", key="eur_del_cfm", type="primary", use_container_width=True):
                    cfg["eur_usd_rates"] = [r for r in eur_rates if r["month"] != _eur_del_m]
                    save_cfg(cfg); st.toast(f"✅ {_eur_del_m} 삭제"); st.rerun()
    else:
        st.info("등록된 EUR/USD 환율이 없습니다. 기본값 1.10이 적용됩니다.")

    st.divider()

    # ── 자동 조회 (Frankfurter API) ───────────────────────────────────────────
    st.subheader("🔄 EUR/USD 자동 조회")
    st.caption("Frankfurter.app (무료 API, 유럽중앙은행 기준) — API 키 불필요")

    _af1, _af2, _af3 = st.columns([2, 2, 3])
    with _af1:
        _fetch_months = st.number_input(
            "조회 개월 수 (최근 N개월)",
            value=3, min_value=1, max_value=24, step=1,
            help="현재 월 포함 최근 N개월의 말일 EUR/USD를 일괄 조회합니다."
        )
    with _af2:
        _overwrite = st.checkbox("기존 값 덮어쓰기", value=False,
                                 help="체크 해제 시 이미 등록된 월은 유지합니다.")
    with _af3:
        st.markdown("&nbsp;", unsafe_allow_html=True)
        if st.button("🔄 자동 조회 & 저장", use_container_width=True, type="primary"):
            import requests as _req
            _existing_months = {r["month"] for r in cfg.get("eur_usd_rates", [])}
            _saved, _skipped, _failed = [], [], []
            _today = date.today()
            for _mi in range(int(_fetch_months)):
                # N개월 전부터 이번 달까지 역순 순회
                _y = _today.year
                _m = _today.month - _mi
                while _m <= 0:
                    _m += 12; _y -= 1
                _month_str = f"{_y}-{_m:02d}"
                if not _overwrite and _month_str in _existing_months:
                    _skipped.append(_month_str)
                    continue
                # 해당 월 마지막 날 환율 조회
                import calendar
                _last_day = calendar.monthrange(_y, _m)[1]
                # 미래 월이면 오늘 날짜 기준
                _fetch_date = min(date(_y, _m, _last_day), _today).isoformat()
                try:
                    _resp = _req.get(
                        f"https://api.frankfurter.app/{_fetch_date}?from=EUR&to=USD",
                        timeout=8
                    )
                    if _resp.status_code == 200:
                        _rate_val = _resp.json()["rates"]["USD"]
                        _actual_date = _resp.json()["date"][:7]  # 실제 데이터 월
                        # 말일이 주말이면 직전 영업일 데이터 반환 → 해당 월로 저장
                        _rest2 = [r for r in cfg.get("eur_usd_rates",[]) if r["month"] != _month_str]
                        _rest2.append({"month": _month_str, "rate": round(_rate_val, 4)})
                        cfg["eur_usd_rates"] = sorted(_rest2, key=lambda x: x["month"])
                        _saved.append(f"{_month_str}: {_rate_val:.4f} (기준일 {_actual_date})")
                    else:
                        _failed.append(f"{_month_str} (HTTP {_resp.status_code})")
                except Exception as _fe:
                    _failed.append(f"{_month_str} ({_fe})")
            if _saved:
                save_cfg(cfg)
                st.success(f"✅ {len(_saved)}개월 저장 완료\n" + "\n".join(_saved))
            if _skipped:
                st.info(f"⏭️ 기존값 유지 {len(_skipped)}개월: {', '.join(_skipped)}")
            if _failed:
                st.error(f"❌ 조회 실패: {', '.join(_failed)}")
            if _saved:
                st.rerun()

    st.divider()
    st.subheader("환율 수동 추가 / 수정")
    with st.form("add_eur"):
        _ec1, _ec2 = st.columns(2)
        with _ec1: _eur_m = st.text_input("기준월 (YYYY-MM)", placeholder="2026-05")
        with _ec2: _eur_r = st.number_input("EUR/USD", value=1.10, step=0.0001, format="%.4f")
        if st.form_submit_button("💾 저장"):
            try: datetime.strptime(_eur_m, "%Y-%m")
            except: st.error("YYYY-MM 형식으로 입력하세요.")
            else:
                _eur_rest = [r for r in cfg.get("eur_usd_rates",[]) if r["month"] != _eur_m]
                _eur_rest.append({"month": _eur_m, "rate": _eur_r})
                cfg["eur_usd_rates"] = sorted(_eur_rest, key=lambda x: x["month"])
                save_cfg(cfg); st.success(f"{_eur_m} 저장 — EUR/USD {_eur_r:.4f}"); st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Excel 보고서 생성 함수
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(cfg, ni, co, xr, ref_month, sel_buyer_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "요약 보고서"

    # ── 스타일 헬퍼 ──
    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def cell(r, c, val=None, bold=False, size=10, color="000000",
             bg=None, align="center", fmt=None, wrap=False):
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = Font(name="Arial", size=size, bold=bold, color=color)
        if bg: cl.fill = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cl.border = thin_border()
        if fmt: cl.number_format = fmt
        return cl

    def merge_cell(r, c1, c2, val, bold=True, size=11, color="FFFFFF",
                   bg="1F4E79", align="left"):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cl = ws.cell(row=r, column=c1, value=val)
        cl.font = Font(name="Arial", size=size, bold=bold, color=color)
        cl.fill = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal=align, vertical="center", indent=1)
        ws.row_dimensions[r].height = 22
        return cl

    COLS = 10
    ROW = 1

    # ── 타이틀 ──
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=COLS)
    t = ws.cell(row=ROW, column=1, value="BP / BM 단가 요약 보고서")
    t.font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ROW].height = 40
    ROW += 1

    # ── 기준 정보 ──
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=COLS)
    info = (f"기준월: {ref_month}   |   Ni INDEX: ${ni:,.2f}/ton   |   "
            f"Co INDEX: ${co:,.2f}/ton   |   환율: {xr:,.0f} KRW/USD   |   작성일: {date.today()}")
    t2 = ws.cell(row=ROW, column=1, value=info)
    t2.font = Font(name="Arial", size=10, color="FFFFFF")
    t2.fill = PatternFill("solid", fgColor="2E75B6")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ROW].height = 20
    ROW += 2

    # ════ SECTION 1: BP/BM 매각 단가 ════
    merge_cell(ROW, 1, 6, "  ① BP/BM 매각 단가 현황", bg="2E75B6")
    ROW += 1
    for ci, h in enumerate(["매입사","품목","Ni 지불율","Co 지불율","단가 ($/kg)","단가 (원/kg)"], 1):
        cell(ROW, ci, h, bold=True, color="FFFFFF", bg="4472C4", size=10)
    ws.row_dimensions[ROW].height = 18
    ROW += 1

    ab = [b for b in cfg["buyers"] if b.get("active", True)]
    for b in ab:
        _, _, tot, pkg = bp_price(ni, co, b["ni_content"], b["co_content"], b["ni_payable"], b["co_payable"])
        bg = "D6E4F0" if b["product"] == "BP" else "D5F5E3"
        cell(ROW, 1, b["name"],        bg=bg, align="left")
        cell(ROW, 2, b["product"],     bg=bg)
        cell(ROW, 3, b["ni_payable"],  bg=bg, fmt="0.00")
        cell(ROW, 4, b["co_payable"],  bg=bg, fmt="0.00")
        cell(ROW, 5, round(pkg, 5),    bg=bg, fmt='"$"#,##0.00000')
        cell(ROW, 6, round(pkg*xr, 0), bg=bg, fmt='"₩"#,##0')
        ROW += 1
    ROW += 1

    # ════ SECTION 2: 원가·마진 매트릭스 ════
    ap  = [p for p in cfg.get("processors", []) if p.get("active", True)]
    as_ = [s for s in cfg.get("scrap_types", []) if s.get("active", True)]
    sel_b = next((b for b in ab if b["id"] == sel_buyer_id), ab[0] if ab else None)

    if ap and as_:
        n_proc = len(ap)
        merge_cell(ROW, 1, 1+n_proc, "  ② 원가 매트릭스  — 재매입 순원가 ($/kg, 임가공비 ÷ 전환율)", bg="2E75B6")
        ROW += 1
        cell(ROW, 1, "스크랩 유형", bold=True, color="FFFFFF", bg="4472C4")
        for ci, p in enumerate(ap, 2):
            cell(ROW, ci, p["name"], bold=True, color="FFFFFF", bg="4472C4")
        ws.row_dimensions[ROW].height = 18
        ROW += 1
        for scrap in as_:
            cell(ROW, 1, scrap["name"], bold=True, bg="F0F2F6", align="left")
            for ci, proc in enumerate(ap, 2):
                cond = proc.get("conditions", {}).get(scrap["id"], {})
                pf = cond.get("processing_fee"); conv = cond.get("conversion_rate")
                bmc = round(pf / (conv / 100), 4) if (pf is not None and conv and conv > 0) else None
                if bmc is not None:
                    cell(ROW, ci, bmc, fmt='"$"#,##0.0000', bg="FDEBD0")
                else:
                    cell(ROW, ci, "—", color="AAAAAA")
            ROW += 1
        ROW += 1

        if sel_b:
            _, _, _, sell_pkg = bp_price(ni, co, sel_b["ni_content"], sel_b["co_content"],
                                         sel_b["ni_payable"], sel_b["co_payable"])
            lbl = (f"  ③ 마진 매트릭스  ($/kg)  —  "
                   f"매각: {sel_b['name']} ({sel_b['product']}) ${sell_pkg:.4f}/kg 기준")
            merge_cell(ROW, 1, 1+n_proc, lbl, bg="2E75B6")
            ROW += 1
            cell(ROW, 1, "스크랩 유형", bold=True, color="FFFFFF", bg="4472C4")
            for ci, p in enumerate(ap, 2):
                cell(ROW, ci, p["name"], bold=True, color="FFFFFF", bg="4472C4")
            ws.row_dimensions[ROW].height = 18
            ROW += 1
            best_margin = -999; best_combo = ""
            for scrap in as_:
                cell(ROW, 1, scrap["name"], bold=True, bg="F0F2F6", align="left")
                for ci, proc in enumerate(ap, 2):
                    cond = proc.get("conditions", {}).get(scrap["id"], {})
                    pf = cond.get("processing_fee"); conv = cond.get("conversion_rate")
                    bmc = round(pf/(conv/100),4) if (pf is not None and conv and conv>0) else None
                    if bmc is not None:
                        margin = round(sell_pkg - bmc, 4)
                        bg_c = "D5F5E3" if margin >= 0 else "FADBD8"
                        cell(ROW, ci, margin, fmt='"$"#,##0.0000;[Red]"-$"#,##0.0000', bg=bg_c)
                        if margin > best_margin:
                            best_margin = margin
                            best_combo = f"{scrap['name']} × {proc['name']}"
                    else:
                        cell(ROW, ci, "—", color="AAAAAA")
                ROW += 1
            ROW += 1
            if best_combo:
                ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=1+n_proc)
                bc = ws.cell(row=ROW, column=1,
                             value=f"★  최고 마진:  {best_combo}  →  ${best_margin:+.4f}/kg  (₩{best_margin*xr:+,.0f}/kg)")
                bc.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
                bc.fill = PatternFill("solid", fgColor="FEF9E7")
                bc.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[ROW].height = 22
            ROW += 2

    # ════ SECTION 4: 월별 손익 요약 ════
    ph_all_xl = cfg.get("processing_history", [])
    ship_map_xl = {s["id"]: s for s in cfg.get("shipments", [])}
    pnl_mo = defaultdict(lambda: {"bp":0.0,"pf":0.0,"eu":0.0,"out":0.0,"cnt":0})
    for r in ph_all_xl:
        sh = ship_map_xl.get(r.get("shipment_id",""), {})
        mo = (sh.get("loading_date") or "미연결")[:7]
        out = float(r.get("output_kg",0) or 0)
        inp = _ph_input_kg(r)   # 일관성: 역산 로직 통일
        pnl_mo[mo]["bp"]  += float(r.get("bp_sale_per_kg",0) or 0) * out
        pnl_mo[mo]["pf"]  += float(r.get("processing_fee_per_kg",0) or 0) * inp
        pnl_mo[mo]["eu"]  += _ph_export_usd(r, cfg)   # BUG FIX: HBL 레벨 수출비 포함
        pnl_mo[mo]["out"] += out
        pnl_mo[mo]["cnt"] += 1

    if pnl_mo:
        merge_cell(ROW, 1, 8, "  ④ 월별 손익 요약  (거래 마진 = BP 매각 − 임가공비 − 수출비, 원료비·보관비 제외)", bg="2E75B6")
        ROW += 1
        pnl_hdrs = ["월","배치수","생산(kg)","BP 매각(USD)","임가공비(USD)","수출비(USD)","거래 마진(USD)","누적 거래 마진(USD)"]
        for ci, h in enumerate(pnl_hdrs, 1):
            cell(ROW, ci, h, bold=True, color="FFFFFF", bg="4472C4", size=9)
        ws.row_dimensions[ROW].height = 18
        ROW += 1
        cum = 0.0
        for mo in sorted(pnl_mo.keys()):
            v = pnl_mo[mo]
            net = v["bp"] - v["pf"] - v["eu"]
            cum += net
            net_bg = "D5F5E3" if net >= 0 else "FADBD8"
            cum_bg = "D5F5E3" if cum >= 0 else "FADBD8"
            cell(ROW, 1, mo, align="left")
            cell(ROW, 2, v["cnt"])
            cell(ROW, 3, round(v["out"],0), fmt="#,##0")
            cell(ROW, 4, round(v["bp"],2),  fmt='"$"#,##0.00')
            cell(ROW, 5, round(v["pf"],2),  fmt='"$"#,##0.00')
            cell(ROW, 6, round(v["eu"],2),  fmt='"$"#,##0.00')
            cell(ROW, 7, round(net,2),       fmt='"$"#,##0.00;[Red]"-$"#,##0.00', bg=net_bg, bold=True)
            cell(ROW, 8, round(cum,2),       fmt='"$"#,##0.00;[Red]"-$"#,##0.00', bg=cum_bg, bold=True)
            ROW += 1
        # 합계행
        tot_bp  = sum(v["bp"]  for v in pnl_mo.values())
        tot_pf  = sum(v["pf"]  for v in pnl_mo.values())
        tot_eu  = sum(v["eu"]  for v in pnl_mo.values())
        tot_net = tot_bp - tot_pf - tot_eu
        tot_bg  = "D5F5E3" if tot_net >= 0 else "FADBD8"
        for ci, val in enumerate(["합계","","",round(tot_bp,2),round(tot_pf,2),round(tot_eu,2),round(tot_net,2),""], 1):
            kw = {"bold":True, "bg":"F0F2F6"}
            if ci==4: kw["fmt"]='"$"#,##0.00'
            elif ci==5: kw["fmt"]='"$"#,##0.00'
            elif ci==6: kw["fmt"]='"$"#,##0.00'
            elif ci==7: kw["bg"]=tot_bg; kw["fmt"]='"$"#,##0.00;[Red]"-$"#,##0.00'
            cell(ROW, ci, val, **kw)
        ROW += 2

    # ════ SECTION 5: HBL별 손익 상세 ════
    if ph_all_xl:
        merge_cell(ROW, 1, 9, "  ⑤ HBL별 손익 상세", bg="2E75B6")
        ROW += 1
        hbl_hdrs = ["HBL","매입사","선적일","투입(kg)","생산(kg)","전환율(%)","BP 매각","임가공비","수출비","거래 마진"]
        for ci, h in enumerate(hbl_hdrs, 1):
            cell(ROW, ci, h, bold=True, color="FFFFFF", bg="4472C4", size=9)
        ws.row_dimensions[ROW].height = 18
        ROW += 1
        buyer_map_xl = {b["id"]:b for b in cfg["buyers"]}
        hagg = defaultdict(lambda: {"bp":0.0,"pf":0.0,"eu":0.0,"out":0.0,"inp":0.0,"cnt":0})
        for r in ph_all_xl:
            sid = r.get("shipment_id","")
            if not sid: continue
            out = float(r.get("output_kg",0) or 0)
            inp_v = _ph_input_kg(r)   # 일관성: 역산 로직 통일
            hagg[sid]["bp"]  += float(r.get("bp_sale_per_kg",0) or 0)*out
            hagg[sid]["pf"]  += float(r.get("processing_fee_per_kg",0) or 0)*inp_v
            hagg[sid]["eu"]  += _ph_export_usd(r, cfg)   # BUG FIX: HBL 레벨 수출비 포함
            hagg[sid]["out"] += out
            hagg[sid]["inp"] += float(inp_v or 0)
        for sid in sorted(hagg, key=lambda x: ship_map_xl.get(x,{}).get("loading_date","")):
            v = hagg[sid]
            sh = ship_map_xl.get(sid,{})
            byr = buyer_map_xl.get(sh.get("buyer_id",""),{})
            net = v["bp"]-v["pf"]-v["eu"]
            net_bg = "D5F5E3" if net>=0 else "FADBD8"
            conv = round(v["out"]/v["inp"]*100,2) if v["inp"]>0 else None
            cell(ROW,1, sh.get("hbl","—"),                                     align="left")
            cell(ROW,2, f"{byr.get('name','?')} ({byr.get('product','?')})",   align="left")
            cell(ROW,3, sh.get("loading_date","—"))
            cell(ROW,4, round(v["inp"],0), fmt="#,##0")
            cell(ROW,5, round(v["out"],0), fmt="#,##0")
            cell(ROW,6, conv,              fmt='#,##0.00"%"' if conv else None)
            cell(ROW,7, round(v["bp"],2),  fmt='"$"#,##0.00')
            cell(ROW,8, round(v["pf"],2),  fmt='"$"#,##0.00')
            cell(ROW,9, round(v["eu"],2),  fmt='"$"#,##0.00')
            cell(ROW,10,round(net,2),      fmt='"$"#,##0.00;[Red]"-$"#,##0.00', bg=net_bg, bold=True)
            ROW += 1
        ROW += 1

    # ════ SECTION 6: 선적 정산 현황 ════
    merge_cell(ROW, 1, 8, "  ⑥ 선적 정산 현황", bg="2E75B6")
    ROW += 1
    ships = cfg.get("shipments", [])
    buyer_map_l = {b["id"]: b for b in cfg["buyers"]}
    stats = {}
    for s in ships:
        k = s.get("status","provisional"); stats[k] = stats.get(k,0)+1
    total_w = sum(s.get("weight_kg",0) for s in ships)
    total_inv_s = sum(s.get("invoice_usd",0) for s in ships)
    for ci, (lbl, val) in enumerate([
        ("총 선적건", f"{len(ships)}건"),
        ("총 중량", f"{total_w/1000:,.1f} ton"),
        ("Provisional 정산", f"{stats.get('provisional',0)}건"),
        ("최종정산", f"{stats.get('final',0)}건"),
        ("입금완료", f"{stats.get('paid',0)}건"),
        ("총 Invoice", f"${total_inv_s:,.0f}"),
    ], 1):
        cell(ROW,   ci, lbl, bold=True, color="FFFFFF", bg="4472C4", size=9)
        cell(ROW+1, ci, val, size=11, bold=True, bg="F0F2F6")
    ws.row_dimensions[ROW].height = 16; ws.row_dimensions[ROW+1].height = 22
    ROW += 3
    if ships:
        for ci, h in enumerate(["HBL","매입사","선적일","중량(kg)","Invoice(USD)","Prov월","Final월","상태"], 1):
            cell(ROW, ci, h, bold=True, color="FFFFFF", bg="4472C4", size=9)
        ws.row_dimensions[ROW].height = 16; ROW += 1
        for s in sorted(ships, key=lambda x: x.get("loading_date",""), reverse=True):
            b2 = buyer_map_l.get(s.get("buyer_id"),{})
            sc = {"provisional":"FEF9E7","final":"D5F5E3","paid":"D6E4F0"}.get(s.get("status","provisional"),"FFFFFF")
            cell(ROW,1, s.get("hbl","—"),                                     bg=sc, align="left")
            cell(ROW,2, f"{b2.get('name','?')} ({b2.get('product','?')})",    bg=sc, align="left")
            cell(ROW,3, s.get("loading_date","—"),                            bg=sc)
            cell(ROW,4, s.get("weight_kg",0),    fmt="#,##0",                 bg=sc)
            cell(ROW,5, s.get("invoice_usd",0),  fmt='"$"#,##0.00',          bg=sc)
            cell(ROW,6, s.get("prov_month","—"),                              bg=sc)
            cell(ROW,7, s.get("final_month","—"),                             bg=sc)
            cell(ROW,8, {"provisional":"Provisional 정산","final":"최종정산","paid":"입금완료"}.get(s.get("status",""),"—"), bg=sc)
            ROW += 1

    # ── 컬럼 너비 및 동결 ──
    from openpyxl.utils import get_column_letter
    col_widths = [18, 20, 12, 13, 14, 13, 14, 16, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 0 — 요약 보고서
# ══════════════════════════════════════════════════════════════════════════════
with t_report:
    st.subheader("요약 보고서")

    # ── 헤더 바 ──────────────────────────────────────────────────────────────
    ro1, ro2 = st.columns([2, 1])
    with ro1:
        _rpt_month_opts = hist_opts if hist_opts else [f"{date.today().year}-{date.today().month:02d}"]
        rpt_month = st.selectbox("기준월", _rpt_month_opts, key="rpt_month_sel")
        st.markdown(f"**작성일**: `{date.today()}`  |  **Ni**: `${NI:,.2f}/t`  **Co**: `${CO:,.2f}/t`  **KRW**: `{XR:,.0f}`")
    with ro2:
        if active_buyers:
            xl_bytes = generate_excel_report(cfg, NI, CO, XR, rpt_month,
                                             active_buyers[0]["id"])
            st.download_button("📥 Excel 보고서 다운로드", data=xl_bytes,
                               file_name=f"BP_BM_요약보고서_{rpt_month}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    st.divider()

    # ── 공통 데이터 준비 ──────────────────────────────────────────────────────
    _rpt_ships    = cfg.get("shipments", [])
    _rpt_ph_all   = cfg.get("processing_history", [])
    _rpt_buyer_m0 = {b["id"]: b for b in cfg["buyers"]}
    _rpt_ship_m0  = {s["id"]: s for s in _rpt_ships}

    # ── 연간 누계 요약 ────────────────────────────────────────────────────────
    _cur_year = date.today().year
    st.markdown(f"#### 📅 {_cur_year}년 누계")
    _yr_ph    = [r for r in _rpt_ph_all
                 if (_rpt_ship_m0.get(r.get("shipment_id",""),{}).get("loading_date","") or "")[:4] == str(_cur_year)]
    _yr_ships = [s for s in _rpt_ships if (s.get("loading_date","") or "")[:4] == str(_cur_year)]
    _yr_bp    = sum(float(r.get("bp_sale_per_kg",0) or 0) * float(r.get("output_kg",0) or 0) for r in _yr_ph)
    _yr_pf    = sum(float(r.get("processing_fee_per_kg",0) or 0) * _ph_input_kg(r) for r in _yr_ph)
    _yr_eu    = sum(_ph_export_usd(r, cfg) for r in _yr_ph)
    _yr_out   = sum(float(r.get("output_kg",0) or 0) for r in _yr_ph)
    _yr_inv   = sum(float(s.get("invoice_usd",0) or 0) for s in _yr_ships)
    _yr_net   = _yr_bp - _yr_pf - _yr_eu
    _ya1,_ya2,_ya3,_ya4,_ya5 = st.columns(5)
    _ya1.metric("선적 건수",    f"{len(_yr_ships)}건")
    _ya2.metric("선적 중량",    f"{sum(float(s.get('weight_kg',0)) for s in _yr_ships):,.0f} kg")
    _ya3.metric("BP 생산",      f"{_yr_out:,.0f} kg")
    _ya4.metric("Invoice 합계", f"${_yr_inv:,.0f}")
    _ya5.metric("거래 마진",     f"${_yr_net:+,.0f}",
                delta_color="normal" if _yr_net >= 0 else "inverse")

    # ── 월별 거래 마진 추이 ──────────────────────────────────────────────────
    _mo_agg = defaultdict(lambda: {"bp": 0.0, "pf": 0.0, "eu": 0.0})
    for r in _yr_ph:
        _mo_key = (_rpt_ship_m0.get(r.get("shipment_id",""), {}).get("loading_date","") or "")[:7]
        if not _mo_key:
            continue
        _mo_agg[_mo_key]["bp"] += float(r.get("bp_sale_per_kg",0) or 0) * float(r.get("output_kg",0) or 0)
        _mo_agg[_mo_key]["pf"] += float(r.get("processing_fee_per_kg",0) or 0) * _ph_input_kg(r)
        _mo_agg[_mo_key]["eu"] += _ph_export_usd(r, cfg)
    if _mo_agg:
        _mo_rows = []
        for _mk in sorted(_mo_agg.keys()):
            _mv = _mo_agg[_mk]
            _mo_net = _mv["bp"] - _mv["pf"] - _mv["eu"]
            _mo_rows.append({"월": _mk, "거래 마진": round(_mo_net, 2),
                              "구분": "흑자" if _mo_net >= 0 else "적자"})
        try:
            import plotly.express as px
            _df_mo = pd.DataFrame(_mo_rows)
            _fig_mo = px.bar(
                _df_mo, x="월", y="거래 마진", color="구분", text="거래 마진",
                color_discrete_map={"흑자": "#2ECC71", "적자": "#E74C3C"},
            )
            _fig_mo.update_traces(texttemplate="$%{y:,.0f}", textposition="outside")
            _fig_mo.add_hline(y=0, line_color="rgba(255,255,255,0.3)", line_width=1)
            _fig_mo.update_layout(
                height=320, showlegend=False,
                margin=dict(l=10, r=10, t=30, b=10),
                plot_bgcolor="#1E1E2E", paper_bgcolor="#16213E",
                font=dict(color="#D0D0E8"),
                xaxis_title="", yaxis_title="",
            )
            _fig_mo.update_yaxes(gridcolor="rgba(255,255,255,0.08)")
            st.markdown(f"#### 📈 {_cur_year}년 월별 거래 마진 추이")
            st.plotly_chart(_fig_mo, use_container_width=True)
        except ImportError:
            pass

    st.divider()

    # ── 재고 회전 계획 ────────────────────────────────────────────────────────
    st.markdown("#### 📋 재고 회전 계획")
    st.caption("보유일수 목표를 설정해 즉시·단기 배출 필요량을 파악합니다. FIFO Lot 기준으로 계산됩니다.")

    _inv_col1, _inv_col2 = st.columns([1, 3])
    with _inv_col1:
        _target_days = st.number_input(
            "목표 보유일수", min_value=1, max_value=365, value=30, step=1,
            help="입고 후 이 기간 내에 배출하는 것을 목표로 합니다."
        )
        _urgent_days = st.number_input(
            "긴급 기준 (잔여일)", min_value=1, max_value=30, value=7, step=1,
            help="목표 기한까지 이 일수 이하면 긴급으로 표시합니다."
        )

    _inv_active_sc = [s for s in cfg.get("scrap_types", []) if s.get("active", True)]
    _today_inv = date.today()

    # 스크랩별 FIFO 잔량 분석
    _plan_summary = []   # 전체 요약용
    _plan_detail  = {}   # 스크랩별 상세

    for _sc_p in _inv_active_sc:
        _scid_p = _sc_p["id"]
        try:
            _, _, _rem_p = _fifo_lot_trace(cfg, _scid_p)
        except Exception:
            continue

        _lots_active = [l for l in _rem_p if l.get("remain", 0) > 0.1]
        if not _lots_active:
            continue

        _rows_p = []
        _overdue_qty = _urgent_qty = _ok_qty = 0.0

        for _lot in _lots_active:
            _lot_date_str = (_lot.get("date") or "")[:10]
            _lot_qty      = float(_lot.get("remain", 0))
            try:
                _lot_dt   = datetime.strptime(_lot_date_str, "%Y-%m-%d").date()
                _elapsed  = (_today_inv - _lot_dt).days
                _deadline = _lot_dt + __import__("datetime").timedelta(days=int(_target_days))
                _remain_d = (_deadline - _today_inv).days
            except Exception:
                _elapsed = _remain_d = None
                _deadline = None

            if _remain_d is not None and _remain_d < 0:
                _status = "🔴 초과"
                _overdue_qty += _lot_qty
            elif _remain_d is not None and _remain_d <= _urgent_days:
                _status = "🟡 긴급"
                _urgent_qty += _lot_qty
            else:
                _status = "🟢 정상"
                _ok_qty += _lot_qty

            _rows_p.append({
                "Lot":          _lot.get("label", _lot_date_str),
                "입고일":       _lot_date_str,
                "잔량(kg)":     round(_lot_qty, 0),
                "경과일":       _elapsed,
                "목표기한":     _deadline.isoformat() if _deadline else "—",
                "잔여일":       _remain_d,
                "상태":         _status,
                "원료단가($/kg)": _lot.get("unit_cost"),
            })

        _plan_detail[_sc_p["name"]] = _rows_p
        _plan_summary.append({
            "스크랩":         _sc_p["name"],
            "🔴 즉시배출(kg)": round(_overdue_qty, 0),
            "🟡 긴급배출(kg)": round(_urgent_qty, 0),
            "🟢 정상(kg)":    round(_ok_qty, 0),
            "총 잔량(kg)":    round(_overdue_qty + _urgent_qty + _ok_qty, 0),
        })

    if _plan_summary:
        # 요약 테이블
        _df_plan_sum = pd.DataFrame(_plan_summary)
        _total_overdue = _df_plan_sum["🔴 즉시배출(kg)"].sum()
        _total_urgent  = _df_plan_sum["🟡 긴급배출(kg)"].sum()
        _total_ok      = _df_plan_sum["🟢 정상(kg)"].sum()

        _sp1, _sp2, _sp3, _sp4 = st.columns(4)
        _sp1.metric("🔴 즉시 배출 필요", f"{_total_overdue:,.0f} kg",
                    help=f"목표 {_target_days}일 이미 초과한 Lot")
        _sp2.metric(f"🟡 {_urgent_days}일 내 배출", f"{_total_urgent:,.0f} kg",
                    help=f"목표 기한까지 {_urgent_days}일 이하")
        _sp3.metric("🟢 정상", f"{_total_ok:,.0f} kg",
                    help=f"목표 {_target_days}일 내 여유 있음")
        _sp4.metric("총 잔량", f"{_total_overdue+_total_urgent+_total_ok:,.0f} kg")

        def _style_plan_sum(row):
            styles = [""] * len(row)
            cols = list(row.index)
            if "🔴 즉시배출(kg)" in cols and row["🔴 즉시배출(kg)"] > 0:
                styles[cols.index("🔴 즉시배출(kg)")] = "color:#c0392b;font-weight:700"
            if "🟡 긴급배출(kg)" in cols and row["🟡 긴급배출(kg)"] > 0:
                styles[cols.index("🟡 긴급배출(kg)")] = "color:#d35400;font-weight:700"
            return styles

        st.dataframe(
            _df_plan_sum.style.apply(_style_plan_sum, axis=1).format({
                "🔴 즉시배출(kg)": "{:,.0f}",
                "🟡 긴급배출(kg)": "{:,.0f}",
                "🟢 정상(kg)":    "{:,.0f}",
                "총 잔량(kg)":    "{:,.0f}",
            }),
            use_container_width=True, hide_index=True,
        )

        # 스크랩별 Lot 상세
        for _sc_nm_p, _rows_p2 in _plan_detail.items():
            with st.expander(f"🔍 {_sc_nm_p} — Lot별 상세", expanded=False):
                _df_p2 = pd.DataFrame(_rows_p2)

                def _style_lot_row(row):
                    st_val = row.get("상태","")
                    if "🔴" in st_val:
                        return ["color:#c0392b;font-weight:600" if c in ("상태","잔여일","잔량(kg)") else "" for c in row.index]
                    if "🟡" in st_val:
                        return ["color:#d35400;font-weight:600" if c in ("상태","잔여일","잔량(kg)") else "" for c in row.index]
                    return [""] * len(row)

                st.dataframe(
                    _df_p2.style.apply(_style_lot_row, axis=1).format(na_rep="—", formatter={
                        "잔량(kg)":        "{:,.0f}",
                        "경과일":          lambda v: f"{v}일" if v is not None else "—",
                        "잔여일":          lambda v: f"{v:+d}일" if v is not None else "—",
                        "원료단가($/kg)":  lambda v: f"${v:.5f}" if v else "—",
                    }),
                    use_container_width=True, hide_index=True,
                )
                _overdue_r = sum(r["잔량(kg)"] for r in _rows_p2 if "🔴" in r["상태"])
                _urgent_r  = sum(r["잔량(kg)"] for r in _rows_p2 if "🟡" in r["상태"])
                if _overdue_r > 0 or _urgent_r > 0:
                    st.warning(
                        f"즉시 배출 필요: **{_overdue_r:,.0f} kg** (목표 초과)"
                        + (f"  |  {_urgent_days}일 내 배출: **{_urgent_r:,.0f} kg**" if _urgent_r > 0 else "")
                    )
    else:
        st.info("FIFO Lot 정보가 없습니다. 입출고 기록 탭에서 기초재고와 입고 이력을 입력하세요.")

    st.divider()

    # ── 공통 데이터 재사용 ───────────────────────────────────────────────────
    _rpt_buyer_m  = {b["id"]: b for b in cfg["buyers"]}
    _rpt_ship_m   = {s["id"]: s for s in _rpt_ships}
    _rpt_sc_m     = {s["id"]: s for s in cfg.get("scrap_types", [])}
    _rpt_proc_m   = {p["id"]: p for p in cfg.get("processors", [])}

    # ── FIFO 원가·보관비 사전 계산 (report 탭 전용) ──────────────────────────
    _rpt_fifo_rmc  = {}  # (sc_id, ship_id) → $/kg
    _rpt_fifo_stor = {}  # (sc_id, ship_id) → USD
    _rpt_fifo_sc_ids = {
        r.get("scrap_type_id","") for r in _rpt_ph_all
        if (r.get("scrap_type_id")
            and cfg.get("raw_material_inventory",{}).get(r.get("scrap_type_id",""),{}).get("opening")
            and any(dr.get("scrap_type_id") == r.get("scrap_type_id","")
                    for dr in cfg.get("dispatch_records",[])))
    }
    for _rsc in _rpt_fifo_sc_ids:
        _rbl, _, _ = _fifo_lot_trace(cfg, _rsc)
        for _rsid, _rsd in _rbl.items():
            _rlots = _rsd.get("lots", {})
            _rqty  = sum(v["qty"]            for v in _rlots.values() if v.get("unit_cost") is not None)
            _ramt  = sum(v.get("amount",0.0) for v in _rlots.values() if v.get("unit_cost") is not None)
            if _rqty > 0:
                _rpt_fifo_rmc[(_rsc, _rsid)] = _ramt / _rqty
            _rpt_fifo_stor[(_rsc, _rsid)] = _rsd.get("storage_cost", 0.0)

    _rpt_batch_inp = defaultdict(float)
    for _rrr in _rpt_ph_all:
        _rpt_batch_inp[(_rrr.get("scrap_type_id",""), _rrr.get("shipment_id",""))] += _ph_input_kg(_rrr)

    def _rpt_raw(rec):
        """배치 원료비 — FIFO 우선, 없으면 선적일 기준 이동평균"""
        _sc3  = rec.get("scrap_type_id","")
        _sid3 = rec.get("shipment_id","") or f"__no_ship__{rec.get('id','')}"
        fifo3 = _rpt_fifo_rmc.get((_sc3, _sid3))
        if fifo3 is not None:
            return fifo3 * _ph_input_kg(rec)
        # 날짜 기준 이동평균 (선적일 → 더 정확한 당시 단가)
        _ref_date3 = _rec_ref_date(rec, cfg)
        avg3, _ = _inv_moving_avg(cfg, _sc3, _ref_date3)
        return (avg3 or 0) * _ph_input_kg(rec)

    def _rpt_stor(rec):
        """배치 보관비 — 수동 우선, 없으면 FIFO 자동 비례 배분"""
        man3 = _ph_storage_cost(rec, cfg)
        if man3:
            return man3
        _sc3  = rec.get("scrap_type_id","")
        _sid3 = rec.get("shipment_id","") or f"__no_ship__{rec.get('id','')}"
        tot3  = _rpt_fifo_stor.get((_sc3, _sid3), 0.0)
        if tot3 <= 0:
            return 0.0
        inp_tot3 = _rpt_batch_inp.get((_sc3, _sid3), 0.0)
        inp_r3   = _ph_input_kg(rec)
        return round(tot3 * (inp_r3 / inp_tot3), 4) if inp_tot3 > 0 else 0.0

    # ── Section 3: 손익 요약 ─────────────────────────────────────────────────
    st.markdown("#### ① 손익 요약")
    st.caption("**거래 마진** = BP 매각 − 임가공비 − 수출비  |  **실질 손익** = 거래 마진 − 원료 취득원가 − 보관비  (원료: FIFO 우선 → 이동평균)")

    _pnl_by_month = defaultdict(lambda: {"bp":0.0,"pf":0.0,"eu":0.0,"raw":0.0,"stor":0.0,"out":0.0,"inp":0.0,"cnt":0})
    for _r in _rpt_ph_all:
        _sh = _rpt_ship_m.get(_r.get("shipment_id",""), {})
        _mo = (_sh.get("loading_date") or "미연결")[:7]
        _out = float(_r.get("output_kg",0) or 0)
        _inp = _ph_input_kg(_r)
        _pnl_by_month[_mo]["bp"]   += float(_r.get("bp_sale_per_kg",0) or 0) * _out
        _pnl_by_month[_mo]["pf"]   += float(_r.get("processing_fee_per_kg",0) or 0) * _inp
        _pnl_by_month[_mo]["eu"]   += _ph_export_usd(_r, cfg)
        _pnl_by_month[_mo]["raw"]  += _rpt_raw(_r)
        _pnl_by_month[_mo]["stor"] += _rpt_stor(_r)
        _pnl_by_month[_mo]["out"]  += _out
        _pnl_by_month[_mo]["inp"]  += _inp
        _pnl_by_month[_mo]["cnt"]  += 1

    if _pnl_by_month:
        _pnl_rows = []
        _cum_net=0.0; _cum_real=0.0
        _tot_bp_r=0.0; _tot_pf_r=0.0; _tot_eu_r=0.0
        _tot_raw_r=0.0; _tot_stor_r=0.0; _tot_out_r=0.0; _tot_inp_r=0.0
        for _mo in sorted(_pnl_by_month.keys()):
            _v    = _pnl_by_month[_mo]
            _net  = _v["bp"] - _v["pf"] - _v["eu"]
            _real = _net - _v["raw"] - _v["stor"]
            _cum_net  += _net;  _cum_real += _real
            _tot_bp_r += _v["bp"];  _tot_pf_r  += _v["pf"];  _tot_eu_r  += _v["eu"]
            _tot_raw_r+= _v["raw"]; _tot_stor_r+= _v["stor"]
            _tot_out_r+= _v["out"]; _tot_inp_r += _v["inp"]
            _pnl_rows.append({
                "월":             _mo,
                "배치수":         _v["cnt"],
                "생산(kg)":       round(_v["out"],0),
                "BP 매각(USD)":   round(_v["bp"],2),
                "임가공비(USD)":  round(_v["pf"],2),
                "수출비(USD)":    round(_v["eu"],2),
                "거래 마진(USD)":  round(_net,2),
                "원료비(USD)":    round(_v["raw"],2),
                "보관비(USD)":    round(_v["stor"],2) if _v["stor"] else None,
                "실질 손익(USD)": round(_real,2),
                "누적 실질 손익": round(_cum_real,2),
            })

        # 요약 메트릭
        _pm1,_pm2,_pm3,_pm4,_pm5 = st.columns(5)
        _pm1.metric("총 BP 매각",    f"${_tot_bp_r:,.0f}")
        _pm2.metric("총 임가공비",   f"${_tot_pf_r:,.0f}")
        _pm3.metric("총 수출비",     f"${_tot_eu_r:,.0f}")
        _pm4.metric("누적 거래 마진", f"${_cum_net:+,.0f}",
                    delta_color="normal" if _cum_net >= 0 else "inverse")
        _pm5.metric("누적 실질 손익",f"${_cum_real:+,.0f}",
                    delta_color="normal" if _cum_real >= 0 else "inverse")

        # 원료 원가 기준 안내
        _rmc_fifo_cnt_r = sum(
            1 for r in _rpt_ph_all
            if _rpt_fifo_rmc.get((r.get("scrap_type_id",""),
                r.get("shipment_id","") or f"__no_ship__{r.get('id','')}")) is not None)
        _rmc_mavg_cnt_r = len(_rpt_ph_all) - _rmc_fifo_cnt_r
        _rmc_src_r = []
        if _rmc_fifo_cnt_r: _rmc_src_r.append(f"{_rmc_fifo_cnt_r}건 FIFO")
        if _rmc_mavg_cnt_r: _rmc_src_r.append(f"{_rmc_mavg_cnt_r}건 이동평균")
        if _rmc_src_r:
            st.caption(f"📌 원료비 원가 기준: {' / '.join(_rmc_src_r)}")

        def _hl_pnl(row):
            styles = [""] * len(row)
            _ci = list(row.index)
            v_net  = row.get("거래 마진(USD)",0) or 0
            v_real = row.get("실질 손익(USD)",0) or 0
            c_net  = ("background-color:#2E75B6;color:white;font-weight:600" if v_net>=0
                      else "background-color:#922b21;color:white;font-weight:600")
            c_real = ("background-color:#1E8449;color:white;font-weight:600" if v_real>=0
                      else "background-color:#C0392B;color:white;font-weight:600")
            if "거래 마진(USD)"  in _ci: styles[_ci.index("거래 마진(USD)")]  = c_net
            if "실질 손익(USD)" in _ci: styles[_ci.index("실질 손익(USD)")] = c_real
            return styles

        st.dataframe(
            pd.DataFrame(_pnl_rows).style.apply(_hl_pnl, axis=1).format(na_rep="—", formatter={
                "생산(kg)":       "{:,.0f}",
                "BP 매각(USD)":   "${:,.2f}",
                "임가공비(USD)":  "${:,.2f}",
                "수출비(USD)":    "${:,.2f}",
                "거래 마진(USD)":  "${:+,.2f}",
                "원료비(USD)":    "${:,.2f}",
                "보관비(USD)":    lambda v: f"${v:,.2f}" if v else "—",
                "실질 손익(USD)": "${:+,.2f}",
                "누적 실질 손익": "${:+,.2f}",
            }),
            use_container_width=True, hide_index=True,
        )

        # BP 1kg당 비용 분해 (FIFO vs 이동평균)
        if _tot_out_r > 0:
            _raw_mavg_r = sum(
                (_inv_moving_avg(cfg, r.get("scrap_type_id",""))[0] or 0) * _ph_input_kg(r)
                for r in _rpt_ph_all
            )
            _conv_r = _tot_out_r / _tot_inp_r * 100 if _tot_inp_r else 0
            _bp_pk_r  = _tot_bp_r  / _tot_out_r
            _pf_pk_r  = _tot_pf_r  / _tot_out_r
            _eu_pk_r  = _tot_eu_r  / _tot_out_r
            _gm_pk_r  = _bp_pk_r - _pf_pk_r - _eu_pk_r
            _rf_pk_r  = _tot_raw_r  / _tot_out_r
            _rm_pk_r  = _raw_mavg_r / _tot_out_r
            _st_pk_r  = _tot_stor_r / _tot_out_r
            _bpkg_r = [
                ("BP 매각가",            _bp_pk_r, _bp_pk_r),
                ("  (−) 임가공비",       _pf_pk_r, _pf_pk_r),
                ("  (−) 수출비",         _eu_pk_r, _eu_pk_r),
                ("= 거래 마진",          _gm_pk_r, _gm_pk_r),
                ("  (−) 원료 취득원가",  _rf_pk_r, _rm_pk_r),
                ("  (−) 보관비",         _st_pk_r, _st_pk_r),
                ("= 실질 손익",          _gm_pk_r - _rf_pk_r - _st_pk_r,
                                         _gm_pk_r - _rm_pk_r - _st_pk_r),
            ]
            with st.expander(f"📐 BP 1kg당 비용 분해  (전환율 가중평균 {_conv_r:.1f}%)", expanded=False):
                _df_bpkg_r = pd.DataFrame(_bpkg_r, columns=["항목","FIFO ($/kg BP)","이동평균 ($/kg BP)"])
                def _hl_bpkg_r(row):
                    lbl = row["항목"]; fv = row["FIFO ($/kg BP)"]; mv = row["이동평균 ($/kg BP)"]
                    if "실질 손익" in lbl:
                        bg = "#1E8449" if fv >= 0 else "#922b21"
                        s  = f"background-color:{bg};color:white;font-weight:700"
                        return ["font-weight:700", s, s]
                    if lbl.startswith("="):   return ["font-weight:700","font-weight:700","font-weight:700"]
                    if "BP 매각가" in lbl:    return ["font-weight:700","color:#1565c0;font-weight:600","color:#1565c0;font-weight:600"]
                    if "원료" in lbl and abs(fv - mv) > 0.0001:
                        fc = "color:#1E8449;font-weight:600" if fv<=mv else "color:#E74C3C;font-weight:600"
                        mc = "color:#1E8449;font-weight:600" if mv<=fv else "color:#E74C3C;font-weight:600"
                        return ["color:#555", fc, mc]
                    return ["color:#555","",""]
                st.dataframe(
                    _df_bpkg_r.style.apply(_hl_bpkg_r, axis=1)
                              .format({"FIFO ($/kg BP)":"${:+.4f}", "이동평균 ($/kg BP)":"${:+.4f}"}),
                    use_container_width=True, hide_index=True,
                )

        # 직접판매 손익 요약
        _ds_rpt = [ds for ds in cfg.get("direct_sales",[]) if ds.get("sale_price_per_kg") is not None]
        if _ds_rpt:
            _ds_rev_r = sum(float(ds.get("sale_price_per_kg",0)) * float(ds.get("quantity_kg",0)) for ds in _ds_rpt)
            _ds_qty_r = sum(float(ds.get("quantity_kg",0)) for ds in _ds_rpt)
            _ds_raw_r = sum(
                (_inv_moving_avg(cfg, ds.get("scrap_type_id",""))[0] or 0) * float(ds.get("quantity_kg",0))
                for ds in _ds_rpt
            )
            _ds_net_r = _ds_rev_r - _ds_raw_r
            with st.expander(f"🏷️ 직접판매 손익 요약  ({len(_ds_rpt)}건, {_ds_qty_r:,.0f} kg)", expanded=False):
                _dp1,_dp2,_dp3 = st.columns(3)
                _dp1.metric("직판 매출액",   f"${_ds_rev_r:,.0f}")
                _dp2.metric("원료 취득원가", f"${_ds_raw_r:,.0f}", help="이동평균 원가 기준")
                _dp3.metric("직판 매출이익", f"${_ds_net_r:+,.0f}",
                            delta_color="normal" if _ds_net_r >= 0 else "inverse")
    else:
        st.info("처리이력(임가공사 관리 탭)을 등록하면 손익 요약이 표시됩니다.")

    st.divider()

    # ── Section 4: 선적 정산 현황 ────────────────────────────────────────────
    st.markdown("#### ② 선적 정산 현황")
    if not _rpt_ships:
        st.info("등록된 선적건이 없습니다.")
    else:
        _prov_c  = sum(1 for s in _rpt_ships if s.get("status")=="provisional")
        _final_c = sum(1 for s in _rpt_ships if s.get("status")=="final")
        _paid_c  = sum(1 for s in _rpt_ships if s.get("status")=="paid")
        _tot_w   = sum(s.get("weight_kg",0) for s in _rpt_ships)
        _tot_inv = sum(s.get("invoice_usd",0) for s in _rpt_ships)
        _sc1,_sc2,_sc3,_sc4,_sc5,_sc6 = st.columns(6)
        _sc1.metric("총 선적건",    f"{len(_rpt_ships)}건")
        _sc2.metric("총 중량",      f"{_tot_w/1000:,.1f} t")
        _sc3.metric("Provisional 정산",     f"{_prov_c}건",
                    delta="미확정" if _prov_c else None,
                    delta_color="inverse" if _prov_c else "off")
        _sc4.metric("최종정산",     f"{_final_c}건")
        _sc5.metric("입금완료",     f"{_paid_c}건")
        _sc6.metric("총 Invoice",   f"${_tot_inv:,.0f}")

        _ship_tbl = []
        for _s in sorted(_rpt_ships, key=lambda x: x.get("loading_date",""), reverse=True):
            _sb = _rpt_buyer_m.get(_s.get("buyer_id"),{})
            _stat_lbl = {"provisional":"🟡 Provisional","final":"🟢 최종","paid":"🔵 입금"}.get(_s.get("status",""),"—")
            _ship_tbl.append({
                "HBL":         _s.get("hbl","—"),
                "매입사":      f"{_sb.get('name','?')} ({_sb.get('product','?')})",
                "선적일":      _s.get("loading_date",""),
                "중량(kg)":    _s.get("weight_kg",0),
                "Invoice(USD)":_s.get("invoice_usd",0),
                "Prov 월":     _s.get("prov_month","—"),
                "Final 월":    _s.get("final_month","—"),
                "상태":        _stat_lbl,
            })
        st.dataframe(pd.DataFrame(_ship_tbl).style.format(
            {"중량(kg)":"{:,.0f}","Invoice(USD)":"${:,.2f}"}),
            use_container_width=True, hide_index=True)

        # HBL별 손익 연결
        _hbl_linked = [r for r in _rpt_ph_all if r.get("shipment_id","")]
        if _hbl_linked:
            st.markdown("---")
            st.markdown("**HBL별 실질 손익** (처리이력 연결 기준)")
            _hbl4_agg = defaultdict(lambda: {"bp":0.0,"pf":0.0,"eu":0.0,"raw":0.0,"stor":0.0,"out":0.0})
            for _r4 in _hbl_linked:
                _hid4 = _r4["shipment_id"]
                _op4  = float(_r4.get("output_kg",0) or 0)
                _ip4  = _ph_input_kg(_r4)
                _hbl4_agg[_hid4]["bp"]   += float(_r4.get("bp_sale_per_kg",0) or 0) * _op4
                _hbl4_agg[_hid4]["pf"]   += float(_r4.get("processing_fee_per_kg",0) or 0) * _ip4
                _hbl4_agg[_hid4]["eu"]   += _ph_export_usd(_r4, cfg)
                _hbl4_agg[_hid4]["raw"]  += _rpt_raw(_r4)
                _hbl4_agg[_hid4]["stor"] += _rpt_stor(_r4)
                _hbl4_agg[_hid4]["out"]  += _op4

            _hbl4_rows = []
            for _hid4, _hv4 in sorted(
                _hbl4_agg.items(),
                key=lambda x: _rpt_ship_m.get(x[0],{}).get("loading_date",""),
                reverse=True,
            ):
                _hs4    = _rpt_ship_m.get(_hid4, {})
                _hb4    = _rpt_buyer_m.get(_hs4.get("buyer_id",""), {})
                _trade4 = _hv4["bp"] - _hv4["pf"] - _hv4["eu"]
                _real4  = _trade4 - _hv4["raw"] - _hv4["stor"]
                _mgr4   = _real4 / _hv4["bp"] * 100 if _hv4["bp"] > 0 else None
                _hbl4_rows.append({
                    "HBL":        _hs4.get("hbl","—"),
                    "선적일":     _hs4.get("loading_date",""),
                    "매입사":     f"{_hb4.get('name','?')} ({_hb4.get('product','?')})",
                    "BP(kg)":     round(_hv4["out"],0),
                    "거래 마진":  round(_trade4,2),
                    "원료비":     round(_hv4["raw"],2),
                    "보관비":     round(_hv4["stor"],2) if _hv4["stor"] else None,
                    "실질 손익":  round(_real4,2),
                    "마진율(%)":  round(_mgr4,2) if _mgr4 is not None else None,
                })

            def _hl_hbl4(row):
                styles = [""] * len(row)
                _ci4 = list(row.index)
                v4 = row.get("실질 손익", 0) or 0
                if "실질 손익" in _ci4:
                    styles[_ci4.index("실질 손익")] = (
                        "background-color:#1E8449;color:white;font-weight:600" if v4 >= 0
                        else "background-color:#C0392B;color:white;font-weight:600")
                return styles

            st.dataframe(
                pd.DataFrame(_hbl4_rows).style.apply(_hl_hbl4, axis=1).format(na_rep="—", formatter={
                    "BP(kg)":    "{:,.0f}",
                    "거래 마진": "${:+,.2f}",
                    "원료비":    "${:,.2f}",
                    "보관비":    lambda v: f"${v:,.2f}" if v else "—",
                    "실질 손익": "${:+,.2f}",
                    "마진율(%)": lambda v: f"{v:+.2f}%" if v is not None else "—",
                }),
                use_container_width=True, hide_index=True,
            )

    st.divider()

    # ── Section 5: 원료 재고 현황 ────────────────────────────────────────────
    st.markdown("#### ③ 원료 재고 현황")
    _inv_cfg = cfg.get("raw_material_inventory", {})
    _inv_sc_list = [s for s in cfg.get("scrap_types",[]) if _inv_cfg.get(s["id"],{}).get("opening")]
    if not _inv_sc_list:
        st.info("스크랩 유형 관리 탭에서 기초재고를 설정하면 재고 현황이 표시됩니다.")
    else:
        # FIFO 잔여 Lot 가중평균 단가 계산
        _rpt_fifo_rem = {}  # sc_id → ($/kg, remain_kg)
        for _isc5 in _inv_sc_list:
            _, _, _lot_rem5 = _fifo_lot_trace(cfg, _isc5["id"])
            _rqty5 = sum(lr["remain"] for lr in _lot_rem5 if lr.get("unit_cost") is not None)
            _ramt5 = sum(lr["remain"] * lr["unit_cost"] for lr in _lot_rem5 if lr.get("unit_cost") is not None)
            _rpt_fifo_rem[_isc5["id"]] = (_ramt5 / _rqty5 if _rqty5 > 0 else None, _rqty5)

        _inv_rows = []
        for _isc in _inv_sc_list:
            _avg, _total_in = _inv_moving_avg(cfg, _isc["id"])
            _dispatched  = sum(float(dr.get("quantity_kg",0))
                               for dr in cfg.get("dispatch_records",[])
                               if dr.get("scrap_type_id")==_isc["id"])
            _direct_sold = sum(float(ds.get("quantity_kg",0))
                               for ds in cfg.get("direct_sales",[])
                               if ds.get("scrap_type_id")==_isc["id"])
            _balance  = round(_total_in - _dispatched - _direct_sold, 0)
            _est_val  = round(_balance * _avg, 0)      if (_avg   and _balance > 0) else None
            _fifo_avg5, _ = _rpt_fifo_rem.get(_isc["id"], (None, 0))
            _fifo_val5 = round(_balance * _fifo_avg5, 0) if (_fifo_avg5 and _balance > 0) else None
            _inv_rows.append({
                "스크랩 유형":        _isc["name"],
                "누적 입고(kg)":      round(_total_in, 0),
                "임가공 출고(kg)":    round(_dispatched, 0),
                "직접판매(kg)":       round(_direct_sold, 0),
                "창고 잔량(kg)":      _balance,
                "이동평균 단가":      round(_avg,      5) if _avg      else None,
                "FIFO 잔여 단가":     round(_fifo_avg5,5) if _fifo_avg5 else None,
                "평가액(이동평균)":   _est_val,
                "평가액(FIFO)":       _fifo_val5,
            })

        def _hl_inv(row):
            styles = [""] * len(row)
            v = row.get("창고 잔량(kg)", 0) or 0
            if v < 0:
                styles[4] = "background-color:#fadbd8"
            return styles

        st.dataframe(
            pd.DataFrame(_inv_rows).style.apply(_hl_inv, axis=1).format(na_rep="—", formatter={
                "누적 입고(kg)":    "{:,.0f}",
                "임가공 출고(kg)":  "{:,.0f}",
                "직접판매(kg)":     "{:,.0f}",
                "창고 잔량(kg)":    "{:,.0f}",
                "이동평균 단가":    lambda v: f"${v:.5f}" if v is not None else "—",
                "FIFO 잔여 단가":   lambda v: f"${v:.5f}" if v is not None else "—",
                "평가액(이동평균)": lambda v: f"${v:,.0f}" if v is not None else "—",
                "평가액(FIFO)":     lambda v: f"${v:,.0f}" if v is not None else "—",
            }),
            use_container_width=True, hide_index=True,
        )
        st.caption("창고 잔량 = 누적 입고 − 임가공 출고 − 직접판매  |  FIFO 잔여 단가 = 미소진 Lot 가중평균  |  임가공사 보유 현황은 ⑥ 참조")

    st.divider()

    # ── Section 6: 임가공사 보유 현황 ───────────────────────────────────────────
    st.markdown("#### ④ 임가공사 보유 현황")
    st.caption("임가공 출고 누적 vs B/L 처리 투입 누적 — 차이가 현재 임가공사 보유 중인 스크랩 추정량입니다.")
    _dr_all = cfg.get("dispatch_records", [])
    if not _dr_all:
        st.info("출고 기록 탭에서 임가공 출고를 입력하면 파이프라인 현황이 표시됩니다.")
    else:
        # 프로세서 × 스크랩유형 집계
        _pipe_agg = defaultdict(lambda: {"dispatched": 0.0, "processed": 0.0})
        for _dr in _dr_all:
            _k = (_dr.get("processor_id",""), _dr.get("scrap_type_id",""))
            _pipe_agg[_k]["dispatched"] += float(_dr.get("quantity_kg",0))
        for _r in _rpt_ph_all:
            _k = (_r.get("processor_id",""), _r.get("scrap_type_id",""))
            _pipe_agg[_k]["processed"] += _ph_input_kg(_r)

        _pipe_rows = []
        for (_pid, _sid), _pv in sorted(_pipe_agg.items(),
             key=lambda x: (_rpt_proc_m.get(x[0][0],{}).get("name",""),
                            _rpt_sc_m.get(x[0][1],{}).get("name",""))):
            _pr_name = _rpt_proc_m.get(_pid,{}).get("name","—")
            _sc_name = _rpt_sc_m.get(_sid,{}).get("name","—")
            _remain  = _pv["dispatched"] - _pv["processed"]
            _pipe_rows.append({
                "임가공사":        _pr_name,
                "스크랩 유형":     _sc_name,
                "출고 누적(kg)":   round(_pv["dispatched"],0),
                "B/L 투입(kg)":    round(_pv["processed"],0),
                "임가공사 보유(kg)":round(_remain, 0),
            })
        def _hl_pipe(row):
            v = row.get("임가공사 보유(kg)",0) or 0
            if v < -1: return ["","","","","color:#b71c1c;font-weight:600"]
            if v > 0:  return ["","","","","color:#1565c0;font-weight:600"]
            return [""]*5
        st.dataframe(pd.DataFrame(_pipe_rows).style
            .apply(_hl_pipe, axis=1)
            .format({
                "출고 누적(kg)":    "{:,.0f}",
                "B/L 투입(kg)":     "{:,.0f}",
                "임가공사 보유(kg)": "{:+,.0f}",
            }), use_container_width=True, hide_index=True)

    st.divider()

    # ── Section 7: 보관비 현황 ───────────────────────────────────────────────
    st.markdown("#### ⑤ 보관비 현황")
    st.caption("수동 입력(storage_days) 및 FIFO 자동 계산 보관비를 스크랩 유형별로 집계합니다.")
    if not _rpt_ph_all:
        st.info("처리이력을 등록하면 보관비 현황이 표시됩니다.")
    else:
        _stor7_by_sc  = defaultdict(lambda: {"manual":0.0, "fifo_auto":0.0, "manual_cnt":0, "auto_cnt":0})
        _stor7_by_mon = defaultdict(lambda: {"manual":0.0, "fifo_auto":0.0})
        for _r7 in _rpt_ph_all:
            _scid7 = _r7.get("scrap_type_id","")
            _sh7   = _rpt_ship_m.get(_r7.get("shipment_id",""), {})
            _mo7   = (_sh7.get("loading_date") or "미상")[:7]
            _man7  = _ph_storage_cost(_r7, cfg)
            _auto7 = _rpt_stor(_r7) if not _man7 else 0.0
            if _man7:
                _stor7_by_sc[_scid7]["manual"]     += _man7
                _stor7_by_sc[_scid7]["manual_cnt"] += 1
                _stor7_by_mon[_mo7]["manual"]      += _man7
            elif _auto7:
                _stor7_by_sc[_scid7]["fifo_auto"]  += _auto7
                _stor7_by_sc[_scid7]["auto_cnt"]   += 1
                _stor7_by_mon[_mo7]["fifo_auto"]   += _auto7

        # 스크랩 유형별 합계
        _stor7_sc_rows = []
        _tot_man7 = 0.0; _tot_auto7 = 0.0
        for _scid7, _sv7 in _stor7_by_sc.items():
            _sc7nm = _rpt_sc_m.get(_scid7, {}).get("name","—")
            _tot7  = _sv7["manual"] + _sv7["fifo_auto"]
            _stor7_sc_rows.append({
                "스크랩 유형":      _sc7nm,
                "수동 보관비(USD)": round(_sv7["manual"],2)    if _sv7["manual"]    else None,
                "수동 건수":        _sv7["manual_cnt"]          if _sv7["manual_cnt"] else None,
                "FIFO 자동(USD)":   round(_sv7["fifo_auto"],2)  if _sv7["fifo_auto"] else None,
                "자동 건수":        _sv7["auto_cnt"]            if _sv7["auto_cnt"]  else None,
                "합계(USD)":        round(_tot7, 2),
            })
            _tot_man7 += _sv7["manual"]; _tot_auto7 += _sv7["fifo_auto"]

        if _stor7_sc_rows:
            _s7c1, _s7c2, _s7c3 = st.columns(3)
            _s7c1.metric("수동 보관비 합계",    f"${_tot_man7:,.2f}")
            _s7c2.metric("FIFO 자동 보관비 합계",f"${_tot_auto7:,.2f}")
            _s7c3.metric("총 보관비",            f"${_tot_man7+_tot_auto7:,.2f}")

            st.dataframe(
                pd.DataFrame(_stor7_sc_rows).style.format(na_rep="—", formatter={
                    "수동 보관비(USD)": lambda v: f"${v:,.2f}" if v else "—",
                    "FIFO 자동(USD)":   lambda v: f"${v:,.2f}" if v else "—",
                    "합계(USD)":        "${:,.2f}",
                }),
                use_container_width=True, hide_index=True,
            )

        # 월별 보관비 집계
        if _stor7_by_mon:
            with st.expander("📅 월별 보관비 집계", expanded=False):
                _stor7_mon_rows = []
                for _mo7 in sorted(_stor7_by_mon.keys()):
                    _mv7  = _stor7_by_mon[_mo7]
                    _tot7m = _mv7["manual"] + _mv7["fifo_auto"]
                    _stor7_mon_rows.append({
                        "월":             _mo7,
                        "수동(USD)":      round(_mv7["manual"],2)    if _mv7["manual"]    else None,
                        "FIFO 자동(USD)": round(_mv7["fifo_auto"],2) if _mv7["fifo_auto"] else None,
                        "합계(USD)":      round(_tot7m, 2),
                    })
                st.dataframe(
                    pd.DataFrame(_stor7_mon_rows).style.format(na_rep="—", formatter={
                        "수동(USD)":      lambda v: f"${v:,.2f}" if v else "—",
                        "FIFO 자동(USD)": lambda v: f"${v:,.2f}" if v else "—",
                        "합계(USD)":      "${:,.2f}",
                    }),
                    use_container_width=True, hide_index=True,
                )

# ══════════════════════════════════════════════════════════════════════════════
# TAB — 문서 보관함  (bp_documents.json 별도 Drive 파일 사용)
# ══════════════════════════════════════════════════════════════════════════════
with t_docs:
    import base64 as _b64

    st.markdown("### 📎 문서 보관함")
    st.caption("계약서·CoA 등 관련 파일을 별도 Drive 파일(bp_documents.json)에 보관합니다. 파일당 최대 10 MB.")

    _DOC_CATS  = ["계약서", "CoA", "인보이스", "포워더 서류", "기타"]
    _DOC_LIMIT = 10 * 1024 * 1024  # 10 MB

    try:
        _docs_list = _load_docs_drive()
    except Exception as _docs_err:
        st.error(f"문서 파일 로드 실패: {_docs_err}")
        _docs_list = []

    # ── 업로드 ────────────────────────────────────────────────────────────────
    with st.expander("➕ 파일 저장", expanded=not _docs_list):
        _dup_file  = st.file_uploader("파일 선택 (최대 10 MB)", key="docs_uploader",
                                       help="PDF, Excel, 이미지 등 모든 파일 형식 지원")
        _dup_cat   = st.selectbox("카테고리", _DOC_CATS, key="docs_cat")
        _dup_tag   = st.text_input("태그 (선택)", placeholder="예: TGLHUS26060003, ECOPRO", key="docs_tag")
        _dup_notes = st.text_input("메모 (선택)", placeholder="간단한 설명", key="docs_notes")

        if st.button("💾 저장", key="docs_upload_btn", disabled=_dup_file is None):
            if _dup_file:
                _file_bytes = _dup_file.read()
                if len(_file_bytes) > _DOC_LIMIT:
                    st.error(f"파일 크기 {len(_file_bytes)/1024/1024:.1f} MB — 10 MB 이하만 저장 가능합니다.")
                else:
                    with st.spinner("저장 중..."):
                        _new_doc = {
                            "id":          str(uuid.uuid4())[:8],
                            "filename":    _dup_file.name,
                            "mime":        _dup_file.type or "application/octet-stream",
                            "category":    _dup_cat,
                            "tag":         _dup_tag.strip(),
                            "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "notes":       _dup_notes.strip(),
                            "size_bytes":  len(_file_bytes),
                            "content_b64": _b64.b64encode(_file_bytes).decode(),
                        }
                        try:
                            _save_docs_drive(_docs_list + [_new_doc])
                            st.success(f"✅ **{_dup_file.name}** 저장 완료")
                            st.rerun()
                        except Exception as _e:
                            st.error(f"저장 실패: {_e}")

    # ── 문서 목록 ─────────────────────────────────────────────────────────────
    if not _docs_list:
        st.info("저장된 문서가 없습니다.")
    else:
        _total_sz = sum(d.get("size_bytes", 0) for d in _docs_list)
        _sz_total_str = f"{_total_sz/1024:.0f} KB" if _total_sz < 1024*1024 else f"{_total_sz/1024/1024:.1f} MB"
        st.caption(f"총 {len(_docs_list)}건 · {_sz_total_str} 저장됨")

        _df_cat_opts = ["전체"] + sorted({d.get("category","기타") for d in _docs_list})
        _df_cat_f = st.selectbox("카테고리 필터", _df_cat_opts, key="docs_cat_filter")
        _df_tag_f = st.text_input("태그 검색", placeholder="HBL번호, 회사명 등", key="docs_tag_filter").strip().lower()

        _filtered_docs = [
            d for d in sorted(_docs_list, key=lambda x: x.get("uploaded_at",""), reverse=True)
            if (_df_cat_f == "전체" or d.get("category") == _df_cat_f)
            and (_df_tag_f == "" or _df_tag_f in d.get("tag","").lower() or _df_tag_f in d.get("filename","").lower())
        ]

        st.markdown(f"**{len(_filtered_docs)}건**")

        for _doc in _filtered_docs:
            _doc_id = _doc.get("id","")
            _fname  = _doc.get("filename","(파일명 없음)")
            _dcat   = _doc.get("category","기타")
            _dtag   = _doc.get("tag","")
            _dnotes = _doc.get("notes","")
            _dup_at = _doc.get("uploaded_at","")
            _dsz    = _doc.get("size_bytes", 0)
            _sz_str = f"{_dsz/1024:.1f} KB" if _dsz < 1024*1024 else f"{_dsz/1024/1024:.1f} MB"

            with st.container(border=True):
                _dc1, _dc2, _dc3 = st.columns([6, 2, 2])
                with _dc1:
                    st.markdown(f"**{_fname}**")
                    _meta_parts = [f"`{_dcat}`"]
                    if _dtag:   _meta_parts.append(f"🏷️ {_dtag}")
                    if _dnotes: _meta_parts.append(f"📝 {_dnotes}")
                    _meta_parts.append(f"🕐 {_dup_at}")
                    if _dsz:    _meta_parts.append(f"({_sz_str})")
                    st.caption("  ·  ".join(_meta_parts))
                with _dc2:
                    _b64_data = _doc.get("content_b64", "")
                    if _b64_data:
                        st.download_button(
                            label="⬇️ 다운로드",
                            data=_b64.b64decode(_b64_data),
                            file_name=_fname,
                            mime=_doc.get("mime", "application/octet-stream"),
                            key=f"docs_dl_{_doc_id}",
                        )
                with _dc3:
                    with st.popover("🗑️ 삭제"):
                        st.warning(f"**{_fname}** 을(를) 삭제합니다.")
                        if st.button("확인 삭제", key=f"docs_del_ok_{_doc_id}"):
                            with st.spinner("삭제 중..."):
                                try:
                                    _save_docs_drive([d for d in _docs_list if d.get("id") != _doc_id])
                                    st.rerun()
                                except Exception as _e:
                                    st.error(f"삭제 실패: {_e}")

# ══════════════════════════════════════════════════════════════════════════════
# TAB — 계약 이행
# ══════════════════════════════════════════════════════════════════════════════

def _avg_conv_rate(cfg, scrap_id):
    """스크랩 유형별 평균 전환율(%) — 처리 이력 실적 기반."""
    rates = [float(r["conversion_rate_pct"])
             for r in cfg.get("processing_history", [])
             if r.get("scrap_type_id") == scrap_id and r.get("conversion_rate_pct")]
    return round(sum(rates) / len(rates), 1) if rates else 80.0

def _at_processor_raw_kg(cfg, scrap_id, processor_id=None):
    """임가공사에 있는 미처리 원료 추정량(kg) = 출하 누계 − 처리 이력 투입 누계.
    processor_id 지정 시 해당 임가공사 한정."""
    def _pr_ok(pid): return processor_id is None or pid == processor_id
    dispatched = sum(
        float(dr.get("quantity_kg") or 0)
        for dr in cfg.get("dispatch_records", [])
        if dr.get("scrap_type_id") == scrap_id and _pr_ok(dr.get("processor_id"))
    )
    processed = sum(
        _ph_input_kg(r)
        for r in cfg.get("processing_history", [])
        if r.get("scrap_type_id") == scrap_id and _pr_ok(r.get("processor_id"))
    )
    return max(0.0, dispatched - processed)

def _ship_in_period(ship, start, end):
    ld = ship.get("loading_date", "")
    if not ld:
        return False
    return (not start or ld >= start) and ld <= end

def _contract_metrics(cfg, contract):
    """계약 한 건의 이행 현황 지표 dict 반환."""
    contract_id = contract.get("id", "")
    buyer_id  = contract.get("buyer_id", "")
    scrap_id  = contract.get("scrap_type_id", "")
    qty_mt    = float(contract.get("contract_qty_mt") or 0)
    tol       = float(contract.get("tolerance_pct") or 0)
    def _norm_date(d, is_end=False):
        if not d: return d
        d = d.strip()
        if len(d) == 7:
            return d + ("-31" if is_end else "-01")
        return d
    start = _norm_date(contract.get("start_date", ""))
    end   = _norm_date(contract.get("end_date", "") or "9999-12-31", is_end=True) or "9999-12-31"

    # 선적 완료량 — 배분 기록 우선
    # 같은 매입사의 어느 계약이라도 배분 기록이 있으면 전체 배분 모드 강제
    # (없으면 buyer_id + scrap_type_id 기간 합산 fallback)
    _shipments  = cfg.get("shipments", [])
    ship_map    = {s["id"]: s for s in _shipments if s.get("id")}
    ct_allocs   = [a for a in cfg.get("contract_allocations", [])
                   if a.get("contract_id") == contract_id]
    _buyer_contract_ids = {c["id"] for c in cfg.get("contracts", [])
                           if c.get("buyer_id") == buyer_id}
    _buyer_has_any_alloc = any(
        a for a in cfg.get("contract_allocations", [])
        if a.get("contract_id") in _buyer_contract_ids
    )
    if ct_allocs:
        shipped_kg = sum(
            float(a.get("allocated_kg") or 0)
            for a in ct_allocs
            if _ship_in_period(ship_map.get(a.get("shipment_id",""), {}), start, end)
        )
    elif _buyer_has_any_alloc:
        # 다른 계약에 배분이 있으면 이 계약은 배분 미입력 = 0
        shipped_kg = 0.0
    else:
        shipped_kg = sum(
            float(s.get("weight_kg") or 0)
            for s in _shipments
            if s.get("buyer_id") == buyer_id
            and _ship_in_period(s, start, end)
        )
    shipped_mt = shipped_kg / 1000

    # 허용 범위
    min_mt = qty_mt * (1 - tol / 100)
    max_mt = qty_mt * (1 + tol / 100)

    # 이행률 (계약 중심값 기준)
    fulfill_pct = (shipped_mt / qty_mt * 100) if qty_mt else 0

    # 잔여 의무 (min 기준 — 최소 이행해야 할 잔여)
    remaining_mt = max(0.0, min_mt - shipped_mt)

    # 창고 원료 → BP 환산
    conv = _avg_conv_rate(cfg, scrap_id)
    _, _, _lot_rem = _fifo_lot_trace(cfg, scrap_id)
    warehouse_raw_kg  = sum(lot.get("remain", 0) for lot in _lot_rem)
    warehouse_bp_mt   = warehouse_raw_kg * conv / 100 / 1000

    # 임가공사 미처리 → BP 예상 (계약에 임가공사 지정 시 해당 임가공사만)
    _ct_proc_id    = contract.get("processor_id") or None
    at_proc_raw_kg = _at_processor_raw_kg(cfg, scrap_id, _ct_proc_id)
    at_proc_bp_mt  = at_proc_raw_kg * conv / 100 / 1000

    total_avail_mt = warehouse_bp_mt + at_proc_bp_mt

    # 상태 판정
    if shipped_mt >= min_mt:
        status = "complete"        # 이미 최소 충족
    elif total_avail_mt >= remaining_mt:
        status = "ok"              # 재고로 충당 가능
    else:
        status = "short"           # 재고 부족

    return {
        "shipped_mt":       shipped_mt,
        "qty_mt":           qty_mt,
        "min_mt":           min_mt,
        "max_mt":           max_mt,
        "fulfill_pct":      fulfill_pct,
        "remaining_mt":     remaining_mt,
        "conv_pct":         conv,
        "warehouse_raw_kg": warehouse_raw_kg,
        "warehouse_bp_mt":  warehouse_bp_mt,
        "at_proc_raw_kg":   at_proc_raw_kg,
        "at_proc_bp_mt":    at_proc_bp_mt,
        "total_avail_mt":   total_avail_mt,
        "status":           status,
    }

with t_contract:
    st.markdown("### 📋 계약 이행 현황")

    _ct_list    = cfg.setdefault("contracts", [])
    _ct_buyers  = {b["id"]: b for b in cfg.get("buyers", [])}
    _ct_scraps  = {s["id"]: s for s in cfg.get("scrap_types", [])}
    _ct_procs   = {p["id"]: p for p in cfg.get("processors", [])}
    _ct_sc_opts = {s["name"]: s["id"] for s in cfg.get("scrap_types", []) if s.get("active", True)}
    _ct_by_opts = {}
    for _b in cfg.get("buyers", []):
        if not _b.get("active", True):
            continue
        _lbl = f"{_b['name']} ({_b['product']})"
        if _lbl in _ct_by_opts:
            _old_id = _ct_by_opts.pop(_lbl)
            _ct_by_opts[f"{_lbl} [{_old_id[:4]}]"] = _old_id
            _lbl = f"{_lbl} [{_b['id'][:4]}]"
        _ct_by_opts[_lbl] = _b["id"]
    _ct_pr_opts = {"전체 (구분 없음)": ""} | {p["name"]: p["id"] for p in cfg.get("processors", []) if p.get("active", True)}

    # ── 계약 등록 ────────────────────────────────────────────────────────────
    with st.expander("➕ 계약 등록", expanded=not _ct_list):
        _cf1, _cf2 = st.columns(2)
        _ct_buyer_sel = _cf1.selectbox("매입사",   list(_ct_by_opts), key="ct_buyer")
        _ct_sc_sel    = _cf2.selectbox("원료 유형", list(_ct_sc_opts), key="ct_sc")
        _cf3, _cf4, _cf5 = st.columns(3)
        _ct_qty  = _cf3.number_input("계약량 (MT)", min_value=0.0, step=1.0, format="%.1f", key="ct_qty")
        _ct_tol  = _cf4.number_input("허용 오차 (%)", min_value=0.0, max_value=20.0, value=5.0, step=1.0, format="%.0f", key="ct_tol")
        _ct_prod = _cf5.selectbox("제품", ["BP", "BM"], key="ct_prod")
        _cf6, _cf7, _cf8 = st.columns(3)
        _ct_start  = _cf6.text_input("계약 시작일 (YYYY-MM-DD)", key="ct_start")
        _ct_end    = _cf7.text_input("계약 종료일 (YYYY-MM-DD)", key="ct_end")
        _ct_pr_sel = _cf8.selectbox("임가공사 (선택)", list(_ct_pr_opts), key="ct_proc",
                                     help="지정 시 해당 임가공사 재고만 충당에 반영")
        _ct_notes = st.text_input("메모", key="ct_notes")
        if st.button("💾 계약 등록", key="ct_add_btn"):
            def _is_valid_date(s):
                if not s.strip(): return True
                try: date.fromisoformat(s.strip()); return True
                except ValueError: return False
            if not _ct_buyer_sel or not _ct_sc_sel or _ct_qty <= 0:
                st.error("매입사, 원료 유형, 계약량을 입력하세요.")
            elif not _is_valid_date(_ct_start) or not _is_valid_date(_ct_end):
                st.error("날짜 형식 오류 — YYYY-MM-DD 형식으로 입력하세요. (예: 2026-05-01)")
            else:
                _ct_list.append({
                    "id":              str(uuid.uuid4())[:8],
                    "buyer_id":        _ct_by_opts[_ct_buyer_sel],
                    "product":         _ct_prod,
                    "scrap_type_id":   _ct_sc_opts[_ct_sc_sel],
                    "processor_id":    _ct_pr_opts[_ct_pr_sel],
                    "contract_qty_mt": _ct_qty,
                    "tolerance_pct":   _ct_tol,
                    "start_date":      _ct_start.strip(),
                    "end_date":        _ct_end.strip(),
                    "notes":           _ct_notes.strip(),
                })
                cfg["contracts"] = _ct_list
                save_cfg(cfg)
                st.success("계약 등록 완료")
                st.rerun()

    if not _ct_list:
        st.info("등록된 계약이 없습니다.")
    else:
        # ── 재고 현황 요약 (임가공사별) ───────────────────────────────────────
        st.markdown("#### 📦 가용 재고 현황")
        _inv_cols = st.columns(len(active_scraps)) if active_scraps else []
        for _ci, _sc in enumerate(active_scraps):
            _cv = _avg_conv_rate(cfg, _sc["id"])
            _, _, _lr = _fifo_lot_trace(cfg, _sc["id"])
            _wh = sum(lot.get("remain", 0) for lot in _lr)
            with _inv_cols[_ci]:
                st.markdown(f"**{_sc['name']}**  `전환율 {_cv:.0f}%`")
                _ic1, _ic2 = st.columns(2)
                _ic1.metric("창고 원료", f"{_wh/1000:,.2f} MT", help="FIFO 잔량")
                # 임가공사별 분류
                _ap_total = 0.0
                _ap_rows  = []
                for _pr in active_procs:
                    _ap_pr = _at_processor_raw_kg(cfg, _sc["id"], _pr["id"])
                    if _ap_pr > 0:
                        _ap_total += _ap_pr
                        _ap_rows.append((_pr["name"], _ap_pr))
                _ic2.metric("임가공사 (원료 합계)", f"{_ap_total/1000:,.2f} MT",
                             help="출하 누계 − 처리 이력 투입 누계")
                if _ap_rows:
                    for _prname, _apkg in _ap_rows:
                        st.caption(f"  └ {_prname}: {_apkg/1000:,.2f} MT → BP {_apkg*_cv/100/1000:,.2f} MT")
                st.caption(f"BP 환산 합계: **{(_wh+_ap_total)*_cv/100/1000:,.2f} MT**")

        st.divider()

        # ── 계약별 이행 현황 ──────────────────────────────────────────────────
        st.markdown("#### 계약별 이행 현황")
        _ship_map_disp = {s["id"]: s for s in cfg.get("shipments", [])}
        for _ct in _ct_list:
            _ct_id   = _ct.get("id","")
            _ct_buyer_obj = _ct_buyers.get(_ct.get("buyer_id",""), {})
            _bname   = f"{_ct_buyer_obj.get('name','—')} ({_ct_buyer_obj.get('product','')})" if _ct_buyer_obj else "—"
            _scname  = _ct_scraps.get(_ct.get("scrap_type_id",""), {}).get("name", "—")
            _prname  = _ct_procs.get(_ct.get("processor_id",""), {}).get("name", "") if _ct.get("processor_id") else ""
            _m       = _contract_metrics(cfg, _ct)

            _stat_color = {"complete": "🟢", "ok": "🟡", "short": "🔴"}.get(_m["status"], "⚪")
            _stat_label = {"complete": "이행 완료", "ok": "재고 충분", "short": "재고 부족"}.get(_m["status"], "—")

            with st.expander(
                f"{_stat_color} **{_bname}**  {_ct.get('product','BP')} / {_scname}"
                + (f" / {_prname}" if _prname else "")
                + f"  |  계약 {_m['qty_mt']:,.1f} MT ±{_ct.get('tolerance_pct',0):.0f}%"
                + f"  |  선적 {_m['shipped_mt']:,.2f} MT ({_m['fulfill_pct']:.1f}%)"
                + f"  |  {_stat_label}",
                expanded=_m["status"] == "short",
            ):
                # 이행률 프로그레스
                _prog_val = min(1.0, _m["shipped_mt"] / _m["max_mt"]) if _m["max_mt"] else 0
                st.progress(_prog_val)

                _mc1, _mc2, _mc3, _mc4 = st.columns(4)
                _mc1.metric("계약량", f"{_m['qty_mt']:,.1f} MT",
                             help=f"허용 범위: {_m['min_mt']:,.1f} ~ {_m['max_mt']:,.1f} MT")
                _mc2.metric("선적 완료", f"{_m['shipped_mt']:,.2f} MT",
                             delta=f"{_m['fulfill_pct']:.1f}%")
                _mc3.metric("잔여 의무 (최소 기준)",
                             f"{_m['remaining_mt']:,.2f} MT" if _m["remaining_mt"] > 0 else "충족",
                             delta_color="inverse" if _m["remaining_mt"] > 0 else "normal")
                _mc4.metric("충당 가능 (BP 환산)",
                             f"{_m['total_avail_mt']:,.2f} MT",
                             delta=f"{_m['total_avail_mt']-_m['remaining_mt']:+.2f} MT",
                             delta_color="normal" if _m["total_avail_mt"] >= _m["remaining_mt"] else "inverse")

                _pr_scope = f"임가공사: {_prname}" if _prname else "임가공사: 전체"
                st.caption(
                    f"전환율 {_m['conv_pct']:.0f}% 적용  ·  {_pr_scope}  ·  "
                    f"창고 원료 {_m['warehouse_raw_kg']/1000:,.2f} MT → BP {_m['warehouse_bp_mt']:,.2f} MT  ·  "
                    f"임가공사 미처리 {_m['at_proc_raw_kg']/1000:,.2f} MT → BP {_m['at_proc_bp_mt']:,.2f} MT"
                )
                if _ct.get("notes"):
                    st.caption(f"📝 {_ct['notes']}")

                # 기간 표시
                if _ct.get("start_date") or _ct.get("end_date"):
                    st.caption(f"📅 계약 기간: {_ct.get('start_date','—')} ~ {_ct.get('end_date','—')}")

                # ── 선적 배분 관리 ──────────────────────────────────────────
                st.divider()
                st.markdown("**📦 선적 배분 (계약별 부분 배정)**")

                _allocs = cfg.setdefault("contract_allocations", [])
                _ct_allocs_cur = [a for a in _allocs if a.get("contract_id") == _ct_id]

                if _ct_allocs_cur:
                    _alloc_rows = []
                    for _a in _ct_allocs_cur:
                        _s = _ship_map_disp.get(_a.get("shipment_id", ""), {})
                        _alloc_rows.append({
                            "HBL": _s.get("hbl") or _s.get("hbl_number") or "—",
                            "선적일": _s.get("loading_date", "—"),
                            "배분량 (MT)": round(float(_a.get("allocated_kg") or 0) / 1000, 3),
                            "_alloc_id": _a.get("id", ""),
                        })
                    for _ar in _alloc_rows:
                        _ac1, _ac2, _ac3, _ac4 = st.columns([3, 2, 2, 1])
                        _ac1.write(_ar["HBL"])
                        _ac2.write(_ar["선적일"])
                        _ac3.write(f"{_ar['배분량 (MT)']:,.3f} MT")
                        if _ac4.button("삭제", key=f"alloc_del_{_ar['_alloc_id']}"):
                            cfg["contract_allocations"] = [
                                x for x in _allocs if x.get("id") != _ar["_alloc_id"]
                            ]
                            save_cfg(cfg)
                            st.rerun()
                else:
                    st.caption("배분 기록 없음 — 아래에서 추가하세요.")

                # 배분 추가 폼
                _buyer_ships = [
                    s for s in cfg.get("shipments", [])
                    if s.get("buyer_id") == _ct.get("buyer_id", "")
                ]
                if _buyer_ships:
                    _ship_opts = {
                        f"{s.get('hbl') or s.get('hbl_number') or '—'}  ({s.get('loading_date','—')}, {float(s.get('weight_kg') or 0)/1000:,.2f} MT)": s["id"]
                        for s in _buyer_ships if s.get("id")
                    }
                    with st.form(key=f"alloc_form_{_ct_id}"):
                        _fa1, _fa2 = st.columns([4, 2])
                        _sel_ship_label = _fa1.selectbox("선적건 선택", list(_ship_opts), key=f"alloc_ship_{_ct_id}")
                        _alloc_kg_input = _fa2.number_input("배분량 (MT)", min_value=0.0, step=0.001, format="%.3f", key=f"alloc_kg_{_ct_id}")
                        # 선택한 선적건의 기배분/잔여 표시
                        _preview_sid = _ship_opts.get(_sel_ship_label, "")
                        _preview_total = float(next((s.get("weight_kg",0) for s in _buyer_ships if s.get("id")==_preview_sid), 0) or 0)
                        _preview_used  = sum(float(a.get("allocated_kg",0)) for a in _allocs if a.get("shipment_id")==_preview_sid)
                        _preview_rem   = _preview_total - _preview_used
                        st.caption(f"선적 합계 {_preview_total/1000:,.3f} MT  ·  기배분 {_preview_used/1000:,.3f} MT  ·  잔여 **{_preview_rem/1000:,.3f} MT**")
                        _add_alloc = st.form_submit_button("➕ 배분 추가")
                    if _add_alloc:
                        _sel_sid = _ship_opts[_sel_ship_label]
                        _ship_total_kg = float(next((s.get("weight_kg",0) for s in _buyer_ships if s.get("id")==_sel_sid), 0) or 0)
                        _already_kg    = sum(float(a.get("allocated_kg",0)) for a in _allocs if a.get("shipment_id")==_sel_sid)
                        _remain_kg     = _ship_total_kg - _already_kg
                        if _alloc_kg_input <= 0:
                            st.error("배분량은 0보다 커야 합니다.")
                        elif _alloc_kg_input * 1000 > _remain_kg + 0.1:
                            st.error(f"배분량이 잔여량을 초과합니다. 잔여: {_remain_kg/1000:,.3f} MT")
                        else:
                            _allocs.append({
                                "id":           str(uuid.uuid4())[:8],
                                "contract_id":  _ct_id,
                                "shipment_id":  _sel_sid,
                                "allocated_kg": round(_alloc_kg_input * 1000, 3),
                            })
                            cfg["contract_allocations"] = _allocs
                            save_cfg(cfg)
                            st.rerun()
                else:
                    st.caption("이 매입사의 선적 기록이 없습니다.")

                # 삭제
                with st.popover("🗑️ 계약 삭제"):
                    st.warning(f"**{_bname}** {_ct.get('product','')} {_m['qty_mt']:,.0f} MT 계약을 삭제합니다.")
                    if st.button("확인 삭제", key=f"ct_del_{_ct_id}"):
                        cfg["contracts"] = [c for c in _ct_list if c.get("id") != _ct_id]
                        cfg["contract_allocations"] = [
                            a for a in cfg.get("contract_allocations", []) if a.get("contract_id") != _ct_id
                        ]
                        save_cfg(cfg)
                        st.rerun()
